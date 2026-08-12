#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
consolidar_dw_da.py
===================

Consolida os dados do DW (SIASG) e do DA (Dados Abertos / Compras.gov) em uma
planilha por Classe, com duas abas (dw-XXXX e da-XXXX), removendo do DA todo
registro que já exista no DW.

Regra de duplicidade
--------------------
A chave é a identificação do item da compra, com 22 dígitos:

    DW  -> coluna "Identif Item Compra" (já vem com 22 dígitos)
    DA  -> "idCompra" (16 ou 17 dígitos -> completar com zero à esquerda até 17)
           + "numeroItemCompra" (1 a 5 dígitos -> completar até 5)

Se a chave do DA existir no DW, a linha do DA é descartada (o DW é a fonte
preferencial, por ter mais informação).

Saída
-----
Um arquivo por classe, com as abas dw-XXXX e da-XXXX, mais:
    Relatorio_Consolidacao.xlsx  - contagens por classe
    linhas_em_quarentena.csv     - linhas corrompidas na origem (colunas
                                   desalinhadas), preservadas na íntegra e
                                   fora das planilhas
    duplicatas_removidas.csv     - com --salvar-duplicatas, lista o que saiu do DA

Uso
---
    python consolidar_dw_da.py -e PASTA_COM_OS_ARQUIVOS -s PASTA_DE_SAIDA

    # várias pastas de entrada
    python consolidar_dw_da.py -e "C:\\DW\\6505" -e "C:\\DA" -s "C:\\Saida"

    # informando explicitamente o que é DW e o que é DA
    python consolidar_dw_da.py --dw "C:\\DW\\*.csv" --da "C:\\DA\\*.csv" -s "C:\\Saida"

    # gerando também o arquivo de auditoria com as duplicatas removidas
    python consolidar_dw_da.py -e PASTA -s SAIDA --salvar-duplicatas

Requisitos: Python 3.8+ e openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import argparse
import csv
import glob
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# =============================================================================
# CONFIGURAÇÃO  (ajuste aqui se precisar)
# =============================================================================

# Limite físico de linhas de uma aba do Excel (inclui o cabeçalho).
# Ao estourar, o script cria automaticamente "dw-6505 (2)", "dw-6505 (3)"...
LIMITE_LINHAS_PLANILHA = 1_048_576

# Formatos aplicados na planilha final (iguais aos do modelo enviado).
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_QTDE = "#,##0"
FORMATO_MOEDA = 'R$ #,##0.00'

# True  -> CATMAT gravado como texto, preservando zeros à esquerda ("000183").
# False -> CATMAT gravado como número (183). Atenção: perde os zeros.
CATMAT_COMO_TEXTO = True

# Data do DA usada para preencher as colunas "Ano" e "dataCompra".
# Alterne para "dataResultado" se preferir alinhar com o "Ano Resultado Compra"
# do DW.
CAMPO_DATA_DA = "dataCompra"

# Formatar a data em dd/mm/aaaa custa ~35% de tempo a mais na gravação.
# Com False, a data sai como aaaa-mm-dd (mais rápido, ainda é data de verdade).
FORMATAR_DATA_BR = True

# Modalidades do DA (códigos do SIASG). Complete se aparecerem códigos novos —
# o script avisa no fim quais códigos não estavam mapeados.
MAPA_MODALIDADE_DA = {
    "1": "Convite",
    "2": "Tomada de Preços",
    "3": "Concorrência",
    "4": "Concorrência Internacional",
    "5": "Pregão",
    "6": "Dispensa de Licitação",
    "7": "Inexigibilidade de Licitação",
    "20": "Concurso",
    "22": "Regime Diferenciado de Contratações",
    "99": "Não informado",
}

# Sigla -> nome da unidade de fornecimento (mesmo mapa do Extrator de CATMATs).
_SIGLA_NOME = {
    "FR-AM": "Frasco-Ampola", "FR": "Frasco", "CAPS": "Cápsula",
    "COMPR": "Comprimido", "AM": "Ampola", "UN": "Unidade",
    "SER": "Seringa", "BIS": "Bisnaga", "BLIS": "Blister",
    "BOL": "Bolsa", "BOM": "Bombona", "CA": "Cartucho",
    "CI": "Curie", "CJ": "Conjunto", "DOSE(S)": "Dose(s)",
    "DOSES": "Dose(s)", "DRAG": "Drágea", "EMB": "Embalagem",
    "EMP": "Emplastro", "ENV": "Envelope", "FLAC": "Flaconete",
    "G": "Grama", "GL": "Galão", "GLOB": "Glóbulo",
    "KG": "Quilograma", "L": "Litro", "MCG": "Micrograma",
    "MCU": "Milicurie", "MG": "Miligrama",
    "MIL CTE": "Milheiro de Cartelas", "ML": "Mililitro",
    "PAST": "Pastilha", "POTE": "Pote", "RO": "Rolo",
    "SAC": "Sachê", "SUP": "Supositório", "TAB": "Tablete",
    "TBO": "Tubo", "TBTE": "Tubete", "UI": "Unid. Internacional",
}

_MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# Em algumas exportações do DW as colunas de código vêm com o cabeçalho EM BRANCO,
# logo depois da coluna descritiva correspondente (ex.: "Descrição Material
# Servico" seguida de uma coluna sem nome que contém o CATMAT). Este mapa
# reconstrói esses nomes.
_DW_CODIGO_SEGUINTE = {
    "descricao material servico": "catmat",
    "classe material": "classe",
    "padrao desc material": "inc",
    "orgao sup unid partic": "uasgsup",
    "orgao unid partic": "uasguni",
    "orgao vinc uresp compra": "uasgvinc",
    "nome uresp compra": "uasg",
}

# Tamanhos aceitos para as chaves (fora disso a linha vai para a quarentena).
TAM_CHAVE_DW = 22
TAM_ID_COMPRA_DA = (16, 17)
TAM_ITEM_DA = (1, 5)

# -----------------------------------------------------------------------------
# Layout das abas de saída: (título, tipo, largura)
# -----------------------------------------------------------------------------
COLS_DW = [
    ("IdentififItemCompra", "texto", 24),
    ("Catmat", "texto", 12),
    ("Descrição", "texto", 60),
    ("U.F.", "texto", 18),
    ("Classe", "inteiro", 9),
    ("Órgão Sup Unid Partic", "texto", 32),
    ("Órgão Unid Partic", "texto", 32),
    ("Órgão Vinc UResp Compra", "texto", 32),
    ("Nome UResp Compra", "texto", 36),
    ("Municipio UResp Compra", "texto", 22),
    ("UF UResp Compra", "texto", 8),
    ("Esfera Unid Partic", "texto", 14),
    ("CPF/CNPJ Fornecedor", "texto", 18),
    ("Nome Fornecedor", "texto", 40),
    ("Fabric Material Compra", "texto", 24),
    ("Marca Material Compra", "texto", 24),
    ("Ano Resultado Compra", "inteiro", 10),
    ("Dia Resultado Compra", "data", 14),
    ("Modalidade Compra", "texto", 26),
    ("Qtde Comprada Item", "qtde", 12),
    ("Valor Preço Unit Item", "moeda", 14),
    ("Preço Total", "moeda", 14),
]

COLS_DA = [
    ("IdentififItemCompra", "texto", 24),
    ("Catmat", "texto", 12),
    ("Descrição", "texto", 60),
    ("U.F.", "texto", 18),
    ("Classe", "inteiro", 9),
    ("nomeUasg", "texto", 36),
    ("municipio", "texto", 22),
    ("estado", "texto", 8),
    ("nomeOrgao", "texto", 32),
    ("poder", "texto", 8),
    ("esfera", "texto", 8),
    ("niFornecedor", "texto", 18),
    ("nomeFornecedor", "texto", 40),
    ("marca", "texto", 24),
    ("Ano", "inteiro", 8),
    ("dataCompra", "data", 14),
    ("modalidade", "texto", 26),
    ("quantidade", "qtde", 12),
    ("precoUnitario", "moeda", 14),
    ("Preço Total", "moeda", 14),
]

_FORMATO_POR_TIPO = {"data": FORMATO_DATA, "qtde": FORMATO_QTDE, "moeda": FORMATO_MOEDA}


# =============================================================================
# UTILIDADES
# =============================================================================

def norm(s) -> str:
    """Normaliza nome de coluna: sem acento, minúsculo, espaços colapsados."""
    if s is None:
        return ""
    s = str(s).replace("\ufeff", "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def txt(v) -> str:
    """Converte qualquer valor de célula em texto limpo, sem '.0' de float."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:                      # NaN
            return ""
        if v.is_integer():
            return str(int(v))
        return repr(v)
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else s


def num(v):
    """Converte texto em número, aceitando '1.234,56', '1234.56', '165,00', '1 '."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(" ", "").replace("\xa0", "")
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    neg = s.startswith("-")
    s = s.lstrip("+-").replace("R$", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")       # 1.234,56
    elif "," in s:
        s = s.replace(",", ".")                        # 1234,56
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3 and len(s.split(".")[0]) <= 3:
        s = s.replace(".", "")                         # 1.234 -> milhar
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def inteiro(v):
    f = num(v)
    return int(f) if f is not None else None


def data_dw(v):
    """'19 Out 1999' -> date(1999,10,19). Aceita também 19/10/1999 e ISO."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = txt(v)
    if not s:
        return None
    partes = s.replace("/", " ").replace("-", " ").split()
    if len(partes) == 3:
        d, m, a = partes
        mes = _MESES_PT.get(norm(m)[:3])
        if mes is None:
            try:
                mes = int(m)
            except ValueError:
                mes = None
        try:
            if mes and len(d) == 4:                    # formato ISO: aaaa mm dd
                return date(int(d), mes, int(a))
            if mes:
                return date(int(a), mes, int(d))
        except ValueError:
            return None
    return None


def data_iso(v):
    """'2025-07-04' ou '2025-07-04T12:00:00' -> date."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = txt(v)
    if len(s) >= 10:
        try:
            return date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except ValueError:
            pass
    return None


def so_digitos(s: str) -> str:
    return "".join(c for c in s if c.isdigit())


# =============================================================================
# LEITURA DE ARQUIVOS (CSV e XLSX) EM STREAMING
# =============================================================================

try:
    csv.field_size_limit(sys.maxsize)
except OverflowError:
    csv.field_size_limit(2 ** 31 - 1)


def _detectar_encoding(caminho: Path) -> str:
    with open(caminho, "rb") as fb:
        amostra = fb.read(1_048_576)
    if amostra.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    for corte in (0, 1, 2, 3):                 # evita erro por caractere partido
        try:
            (amostra[: len(amostra) - corte]).decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_separador(linha: str) -> str:
    # O DW exporta com ";" e o Extrator de CATMATs usa "@" nos CSVs do DA.
    candidatos = {sep: linha.count(sep) for sep in ("@", ";", "\t", "|", ",")}
    sep = max(candidatos, key=candidatos.get)
    return sep if candidatos[sep] > 0 else ";"


def ler_tabela(caminho: Path):
    """Gerador: entrega o cabeçalho (lista) e, em seguida, cada linha (lista)."""
    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb["Dados CATMAT"] if "Dados CATMAT" in wb.sheetnames else wb[wb.sheetnames[0]]
        try:
            for linha in ws.iter_rows(values_only=True):
                yield list(linha)
        finally:
            wb.close()
    else:
        enc = _detectar_encoding(caminho)
        with open(caminho, "r", encoding=enc, newline="", errors="replace") as f:
            primeira = f.readline()
            sep = _detectar_separador(primeira)
            f.seek(0)
            for linha in csv.reader(f, delimiter=sep):
                yield linha


def indices(cabecalho) -> dict:
    """{nome_normalizado: posição}, reconstruindo cabeçalhos vazios do DW."""
    nomes = [norm(c) for c in cabecalho]
    for i in range(1, len(nomes)):
        if not nomes[i]:
            nomes[i] = _DW_CODIGO_SEGUINTE.get(nomes[i - 1], "")
    idx = {}
    for i, nome in enumerate(nomes):
        if nome and nome not in idx:
            idx[nome] = i
    return idx


def detectar_fonte(idx: dict) -> str | None:
    if "identif item compra" in idx:
        return "DW"
    if "idcompra" in idx and "numeroitemcompra" in idx:
        return "DA"
    return None


# =============================================================================
# TRANSFORMAÇÃO DAS LINHAS
# =============================================================================

def _leitor(linha, idx):
    def g(nome, padrao=""):
        i = idx.get(nome)
        if i is None or i >= len(linha):
            return padrao
        return txt(linha[i])
    return g


def _catmat(valor: str):
    if CATMAT_COMO_TEXTO or not valor:
        return valor
    return inteiro(valor) if so_digitos(valor) == valor else valor


def transformar_dw(linha, idx):
    """Devolve (chave22, classe, linha_de_saida, erro) a partir de uma linha do DW."""
    g = _leitor(linha, idx)

    chave = g("identif item compra")
    classe = g("classe")
    erro = ""
    if not (chave.isdigit() and len(chave) == TAM_CHAVE_DW):
        erro = f"'Identif Item Compra' deveria ter {TAM_CHAVE_DW} dígitos: {chave!r}"
    elif classe and not classe.isdigit():
        erro = f"coluna 'Classe' não numérica ({classe!r}) - provável desalinhamento"

    qtde = num(g("qtde comprada item"))
    preco = num(g("valor preco unit item"))
    total = round(qtde * preco, 2) if (qtde is not None and preco is not None) else None
    saida = [
        chave,
        _catmat(g("catmat")),
        g("descricao material servico"),
        g("unidade fornecimento"),
        inteiro(classe),
        g("orgao sup unid partic"),
        g("orgao unid partic"),
        g("orgao vinc uresp compra"),
        g("nome uresp compra"),
        g("municipio uresp compra"),
        g("uf uresp compra"),
        g("esfera unid partic"),
        g("cpf/cnpj fornecedor"),
        g("nome fornecedor"),
        g("fabric material compra"),
        g("marca material compra"),
        inteiro(g("ano resultado compra")),
        data_dw(g("dia resultado compra")),
        g("modalidade compra"),
        qtde,
        preco,
        total,
    ]
    return chave, classe, saida, erro


def _unidade_fornecimento_da(g):
    """Reproduz a coluna 'Unidade de Fornecimento' do Extrator de CATMATs."""
    pronta = g("unidade de fornecimento")
    if pronta:
        return pronta
    nome = g("nomeunidadefornecimento")
    sigla = g("siglaunidadefornecimento")
    cap = g("capacidadeunidadefornecimento")
    medida = g("siglaunidademedida")
    if not nome and sigla:
        nome = _SIGLA_NOME.get(sigla.upper(), sigla)
    cap_num = num(cap) or 0
    partes = [nome] if nome else []
    if cap and cap_num != 0:
        partes.append(cap)
        if medida:
            partes.append(medida)
    return " ".join(partes)


def transformar_da(linha, idx, modalidades_desconhecidas):
    """Devolve (chave22, classe, linha_de_saida, erro) a partir de uma linha do DA."""
    g = _leitor(linha, idx)

    id_compra = g("idcompra")
    n_item = g("numeroitemcompra")
    classe = g("codigoclasse")
    chave, erro = "", ""

    # Validação estrita: é ela que separa a linha boa da linha corrompida.
    # Linhas desalinhadas costumam trazer o idItemCompra no lugar do idCompra,
    # o que geraria uma chave errada (e uma duplicata não detectada).
    if not (id_compra.isdigit() and TAM_ID_COMPRA_DA[0] <= len(id_compra) <= TAM_ID_COMPRA_DA[1]):
        erro = (f"'idCompra' deveria ter {TAM_ID_COMPRA_DA[0]} ou "
                f"{TAM_ID_COMPRA_DA[1]} dígitos: {id_compra!r}")
    elif not (n_item.isdigit() and TAM_ITEM_DA[0] <= len(n_item) <= TAM_ITEM_DA[1]):
        erro = f"'numeroItemCompra' inválido: {n_item!r}"
    elif classe and not classe.isdigit():
        erro = f"'codigoClasse' não numérico ({classe!r}) - provável desalinhamento"
    else:
        chave = id_compra.zfill(17) + n_item.zfill(5)

    dt = data_iso(g(norm(CAMPO_DATA_DA))) or data_iso(g("datacompra")) or data_iso(g("dataresultado"))
    qtde = num(g("quantidade"))
    preco = num(g("precounitario"))
    total = round(qtde * preco, 2) if (qtde is not None and preco is not None) else None

    cod_mod = g("modalidade")
    modalidade = MAPA_MODALIDADE_DA.get(cod_mod, "")
    if not modalidade and cod_mod:
        if not erro:                       # linha corrompida não conta como código novo
            modalidades_desconhecidas.add(cod_mod)
        modalidade = f"Modalidade {cod_mod}"

    saida = [
        chave,
        _catmat(g("codigoitemcatalogo")),
        g("descricaoitem"),
        _unidade_fornecimento_da(g),
        inteiro(classe),
        g("nomeuasg"),
        g("municipio"),
        g("estado"),
        g("nomeorgao"),
        g("poder"),
        g("esfera"),
        g("nifornecedor"),
        g("nomefornecedor"),
        g("marca"),
        dt.year if dt else None,
        dt,
        modalidade,
        qtde,
        preco,
        total,
    ]
    return chave, classe, saida, erro


# =============================================================================
# ESCRITA DAS PLANILHAS
# =============================================================================

_FONTE_CAB = Font(bold=True, color="FFFFFF")
_FUNDO_CAB = PatternFill("solid", fgColor="1F4E79")
_ALINHA_CAB = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _letra(i: int) -> str:
    letra = ""
    while i >= 0:
        letra = chr(ord("A") + i % 26) + letra
        i = i // 26 - 1
    return letra


class SaidaClasse:
    """Um arquivo .xlsx por classe, com as abas dw-XXXX e da-XXXX."""

    def __init__(self, classe: str):
        self.classe = classe
        self.wb = Workbook(write_only=True)
        self.abas = {"dw": [], "da": []}
        self._nova_aba("dw")                      # garante a aba do DW mesmo vazia

    def _nova_aba(self, tipo: str):
        cols = COLS_DW if tipo == "dw" else COLS_DA
        n = len(self.abas[tipo]) + 1
        nome = f"{tipo}-{self.classe}" + (f" ({n})" if n > 1 else "")
        ws = self.wb.create_sheet(nome[:31])
        for i, (_titulo, tipo_col, largura) in enumerate(cols):
            dim = ws.column_dimensions[_letra(i)]
            dim.width = largura
            if tipo_col in _FORMATO_POR_TIPO and tipo_col != "data":
                dim.number_format = _FORMATO_POR_TIPO[tipo_col]
        cabecalho = []
        for titulo, _t, _l in cols:
            c = WriteOnlyCell(ws, value=titulo)
            c.font, c.fill, c.alignment = _FONTE_CAB, _FUNDO_CAB, _ALINHA_CAB
            cabecalho.append(c)
        # Em modo write_only, tudo o que vem antes de <sheetData> (painéis,
        # larguras, formatos de coluna) precisa ser definido ANTES do primeiro
        # append; o atributo simples ws.freeze_panes não é serializado.
        Worksheet.freeze_panes.fset(ws, "A2")
        ws.append(cabecalho)
        registro = [ws, 1]
        self.abas[tipo].append(registro)
        return registro

    def append(self, tipo: str, valores: list):
        registro = self.abas[tipo][-1] if self.abas[tipo] else self._nova_aba(tipo)
        if registro[1] >= LIMITE_LINHAS_PLANILHA:
            registro = self._nova_aba(tipo)
        ws = registro[0]
        if FORMATAR_DATA_BR:
            cols = COLS_DW if tipo == "dw" else COLS_DA
            valores = list(valores)
            for i, (_t, tipo_col, _l) in enumerate(cols):
                if tipo_col == "data" and valores[i] is not None:
                    c = WriteOnlyCell(ws, value=valores[i])
                    c.number_format = FORMATO_DATA
                    valores[i] = c
        ws.append(valores)
        registro[1] += 1

    def salvar(self, caminho: Path):
        for tipo, cols in (("dw", COLS_DW), ("da", COLS_DA)):
            if not self.abas[tipo]:
                self._nova_aba(tipo)
            ultima = _letra(len(cols) - 1)
            for ws, linhas in self.abas[tipo]:
                ws.auto_filter.ref = f"A1:{ultima}{max(linhas, 1)}"
        self.wb.save(caminho)


# =============================================================================
# PROCESSAMENTO
# =============================================================================

def coletar_arquivos(entradas, padroes=("*.csv", "*.CSV", "*.xlsx", "*.xlsm")) -> list:
    achados = []
    for entrada in entradas:
        p = Path(entrada)
        if p.is_dir():
            for padrao in padroes:
                achados.extend(sorted(p.rglob(padrao)))
        elif p.exists():
            achados.append(p)
        else:
            achados.extend(sorted(Path(x) for x in glob.glob(entrada)))
    vistos, unicos = set(), []
    for a in achados:
        chave = str(a.resolve()).lower()
        if chave not in vistos and a.suffix.lower() in (".csv", ".xlsx", ".xlsm"):
            vistos.add(chave)
            unicos.append(a)
    return unicos


def _est(estatisticas, classe):
    return estatisticas.setdefault(classe, {
        "dw": 0, "da_lidas": 0, "da_dup": 0, "da_mantidas": 0,
        "dw_invalidas": 0, "da_invalidas": 0, "da_dup_interna": 0,
    })


class Quarentena:
    """Guarda as linhas corrompidas em CSV, sem perder nada do conteúdo original."""

    def __init__(self, caminho: Path):
        self.caminho = caminho
        self._arq = None
        self._csv = None
        self.total = 0

    def registrar(self, fonte, arquivo, n_linha, motivo, linha):
        if self._arq is None:
            self._arq = open(self.caminho, "w", encoding="utf-8-sig", newline="")
            self._csv = csv.writer(self._arq, delimiter=";")
            self._csv.writerow(["Fonte", "Arquivo", "Linha", "Motivo", "Conteúdo original ->"])
        self._csv.writerow([fonte, arquivo, n_linha, motivo] + [txt(v) for v in linha])
        self.total += 1

    def fechar(self):
        if self._arq:
            self._arq.close()


def processar_dw(arquivos, saidas, estatisticas, chaves_dw, ano_min, ano_max,
                 quarentena, verbose=True):
    for arq in arquivos:
        gen = ler_tabela(arq)
        try:
            idx = indices(next(gen))
        except StopIteration:
            continue
        if "classe" not in idx:
            print(f"  ! {arq.name}: coluna 'Classe' não encontrada - as linhas irão "
                  f"para SEM_CLASSE")
        gravadas = ignoradas = invalidas = 0
        for n_linha, linha in enumerate(gen, start=2):
            if not linha or all(v in (None, "") for v in linha):
                continue
            chave, classe, saida, erro = transformar_dw(linha, idx)
            if erro:
                invalidas += 1
                _est(estatisticas, classe if classe.isdigit() else "SEM_CLASSE")["dw_invalidas"] += 1
                quarentena.registrar("DW", arq.name, n_linha, erro, linha)
                continue
            classe = classe or "SEM_CLASSE"
            chaves_dw.add(int(chave))              # int ocupa menos memória que str
            ano = saida[16]
            if (ano_min and ano and ano < ano_min) or (ano_max and ano and ano > ano_max):
                ignoradas += 1
                continue
            if classe not in saidas:
                saidas[classe] = SaidaClasse(classe)
            saidas[classe].append("dw", saida)
            _est(estatisticas, classe)["dw"] += 1
            gravadas += 1
        if verbose:
            extra = f" | {ignoradas} fora do filtro de ano" if ignoradas else ""
            extra += f" | {invalidas} em quarentena" if invalidas else ""
            print(f"  [DW] {arq.name}: {gravadas:,} linhas{extra}".replace(",", "."))


def processar_da(arquivos, saidas, estatisticas, chaves_dw, ano_min, ano_max,
                 dedup_interno, escritor_dup, modalidades_desconhecidas,
                 quarentena, verbose=True):
    vistas_da = set() if dedup_interno else None
    for arq in arquivos:
        gen = ler_tabela(arq)
        try:
            idx = indices(next(gen))
        except StopIteration:
            continue
        lidas = dup = mantidas = ignoradas = invalidas = 0
        for n_linha, linha in enumerate(gen, start=2):
            if not linha or all(v in (None, "") for v in linha):
                continue
            primeira = txt(linha[0]).lower()
            if primeira.startswith(("totalregistros", "totalpaginas", "idcompra")):
                continue                            # rodapé da API ou cabeçalho repetido
            lidas += 1
            chave, classe, saida, erro = transformar_da(linha, idx, modalidades_desconhecidas)
            if erro:
                invalidas += 1
                _est(estatisticas, classe if classe.isdigit() else "SEM_CLASSE")["da_invalidas"] += 1
                quarentena.registrar("DA", arq.name, n_linha, erro, linha)
                continue
            classe = classe or "SEM_CLASSE"
            e = _est(estatisticas, classe)
            e["da_lidas"] += 1

            if int(chave) in chaves_dw:
                dup += 1
                e["da_dup"] += 1
                if escritor_dup:
                    escritor_dup.writerow([classe, chave, arq.name, saida[1],
                                           saida[15] or "", saida[12], saida[18]])
                continue
            if vistas_da is not None:
                if int(chave) in vistas_da:
                    e["da_dup_interna"] += 1
                    continue
                vistas_da.add(int(chave))

            ano = saida[14]
            if (ano_min and ano and ano < ano_min) or (ano_max and ano and ano > ano_max):
                ignoradas += 1
                continue
            if classe not in saidas:
                saidas[classe] = SaidaClasse(classe)
            saidas[classe].append("da", saida)
            e["da_mantidas"] += 1
            mantidas += 1
        if verbose:
            extra = f" | {ignoradas} fora do filtro de ano" if ignoradas else ""
            extra += f" | {invalidas} em quarentena" if invalidas else ""
            print(f"  [DA] {arq.name}: {lidas:,} lidas | {dup:,} duplicadas removidas | "
                  f"{mantidas:,} mantidas{extra}".replace(",", "."))


def gravar_relatorio(caminho: Path, estatisticas: dict, arquivos_gerados: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    cabecalho = ["Classe", "Linhas DW", "DA lidas", "DA duplicadas (removidas)",
                 "DA mantidas", "% removido do DA", "DW em quarentena",
                 "DA em quarentena", "Duplicatas internas DA", "Arquivo gerado"]
    ws.append(cabecalho)
    for c in ws[1]:
        c.font, c.fill, c.alignment = _FONTE_CAB, _FUNDO_CAB, _ALINHA_CAB
    for classe in sorted(estatisticas):
        e = estatisticas[classe]
        pct = (e["da_dup"] / e["da_lidas"] * 100) if e["da_lidas"] else 0
        ws.append([classe, e["dw"], e["da_lidas"], e["da_dup"], e["da_mantidas"],
                   round(pct, 2), e["dw_invalidas"], e["da_invalidas"],
                   e["da_dup_interna"], arquivos_gerados.get(classe, "")])
    larguras = [14, 14, 14, 26, 14, 18, 18, 18, 22, 44]
    for i, w in enumerate(larguras):
        ws.column_dimensions[_letra(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_letra(len(cabecalho) - 1)}{ws.max_row}"
    wb.save(caminho)


# =============================================================================
# MAIN
# =============================================================================

def main():
    ap = argparse.ArgumentParser(
        description="Consolida DW (SIASG) + DA (Compras.gov) por classe, "
                    "removendo do DA os registros já existentes no DW.")
    ap.add_argument("-e", "--entrada", action="append", default=[],
                    help="pasta (ou arquivo/curinga) com os CSVs do DW e do DA. "
                         "Pode repetir. A origem é detectada pelo cabeçalho.")
    ap.add_argument("--dw", action="append", default=[], help="entradas que são do DW")
    ap.add_argument("--da", action="append", default=[], help="entradas que são do DA")
    ap.add_argument("-s", "--saida", required=True, help="pasta de saída")
    ap.add_argument("--prefixo", default="bps_dw_da__Classe_",
                    help="prefixo do nome dos arquivos gerados")
    ap.add_argument("--sufixo", default="", help="sufixo do nome (ex.: _2020_a_2026)")
    ap.add_argument("--ano-min", type=int, help="descarta registros anteriores a este ano")
    ap.add_argument("--ano-max", type=int, help="descarta registros posteriores a este ano")
    ap.add_argument("--salvar-duplicatas", action="store_true",
                    help="gera CSV de auditoria com as linhas do DA removidas")
    ap.add_argument("--dedup-interno-da", action="store_true",
                    help="também remove repetições da mesma chave dentro do próprio DA "
                         "(atenção: costumam ser registros distintos, de fornecedores "
                         "e preços diferentes)")
    args = ap.parse_args()

    dir_saida = Path(args.saida)
    dir_saida.mkdir(parents=True, exist_ok=True)

    arquivos_dw = coletar_arquivos(args.dw)
    arquivos_da = coletar_arquivos(args.da)
    indefinidos = [a for a in coletar_arquivos(args.entrada)
                   if a not in arquivos_dw and a not in arquivos_da]

    if not (arquivos_dw or arquivos_da or indefinidos):
        print("Nenhum arquivo .csv/.xlsx encontrado nas entradas informadas.")
        return 1

    print("Identificando a origem de cada arquivo...")
    ignorados = []
    for arq in indefinidos:
        gen = ler_tabela(arq)
        try:
            fonte = detectar_fonte(indices(next(gen)))
        except StopIteration:
            fonte = None
        gen.close()
        if fonte == "DW":
            arquivos_dw.append(arq)
        elif fonte == "DA":
            arquivos_da.append(arq)
        else:
            ignorados.append(arq)
    for arq in ignorados:
        print(f"  ! ignorado (cabeçalho não reconhecido): {arq.name}")
    print(f"  {len(arquivos_dw)} arquivo(s) do DW | {len(arquivos_da)} arquivo(s) do DA")

    saidas, estatisticas, chaves_dw = {}, {}, set()
    modalidades_desconhecidas = set()
    quarentena = Quarentena(dir_saida / "linhas_em_quarentena.csv")

    print("\n[1/3] Lendo o DW e montando o índice de chaves...")
    processar_dw(arquivos_dw, saidas, estatisticas, chaves_dw, args.ano_min,
                 args.ano_max, quarentena)
    print(f"  -> {len(chaves_dw):,} chaves únicas no DW".replace(",", "."))

    arq_dup = escritor_dup = None
    if args.salvar_duplicatas:
        arq_dup = open(dir_saida / "duplicatas_removidas.csv", "w",
                       encoding="utf-8-sig", newline="")
        escritor_dup = csv.writer(arq_dup, delimiter=";")
        escritor_dup.writerow(["Classe", "IdentififItemCompra", "Arquivo origem",
                               "Catmat", "dataCompra", "nomeFornecedor", "precoUnitario"])

    print("\n[2/3] Lendo o DA e removendo as duplicatas...")
    try:
        processar_da(arquivos_da, saidas, estatisticas, chaves_dw, args.ano_min,
                     args.ano_max, args.dedup_interno_da, escritor_dup,
                     modalidades_desconhecidas, quarentena)
    finally:
        if arq_dup:
            arq_dup.close()
        quarentena.fechar()

    print("\n[3/3] Gravando as planilhas...")
    gerados = {}
    for classe in sorted(saidas):
        nome = f"{args.prefixo}{classe}{args.sufixo}.xlsx"
        caminho = dir_saida / nome
        saidas[classe].salvar(caminho)
        gerados[classe] = nome
        e = _est(estatisticas, classe)
        print(f"  {nome}: dw={e['dw']:,} | da={e['da_mantidas']:,} "
              f"(removidas {e['da_dup']:,})".replace(",", "."))

    gravar_relatorio(dir_saida / "Relatorio_Consolidacao.xlsx", estatisticas, gerados)

    tot_dw = sum(e["dw"] for e in estatisticas.values())
    tot_lidas = sum(e["da_lidas"] for e in estatisticas.values())
    tot_dup = sum(e["da_dup"] for e in estatisticas.values())
    tot_mant = sum(e["da_mantidas"] for e in estatisticas.values())
    print("\n" + "=" * 62)
    print(f"DW gravado ..................... {tot_dw:,}".replace(",", "."))
    print(f"DA lido (linhas válidas) ....... {tot_lidas:,}".replace(",", "."))
    print(f"DA removido (já estava no DW) .. {tot_dup:,}".replace(",", "."))
    print(f"DA mantido ..................... {tot_mant:,}".replace(",", "."))
    print(f"Relatório ...................... Relatorio_Consolidacao.xlsx")
    if quarentena.total:
        print(f"\n! {quarentena.total:,} linha(s) corrompida(s) não entraram nas planilhas."
              .replace(",", "."))
        print(f"  Conteúdo preservado em: {quarentena.caminho.name}")
    if modalidades_desconhecidas:
        print(f"\n! Códigos de modalidade não mapeados: "
              f"{', '.join(sorted(modalidades_desconhecidas))}"
              f"\n  Complete o dicionário MAPA_MODALIDADE_DA no início do script.")
    print("=" * 62)
    return 0


if __name__ == "__main__":
    sys.exit(main())
