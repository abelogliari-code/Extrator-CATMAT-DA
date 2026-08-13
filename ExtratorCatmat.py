"""
Extrator de CATMATs Pro  —  v2.4
Motor gráfico: CustomTkinter  (tema claro/escuro nativo, cantos arredondados)
Identidade visual: inspirada no BPS / DESID (Gov.br)

Arquivo único e autossuficiente: a interface, o motor de extração e o motor de
consolidação DW + DA (aba 3) vivem todos aqui.

    python ExtratorCatmat.py                      -> abre a interface
    python ExtratorCatmat.py -e ENTRADA -s SAIDA  -> consolidação em linha de comando
"""

import re
import csv
import requests
import pandas as pd
from io import StringIO
from typing import Tuple, List, Optional
import os
import sys
import time
import glob
import argparse
import unicodedata
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
import shutil
import json
import threading
import math
from concurrent.futures import ThreadPoolExecutor, as_completed
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import customtkinter as ctk
import tkinter.ttk as ttk

# =============================================================================
# PALETA  —  neutros Gov.br + acento azul Gov + verde BPS + amarelo BPS
# =============================================================================
C_BG         = "#F4F5F7"   # cinza-papel (fundo geral)
C_SURFACE    = "#FFFFFF"   # branco (cards / frames)
C_BORDER     = "#DDE1E9"   # borda sutil
C_TEXT       = "#1A1D23"   # quase-preto
C_TEXT_MED   = "#555B6E"   # texto secundário
C_TEXT_LIGHT = "#8A92A6"   # placeholder / hint
C_ACCENT     = "#1351B4"   # azul Gov.br (primário)
C_ACCENT_H   = "#0C3784"   # hover do azul
C_GREEN      = "#168821"   # verde BPS (sucesso)
C_GREEN_H    = "#0E5C17"   # hover verde
C_YELLOW     = "#FFCD07"   # amarelo BPS (destaque / faixa)
C_ORANGE     = "#E37222"   # aviso
C_RED        = "#C0392B"   # erro / cancelar
C_LOG_BG     = "#13141A"   # terminal escuro
C_LOG_FG     = "#E8EAF0"   # texto terminal

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# =============================================================================
# LÓGICA DE NEGÓCIO
# =============================================================================

pausar_extracao     = threading.Event()
pausar_busca_catmat = threading.Event()
# Estado "set" = liberado. Inicializar aqui evita que um wait() bloqueie para
# sempre em qualquer fluxo que esqueça de chamar .set() antes de começar.
pausar_extracao.set()
pausar_busca_catmat.set()
cancelar_busca_catmat = False

requests.packages.urllib3.disable_warnings(
    requests.packages.urllib3.exceptions.InsecureRequestWarning
)

# Session compartilhada — reutiliza conexões TCP/TLS entre todas as requisições
# Evita o overhead de handshake (~200-400ms) a cada chamada
_http = requests.Session()
_http.verify = False
_http.headers.update({"Accept-Encoding": "gzip, deflate", "Connection": "keep-alive"})
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=4, pool_maxsize=12, max_retries=0
)
_http.mount("https://", _adapter)
_http.mount("http://",  _adapter)

URL_BASE = "https://dadosabertos.compras.gov.br"
TIMEOUT  = 120
_MAX_PAGINAS = 20000   # trava contra laço infinito se a API paginar sem fim

# =============================================================================
# TIPOS DE BUSCA  —  espelha o seletor "tipo" do endpoint de Pesquisa de Preço
#   /modulo-pesquisa-preco/1_consultarMaterial?tipo={tipo}&codigo={codigo}
# =============================================================================
TIPO_CATMAT = "codigoItemCatalogo"
TIPO_PDM    = "codigoPdm"
ROTULO_TIPO = {TIPO_CATMAT: "CATMAT", TIPO_PDM: "PDM"}
TIPO_POR_ROTULO = {v: k for k, v in ROTULO_TIPO.items()}

# Detectado na 1ª requisição e reaproveitado nas demais:
#   None  → ainda não sabemos
#   True  → API aceita a assinatura nova (tipo + codigo)
#   False → API ainda na assinatura antiga (codigoItemCatalogo)
_API_ACEITA_TIPO = None

ordem_final_colunas = [
    "idCompra","idItemCompra","forma","modalidade","criterioJulgamento",
    "numeroItemCompra","descricaoItem","codigoItemCatalogo","codigoPdm","nomeUnidadeFornecimento",
    "siglaUnidadeFornecimento","nomeUnidadeMedida","capacidadeUnidadeFornecimento","siglaUnidadeMedida",
    "Unidade de Fornecimento","capacidade","quantidade","precoUnitario","Preco Total","percentualMaiorDesconto",
    "niFornecedor","nomeFornecedor","marca","codigoUasg","nomeUasg",
    "codigoMunicipio","municipio","estado","codigoOrgao","nomeOrgao",
    "poder","esfera","dataCompra","dataHoraAtualizacaoCompra","dataHoraAtualizacaoItem",
    "dataResultado","dataHoraAtualizacaoUasg","codigoClasse","nomeClasse",
]


class ExcelChunkWriter:
    """
    Escreve .xlsx em modo STREAMING (openpyxl write_only).

    O modo padrão do openpyxl mantém a planilha inteira em memória e só serializa
    tudo no save(), o que concentra o custo no encerramento — exatamente onde o
    usuário fica esperando. Medido com 30 mil linhas x 39 colunas:

        modo padrão   → 5,5 s durante + 9,5 s NO FINAL + 333 MB de RAM
        write_only    → 9,3 s durante + 0,9 s NO FINAL +   0 MB de RAM

    O tempo "durante" é absorvido pelas esperas de rede (0,5 s por página); o
    tempo "no final" é espera pura. Em uma extração de 230 mil registros isso
    troca ~70 s de espera no encerramento por ~7 s, e ~2,5 GB de RAM por nada.

    Contrapartida: em write_only o workbook só pode ser salvo UMA vez, então não
    há como reescrever o arquivo periodicamente. A proteção contra queda no meio
    da execução passa a ser um espelho .parcial.csv, gravado em append (16 ms por
    página de 500 linhas) e apagado quando o .xlsx é fechado com sucesso.
    """

    def __init__(self, base_filename, sheet_name="Dados CATMAT",
                 max_rows_per_file=1_000_000, espelho=True):
        self.base_filename = base_filename
        self.sheet_name    = sheet_name
        self.max_rows      = max_rows_per_file
        self.part          = 1
        self.header: List[str] = []
        self.current_row_count = 0
        self.files_saved: List[str] = []
        self._finalizado   = False
        self._usar_espelho = espelho
        self._espelho_f    = None
        self._espelho_w    = None
        self._new_workbook()

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".xlsx": ext = ".xlsx"
        return f"{base}_part{self.part}{ext}"

    def _espelho_path(self):
        base, _ = os.path.splitext(self.base_filename)
        return f"{base}.parcial.csv"

    def _new_workbook(self):
        # write_only exige create_sheet(); wb.active não existe nesse modo
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(self.sheet_name)
        self.header_written = False; self.current_row_count = 0

    def _ensure_header(self, columns):
        if not self.header: self.header = list(columns)
        if not self.header_written:
            self.ws.append(self.header); self.header_written = True
            self._abrir_espelho()

    def _abrir_espelho(self):
        """Espelho .parcial.csv — rede de segurança enquanto o .xlsx não fecha."""
        if not self._usar_espelho or self._espelho_f is not None:
            return
        try:
            self._espelho_f = open(self._espelho_path(), "w",
                                   encoding="utf-8-sig", newline="")
            self._espelho_w = csv.writer(self._espelho_f, delimiter=";")
            self._espelho_w.writerow(self.header)
        except Exception:
            self._usar_espelho = False      # sem espelho é melhor que falhar
            self._espelho_f = self._espelho_w = None

    def _fechar_espelho(self, apagar):
        if self._espelho_f is None: return
        try:
            self._espelho_f.close()
        except Exception:
            pass
        if apagar:
            try:
                os.remove(self._espelho_path())
            except Exception:
                pass
        self._espelho_f = self._espelho_w = None

    def _rollover_if_needed(self):
        if self.current_row_count + 1 > self.max_rows:
            path = self._filepath(); self.wb.save(path); self.files_saved.append(path)
            self.part += 1; self._new_workbook()
            if self.header:
                self.ws.append(self.header); self.header_written = True

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        self._ensure_header(list(df.columns))
        faltantes = [c for c in self.header if c not in df.columns]
        if faltantes:
            df = df.copy()          # não mutar o DataFrame do chamador
            for col in faltantes: df[col] = pd.NA
        df = df[self.header]
        for linha in df.itertuples(index=False, name=None):
            self._rollover_if_needed()
            # openpyxl levanta IllegalCharacterError em caracteres de controle,
            # frequentes no texto livre vindo da API — sanitiza na gravação
            limpa = [None if pd.isna(v) else
                     (_CTRL_ILEGAIS.sub(" ", v) if isinstance(v, str) else v)
                     for v in linha]
            self.ws.append(limpa)
            if self._espelho_w is not None:
                self._espelho_w.writerow(["" if v is None else v for v in limpa])
            self.current_row_count += 1

    def flush(self, intervalo_min=30, fator=20):
        """
        Em write_only o .xlsx não pode ser reescrito no meio do caminho: o que se
        garante aqui é que o espelho .parcial.csv esteja em disco.
        """
        if self._espelho_f is None: return None
        try:
            self._espelho_f.flush()
        except Exception:
            return None
        return self._espelho_path()

    def _descartar_workbook(self):
        """
        Um workbook write_only coletado sem save() deixa os geradores internos do
        openpyxl abertos, e o lxml despeja 'Exception ignored ... LxmlSyntaxError'
        no stderr durante o garbage collector. close() encerra os streams.
        """
        try:
            self.ws.close()      # encerra os geradores de escrita da planilha
        except Exception:
            pass
        try:
            self.wb.close()
        except Exception:
            pass

    def finalize(self) -> List[str]:
        if self._finalizado:
            return self.files_saved
        self._finalizado = True
        ok = True
        if self.header_written and self.current_row_count > 0:
            path = self._filepath()
            try:
                self.wb.save(path)
                if path not in self.files_saved: self.files_saved.append(path)
            except Exception:
                ok = False          # mantém o espelho: é tudo o que restou
        else:
            self._descartar_workbook()   # nada a salvar: fecha sem ruído
        # Espelho só é descartado quando o .xlsx foi fechado com sucesso
        self._fechar_espelho(apagar=ok)
        return self.files_saved


class CSVChunkWriter:
    def __init__(self, base_filename, sep=";", encoding="utf-8-sig", max_rows_per_file=1_000_000):
        self.base_filename = base_filename; self.sep = sep
        self.encoding = encoding; self.max_rows = max_rows_per_file
        self.part = 1; self.current_row_count = 0
        self.files_saved: List[str] = []; self.header_written = False
        self.header: List[str] = []

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".csv": ext = ".csv"
        return f"{base}_part{self.part}{ext}"

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        # O conjunto de colunas varia entre páginas (processar_dataframe_final
        # descarta colunas 100% vazias). Sem reindexar pelo cabeçalho da 1ª
        # página, o append gravaria valores sob colunas erradas.
        if not self.header:
            self.header = list(df.columns)
        df = df.reindex(columns=self.header)
        if self.current_row_count + len(df) > self.max_rows:
            self.part += 1; self.current_row_count = 0; self.header_written = False
        path = self._filepath()
        df.to_csv(path, sep=self.sep, index=False,
                  mode="a" if self.header_written else "w",
                  header=not self.header_written, encoding=self.encoding)
        self.header_written = True; self.current_row_count += len(df)
        if path not in self.files_saved: self.files_saved.append(path)

    def flush(self, intervalo_min=30, fator=20):
        """CSV já é gravado em append a cada página — nada a fazer."""
        return None

    def finalize(self) -> List[str]:
        return self.files_saved


def converter_data_para_api(data_dd_mm_yyyy: str) -> Optional[str]:
    s = data_dd_mm_yyyy.strip()
    if not s: return None
    try:
        p = s.split("-")
        if len(p) != 3: return None
        dd, mm, yyyy = p
        if len(dd) == 2 and len(mm) == 2 and len(yyyy) == 4:
            int(dd); int(mm); int(yyyy)
            return f"{yyyy}-{mm}-{dd}"
    except (ValueError, AttributeError):
        pass
    return None


def validar_e_obter_datas(ini: str, fim: str):
    i_api = f_api = None
    if ini.strip():
        i_api = converter_data_para_api(ini)
        if i_api is None:
            return None, None, f"Data de Inicio invalida: '{ini}'\nUse DD-MM-AAAA (ex: 01-01-2024)"
    if fim.strip():
        f_api = converter_data_para_api(fim)
        if f_api is None:
            return None, None, f"Data Final invalida: '{fim}'\nUse DD-MM-AAAA (ex: 31-12-2024)"
    return i_api, f_api, None


# =============================================================================
# PARSER DE PÁGINA CSV
# -----------------------------------------------------------------------------
# A API devolve CSV com campos de texto livre (descricaoItem) que podem conter
# quebras de linha cruas, ';' e aspas desbalanceadas. Três armadilhas conhecidas:
#
#   1. str.splitlines() quebra em \x0b \x0c \x1c-\x1e \x85 \u2028 \u2029, que o
#      parser CSV (e o servidor) tratam como texto comum → fragmentos fantasma.
#   2. Contar ';' com str.split ignora aspas → campos citados com ';' interno
#      viram "excesso de colunas" e a linha é remontada torta.
#   3. Descartar linhas que contêm só '"' apaga o fechamento de um campo
#      multilinha → o parser engole as linhas seguintes e some com registros.
#
# A abordagem aqui é: csv.reader sobre o texto bruto (que já resolve quebras
# dentro de campos citados), remontagem no nível de CAMPO e conferência do
# número de registros contra o esperado da página.
# =============================================================================

# Quebras que str.splitlines() reconhece mas o CSV não
_QUEBRAS_FALSAS = re.compile(r"[\x0b\x0c\x1c\x1d\x1e\x85\u2028\u2029]")
# Caracteres de controle que o openpyxl recusa ao gravar .xlsx
_CTRL_ILEGAIS   = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
# Linhas de metadados (totalRegistros / total paginas / paginas restantes)
_RE_METADADO    = re.compile(
    r'^\s*"?\s*total\s*(de\s*)?(registros|p[áa]ginas?|p[áa]ginas?\s+restantes)\s*:',
    re.IGNORECASE)


def _limpar_campo(v):
    """Normaliza um campo: tira controles ilegais e colapsa quebras internas."""
    if not v:
        return ""
    v = v.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    v = _QUEBRAS_FALSAS.sub(" ", v)
    v = _CTRL_ILEGAIS.sub(" ", v)
    return re.sub(r"[ \t]{2,}", " ", v).strip()


def _ler_registros(texto, quoting):
    """csv.reader sobre o texto bruto — resolve quebras dentro de campos citados."""
    try:
        return list(csv.reader(StringIO(texto, newline=""), delimiter=";",
                               quotechar='"', quoting=quoting, strict=False))
    except csv.Error:
        return []


def _remontar(registros, ncols, idx_livre):
    """
    Remonta registros quebrados.
      len < ncols → fragmento: a quebra caiu DENTRO de um campo, então o último
                    campo do fragmento e o primeiro do seguinte são as duas
                    metades do mesmo campo (por isso a junção é por campo, e
                    não pela linha inteira).
      len > ncols → ';' extra em campo sem aspas: recolhe o excedente de volta
                    para o campo de texto livre em vez de descartar a linha.
    Retorna (linhas, reparos, descartes).
    """
    linhas = []; buf = None; reparos = 0; descartes = 0
    for reg in registros:
        remontado = False
        if buf is not None:
            cabeca = (buf[-1] + " " + (reg[0] if reg else "")).strip()
            reg = buf[:-1] + [cabeca] + reg[1:]
            buf = None; remontado = True

        n = len(reg)
        if n < ncols:
            buf = reg                       # ainda incompleto — segue acumulando
            continue
        if n == ncols:
            if remontado: reparos += 1
            linhas.append(reg)
        elif idx_livre is not None and idx_livre < ncols:
            excedente = n - ncols
            reg = (reg[:idx_livre]
                   + [";".join(reg[idx_livre:idx_livre + excedente + 1])]
                   + reg[idx_livre + excedente + 1:])
            reparos += 1
            linhas.append(reg)
        else:
            descartes += 1
    if buf is not None:
        descartes += 1                      # fragmento órfão no fim da página
    return linhas, reparos, descartes


def _montar_pagina(texto, quoting):
    """Extrai (header, linhas, reparos, descartes) de uma página com um dado quoting."""
    registros = [r for r in _ler_registros(texto, quoting) if any(c.strip() for c in r)]
    registros = [r for r in registros if not _RE_METADADO.match(r[0] if r else "")]
    if not registros:
        return None, [], 0, 0
    header = [c.strip() for c in registros[0]]
    ncols  = len(header)
    if ncols < 2:
        return None, [], 0, 0
    idx_livre = header.index("descricaoItem") if "descricaoItem" in header else None
    linhas, reparos, descartes = _remontar(registros[1:], ncols, idx_livre)
    return header, linhas, reparos, descartes


def parse_pagina_csv(csv_text, esperado_na_pagina=None):
    """
    Converte o CSV de uma página em DataFrame com diagnóstico confiável.

    Retorna (df, diag), diag = {
        "linhas", "esperado", "reparos", "descartes", "modo", "ok", "motivo"
    }
    'ok' é False sempre que a página não entregou exatamente os registros
    esperados — é essa conferência (e não uma heurística de ';') que garante
    que nenhuma página problemática passe batido.
    """
    diag = {"linhas": 0, "esperado": esperado_na_pagina, "reparos": 0,
            "descartes": 0, "modo": "aspas", "ok": True, "motivo": ""}
    if not csv_text:
        diag.update(ok=False, motivo="pagina vazia")
        return pd.DataFrame(), diag

    melhor = None
    for modo, quoting in (("aspas", csv.QUOTE_MINIMAL), ("literal", csv.QUOTE_NONE)):
        header, linhas, reparos, descartes = _montar_pagina(csv_text, quoting)
        if header is None:
            continue
        cand = {"header": header, "linhas": linhas, "reparos": reparos,
                "descartes": descartes, "modo": modo}
        # Bateu o esperado no modo com aspas: não precisa da segunda leitura
        if esperado_na_pagina is not None and len(linhas) == esperado_na_pagina:
            melhor = cand
            break
        # Senão, fica com o que recupera mais registros e descarta menos
        if melhor is None or (len(linhas), -descartes) > (len(melhor["linhas"]),
                                                          -melhor["descartes"]):
            melhor = cand

    if melhor is None:
        diag.update(ok=False, motivo="cabecalho nao identificado")
        return pd.DataFrame(), diag

    dados = [[_limpar_campo(c) for c in ln] for ln in melhor["linhas"]]
    df = pd.DataFrame(dados, columns=melhor["header"], dtype=str)

    diag.update(linhas=len(df), reparos=melhor["reparos"],
                descartes=melhor["descartes"], modo=melhor["modo"])
    if esperado_na_pagina is not None and len(df) != esperado_na_pagina:
        diag["ok"] = False
        diag["motivo"] = f"{len(df)} de {esperado_na_pagina} registros"
    elif melhor["descartes"]:
        diag["ok"] = False
        diag["motivo"] = f"{melhor['descartes']} linha(s) descartada(s)"
    return df, diag


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    """Compatibilidade: mantém a assinatura antiga sobre o novo motor."""
    df, _ = parse_pagina_csv(csv_text)
    return df


def ler_pagina_catmat(codigo, pagina, URL_BASE, TAMANHO_PAGINA, TIMEOUT,
                      data_compra_inicio=None, data_compra_fim=None,
                      tipo=TIPO_CATMAT):
    """
    Lê uma página de Registros de Preço.

    tipo — equivale ao seletor "tipo" do endpoint de Pesquisa de Preço:
        TIPO_CATMAT ("codigoItemCatalogo") → codigo é um CATMAT
        TIPO_PDM    ("codigoPdm")          → codigo é um PDM (traz todos os
                                             CATMATs do PDM de uma só vez)

    Envia a assinatura nova (tipo + codigo). Se o servidor recusar (400/404) e a
    busca for por CATMAT, refaz com a assinatura antiga (codigoItemCatalogo),
    mantendo compatibilidade com instâncias ainda não atualizadas da API.
    """
    global _API_ACEITA_TIPO
    URL = f"{URL_BASE}/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"

    base = {"tamanhoPagina": TAMANHO_PAGINA, "pagina": int(pagina)}
    if data_compra_inicio: base["dataCompraInicio"] = data_compra_inicio
    if data_compra_fim:    base["dataCompraFim"]    = data_compra_fim

    def _requisitar(params):
        """Retorna (csv_text, erro, status_http). csv_text=None quando falhou."""
        tentativas = 0
        while tentativas < 2:
            try:
                resp = _http.get(URL, params=params, timeout=TIMEOUT)
                if resp.status_code == 429:
                    time.sleep(15 if tentativas == 0 else 30); tentativas += 1; continue
                if resp.status_code in (400, 404):
                    return None, f"ERRO_REQUISICAO: HTTP {resp.status_code}", resp.status_code
                resp.raise_for_status()
                return resp.content.decode("utf-8-sig", errors="replace"), None, 200
            except requests.exceptions.ConnectionError as e:
                return None, f"ERRO_CONEXAO: {e}", None
            except requests.exceptions.RequestException as e:
                return None, f"ERRO_REQUISICAO: {e}", None
        return None, f"ERRO_REQUISICAO: 429 persistente para {tipo} {codigo}", 429

    # ── 1ª opção: assinatura nova (tipo + codigo) ────────────────────────────
    if _API_ACEITA_TIPO is not False:
        csv_text, erro, status = _requisitar(
            dict(base, tipo=tipo, codigo=str(int(codigo))))
        if csv_text is not None:
            _API_ACEITA_TIPO = True
            return None, csv_text
        # Só cai para o modo legado quando o servidor recusa a assinatura
        if status not in (400, 404):
            return None, erro
        _API_ACEITA_TIPO = False

    # ── 2ª opção: assinatura antiga — existe apenas para CATMAT ──────────────
    if tipo != TIPO_CATMAT:
        return None, ("ERRO_REQUISICAO: esta instância da API não aceita busca "
                      f"por {tipo}. Selecione CATMAT.")

    csv_text, erro, _ = _requisitar(dict(base, codigoItemCatalogo=int(codigo)))
    if csv_text is not None:
        return None, csv_text
    return None, erro


def _normalizar_campo(item: dict, *candidatos, default=""):
    """Retorna o primeiro campo encontrado no dict entre os candidatos."""
    for c in candidatos:
        if c in item and item[c] is not None:
            return item[c]
    return default


def buscar_pdms_por_classe(codigo_classe: int, URL_BASE: str, TIMEOUT: int,
                           max_tentativas: int = 3):
    """Busca todos os PDMs de uma classe com retry automático e backoff."""
    URL = f"{URL_BASE}/modulo-material/3_consultarPdmMaterial"
    TAMANHO_PAGINA = 500
    all_pdms = []; pagina_atual = 1; total_paginas = 1; total_registros_api = 0

    while pagina_atual <= total_paginas:
        tentativa = 0
        sucesso   = False
        data      = None
        while tentativa < max_tentativas and not sucesso:
            try:
                resp = _http.get(URL, params={
                    "codigoClasse": codigo_classe, "pagina": pagina_atual,
                    "tamanhoPagina": TAMANHO_PAGINA, "bps": "false"
                }, timeout=TIMEOUT)
                # Rate-limit: espera antes de tentar de novo
                if resp.status_code == 429:
                    espera = 15 * (tentativa + 1)
                    print(f"Rate-limit classe {codigo_classe} pág {pagina_atual} "
                          f"— aguardando {espera}s (tentativa {tentativa+1})")
                    time.sleep(espera)
                    tentativa += 1
                    continue
                resp.raise_for_status()
                data    = resp.json()
                sucesso = True
            except requests.exceptions.ConnectionError as e:
                espera = 3 * (tentativa + 1)
                print(f"Erro conexão classe {codigo_classe} pág {pagina_atual}: {e} "
                      f"— aguardando {espera}s (tentativa {tentativa+1})")
                time.sleep(espera)
                tentativa += 1
            except Exception as e:
                espera = 2 * (tentativa + 1)
                print(f"Erro classe {codigo_classe} pág {pagina_atual}: {e} "
                      f"— aguardando {espera}s (tentativa {tentativa+1})")
                time.sleep(espera)
                tentativa += 1

        if not sucesso or data is None:
            print(f"Falha definitiva: classe {codigo_classe} pág {pagina_atual} "
                  f"após {max_tentativas} tentativas")
            return None

        if "resultado" in data:
            all_pdms.extend(data["resultado"])
        if pagina_atual == 1:
            total_registros_api = int(data.get("totalRegistros", 0))
            total_paginas = (math.ceil(total_registros_api / TAMANHO_PAGINA)
                             if total_registros_api > 0 else 1)
            print(f"Classe {codigo_classe}: {total_registros_api} PDMs / "
                  f"{total_paginas} página(s)")
        pagina_atual += 1
        time.sleep(0.2)

    if not all_pdms: return None

    # Normalizar campos — a API pode retornar nomes variados
    rows_norm = []
    for item in all_pdms:
        cod  = _normalizar_campo(item, "codigoPdm", "codigo", "id", "codigoItem")
        desc = _normalizar_campo(item, "nomePdm", "nome", "descricao", "descricaoPdm", "descricaoItem")
        # status pode ser bool True/False, string "ATIVO"/"INATIVO", ou inteiro
        raw_status = _normalizar_campo(item, "statusPdm", "status", "ativo", "situacao")
        if isinstance(raw_status, bool):
            status = "Ativo" if raw_status else "Inativo"
        elif isinstance(raw_status, str):
            status = "Ativo" if raw_status.upper() in ("ATIVO", "TRUE", "S", "SIM", "1") else "Inativo"
        elif isinstance(raw_status, (int, float)):
            status = "Ativo" if raw_status == 1 else "Inativo"
        else:
            status = "Ativo"
        rows_norm.append({"codigoPdm": cod, "nomePdm": desc, "statusPdm": status,
                          "_classe": str(codigo_classe)})

    df = pd.DataFrame(rows_norm).drop_duplicates(subset=["codigoPdm"])
    return df, total_registros_api


def buscar_catmats_por_pdm(codigos_pdm, URL_BASE, TIMEOUT, app,
                           log_fn=None, max_workers=5):
    """
    Busca CATMATs de múltiplos PDMs em paralelo usando ThreadPoolExecutor.
    max_workers=5 → ~5x mais rápido sem sobrecarregar a API.
    """
    global cancelar_busca_catmat
    URL   = f"{URL_BASE}/modulo-material/4_consultarItemMaterial"
    total = len(codigos_pdm)
    all_catmats  = []
    pdms_com_erro = []
    completed     = 0
    lock          = threading.Lock()   # protege all_catmats e pdms_com_erro

    def _fetch_pdm(idx_pdm):
        """Worker: busca todas as páginas de CATMATs de um PDM."""
        i, pdm_code = idx_pdm
        if cancelar_busca_catmat:
            return
        resultados = []; pagina_atual = 1; total_paginas = 1
        try:
            while pagina_atual <= total_paginas:
                if cancelar_busca_catmat: break
                resp = _http.get(URL, params={
                    "codigoPdm": pdm_code, "pagina": pagina_atual,
                    "tamanhoPagina": 500, "bps": "false"
                }, timeout=TIMEOUT)
                if resp.status_code == 429:
                    time.sleep(15); continue          # rate-limit: aguarda e repete
                resp.raise_for_status()
                data = resp.json()
                if "resultado" in data:
                    resultados.extend(data["resultado"])
                if pagina_atual == 1:
                    total_paginas = data.get("totalPaginas", 1)
                pagina_atual += 1
                if pagina_atual <= total_paginas:
                    time.sleep(0.2)                   # throttle entre páginas do mesmo PDM
        except Exception:
            with lock:
                pdms_com_erro.append(pdm_code)
            return
        if resultados:
            with lock:
                all_catmats.extend(resultados)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_fetch_pdm, (i, pdm)): (i, pdm)
            for i, pdm in enumerate(codigos_pdm)
        }
        for future in as_completed(futures):
            pausar_busca_catmat.wait()        # respeita pausa
            if cancelar_busca_catmat:
                app.after(0, lambda: app.set_status_explorador("Busca cancelada."))
                executor.shutdown(wait=False, cancel_futures=True)
                break
            completed += 1
            i, pdm_code = futures[future]
            msg = f"PDM {pdm_code} ({completed}/{total})..."
            app.after(0, lambda m=msg: app.set_status_explorador(m))
            if log_fn and (completed % 5 == 0 or completed == total):
                app.after(0, lambda m=msg: log_fn(m))

    return pd.DataFrame(all_catmats) if all_catmats else None, pdms_com_erro



def _int_do_rodape(csv_text, rotulo):
    """
    Lê um inteiro do rodapé da resposta (ex.: 'totalPaginas: 1.234').
    Tolera separador de milhar — '\\d+' sozinho capturaria apenas o '1'.
    Retorna None quando o rótulo não aparece na resposta.
    """
    m = re.search(rotulo + r"\s*:\s*([\d.,]+)", csv_text, re.IGNORECASE)
    if not m:
        return None
    digitos = re.sub(r"\D", "", m.group(1))
    return int(digitos) if digitos else None


def _fetch_catmat_registros(codigo, d_ini, d_fim, salvar_corr, pasta_corr,
                            _pausar_conexao_fn=None, tipo=TIPO_CATMAT,
                            _cancelado_fn=None, TAMANHO_PAGINA=500):
    if _pausar_conexao_fn is None:
        _pausar_conexao_fn = lambda: None  # no-op se não fornecida
    if _cancelado_fn is None:
        _cancelado_fn = lambda: False
    """
    Worker puro: busca e processa todas as páginas de um CATMAT ou de um PDM,
    conforme `tipo` (TIPO_CATMAT | TIPO_PDM).
    Pode rodar em qualquer thread — não acessa estado compartilhado.
    Retorna: (codigo, dfs_e_meta, tipo, reg_esp, pag_corr)
      tipo: "ok" | "vazio" | "erro" | "conexao"
      dfs_e_meta: lista de (df_processado, marca, num_pagina)
                  marca: "" | "reparada" | "perda"
    """
    dfs_e_meta  = []
    pag_corr    = {}
    reg_esp     = 0
    pagina_atual = 1; total_paginas = None

    try:
        while True:
            _, csv_text = ler_pagina_catmat(codigo, 1, URL_BASE, TAMANHO_PAGINA, TIMEOUT,
                                            d_ini, d_fim, tipo=tipo)
            if csv_text and csv_text.startswith("ERRO_CONEXAO"):
                # Pausa automática + contagem regressiva de 60s antes de retentar
                _pausar_conexao_fn()
                for seg_restante in range(60, 0, -1):
                    # Se o usuário clicar Retomar manualmente, interrompe a contagem
                    if pausar_extracao.is_set():
                        break
                    time.sleep(1)
                # Retoma automaticamente ao fim da contagem (ou imediatamente se
                # o usuário já clicou Retomar)
                pausar_extracao.set()
                continue
            break
        if csv_text is None or csv_text.startswith("ERRO_REQUISICAO"):
            return codigo, [], "erro", 0, {}

        reg_esp = _int_do_rodape(csv_text, "totalRegistros") or 0
        if reg_esp == 0:
            return codigo, [], "vazio", 0, {}

        # Total de páginas: o rodapé da resposta é a fonte primária. O cálculo
        # por totalRegistros entra como conferência e, quando o rodapé falta,
        # como substituto — o código antigo caía em 1 nesse caso e truncava a
        # extração nos 500 primeiros registros em silêncio.
        pag_rodape    = _int_do_rodape(csv_text, r"total\s*(?:de\s*)?p[áa]ginas?")
        pag_calculado = max(1, math.ceil(reg_esp / TAMANHO_PAGINA))
        total_paginas = max(pag_rodape or 0, pag_calculado)
        # Sem rodapé não há como conferir a paginação: só nesse caso vale
        # insistir enquanto as páginas voltarem cheias.
        confiar_no_rodape = pag_rodape is not None

        while True:
            # Respeita pausa/cancelamento também ENTRE PÁGINAS de um mesmo código
            pausar_extracao.wait()
            if _cancelado_fn():
                break

            esperado_pag = max(0, min(TAMANHO_PAGINA,
                                      reg_esp - TAMANHO_PAGINA * (pagina_atual - 1)))
            df_pag, diag = parse_pagina_csv(csv_text, esperado_pag or None)
            # "perda"    → não foi possível reconstituir a página fielmente
            # "reparada" → houve conserto, mas todos os registros foram recuperados
            if not diag["ok"]:
                marca = "perda"
            elif diag["reparos"]:
                marca = "reparada"
            else:
                marca = ""

            if marca == "perda":
                pag_corr.setdefault(codigo, []).append(str(pagina_atual))
                if salvar_corr and pasta_corr:
                    dest = os.path.join(pasta_corr,
                        f"cod_{codigo}_pag_{pagina_atual}_corr.csv")
                    try:
                        with open(dest, "w", encoding="utf-8-sig") as f:
                            f.write(csv_text)
                    except Exception:
                        pass

            if df_pag is not None and not df_pag.empty:
                if tipo == TIPO_PDM:
                    # Busca por PDM devolve vários CATMATs: preserva o
                    # codigoItemCatalogo original e apenas anota o PDM de origem
                    df_pag.loc[:, "codigoPdm"] = str(codigo)
                else:
                    df_pag.loc[:, "codigoItemCatalogo"] = str(codigo)
                df_proc = processar_dataframe_final(df_pag, ordem_final_colunas)
                dfs_e_meta.append((df_proc, marca, pagina_atual))

            if total_paginas is None:
                total_paginas = 1

            # Com rodapé, ele manda: nenhuma requisição além do que ele informa.
            # Sem rodapé, página cheia é o único indício de que ainda há dados.
            insistir = (not confiar_no_rodape
                        and df_pag is not None and len(df_pag) >= TAMANHO_PAGINA)
            pagina_atual += 1
            if pagina_atual > total_paginas and not insistir:
                break
            if pagina_atual > _MAX_PAGINAS:      # trava de segurança
                break
            time.sleep(0.5)
            _, csv_text = ler_pagina_catmat(codigo, pagina_atual, URL_BASE,
                                            TAMANHO_PAGINA, TIMEOUT,
                                            d_ini, d_fim, tipo=tipo)
            if csv_text is None or csv_text.startswith("ERRO_"):
                break

        return codigo, dfs_e_meta, ("ok" if dfs_e_meta else "vazio"), reg_esp, pag_corr

    except Exception:
        return codigo, [], "erro", 0, {}


def processar_dataframe_final(df: pd.DataFrame, ordem_colunas: List[str]) -> pd.DataFrame:
    if df.empty: return df
    fc = df.columns[0]
    df = df[~df[fc].astype(str).str.contains("totalRegistros|totalPaginas",
                                              case=False, na=False)].copy()
    if df.empty: return df

    # Mapa sigla → nome completo para preencher nomeUnidadeFornecimento ausente
    _SIGLA_NOME = {
        "FR-AM": "Frasco-Ampola", "FR": "Frasco", "CAPS": "Cápsula",
        "COMPR": "Comprimido", "AM": "Ampola", "UN": "Unidade",
        "SER": "Seringa", "BIS": "Bisnaga", "BLIS": "Blister",
        "BOL": "Bolsa", "BOM": "Bombona", "CA": "Cartucho",
        "CI": "Curie", "CJ": "Conjunto", "DOSE(S)": "Dose(s)",
        "DOSES": "Dose(s)", "DRAG": "Drágea", "EMB": "Embalagem",
        "EMP": "Emplastro", "ENV": "Envelope", "FLAC": "Flaconete",
        "G": "Grama", "GL": "Galão", "GLOB": "Glóbulo",
        "KG": "Quilograma", "L": "Litro", "MCG": "Micrograma",
        "MCU": "Milicurie", "MG": "Miligrama",
        "MIL CTE": "Milheiro de Cartelas", "ML": "Mililitro",
        "PAST": "Pastilha", "POTE": "Pote", "RO": "Rolo",
        "SAC": "Sachê", "SUP": "Supositório", "TAB": "Tablete",
        "TBO": "Tubo", "TBTE": "Tubete", "UI": "Unid. Internacional",
    }

    def _val(row, col):
        v = row.get(col)
        s = str(v).strip() if pd.notna(v) else ""
        return "" if s in ("", "nan", "None", "null") else s

    def uf(row):
        nome  = _val(row, "nomeUnidadeFornecimento")
        sigla = _val(row, "siglaUnidadeFornecimento")
        cap   = _val(row, "capacidadeUnidadeFornecimento")
        medi  = _val(row, "siglaUnidadeMedida")

        # Se nomeUnidade vazio, tentar preencher pela sigla
        if not nome and sigla:
            nome = _SIGLA_NOME.get(sigla.upper(), sigla)

        # capacidade: ignorar se for 0,00 ou 0.00 ou 0
        try:
            cap_num = float(cap.replace(".", "").replace(",", ".")) if cap else 0
        except Exception:
            cap_num = 0
        cap_valida = bool(cap) and cap_num != 0

        # Montar: Nome + capacidade + siglaUnidadeMedida
        # Se não houver capacidade válida, ignorar também siglaUnidadeMedida
        partes = [nome] if nome else []
        if cap_valida:
            partes.append(cap)
            if medi:
                partes.append(medi)
        return " ".join(partes)

    df["Unidade de Fornecimento"] = df.apply(uf, axis=1)

    def tof(v):
        if pd.isna(v): return 0.0
        try: return float(str(v).replace(".", "").replace(",", "."))
        except: return 0.0

    # Sem os .get() a ausência de uma coluna vira KeyError, que o worker
    # captura no except genérico e reporta como "erro de API" — uma mudança de
    # schema na origem apareceria como indisponibilidade, sem pista no log.
    if "precoUnitario" in df.columns and "quantidade" in df.columns:
        df["Preco Total"] = df["precoUnitario"].apply(tof) * df["quantidade"].apply(tof)
    else:
        df["Preco Total"] = 0.0
    for col in ["nomeUnidadeMedida","percentualMaiorDesconto"]:
        if col in df.columns and (df[col].isnull().all() or
                                   df[col].astype(str).str.strip().eq("").all()):
            df = df.drop(columns=[col])
    exist = [c for c in ordem_colunas if c in df.columns]
    extra = [c for c in df.columns if c not in exist]
    return df[exist + extra]


# =============================================================================
# COMPONENTES DE UI  (helpers)
# =============================================================================

WELCOME = """\
Olá! Bem-vindo ao Extrator de CATMATs Pro.

Sua ferramenta para extrair e descobrir dados no Portal de Compras Governamentais!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
O que este programa faz?
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Este programa possui tres funcoes principais em abas separadas:

  1. Extracao por CATMAT (esta aba)
     Se voce ja tem uma lista de codigos de materiais (CATMATs), esta aba
     busca todas as informacoes de compras, corrige problemas nos dados e
     consolida tudo em um arquivo Excel ou CSV.

  2. Extracao por Classes (aba ao lado)
     Se voce quer descobrir novos itens, pode comecar com o codigo de uma
     ou mais Classes, encontrar todos os Padroes Descritivos de Materiais
     (PDMs) dentro delas e, em seguida, listar todos os CATMATs relacionados
     para extracao.

  3. Consolidacao DW + DA (ultima aba)
     Junta o historico do DW (SIASG) com o do DA (Dados Abertos) em uma
     planilha por Classe, removendo do DA todo registro que ja exista no
     DW. A chave e o identificador de 22 digitos do item da compra.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Primeiros Passos
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  - Para uma extracao direta com uma lista pronta, use esta aba.
    Escolha em "Buscar por" se os codigos da planilha sao CATMATs ou PDMs
    (equivale ao parametro "tipo" da API de Pesquisa de Preco):
      CATMAT -> coluna codigoItemCatalogo
      PDM    -> coluna codigoPdm  (traz todos os itens do PDM de uma vez)
    A coluna generica "codigo" tambem e aceita nos dois modos.

  - Para descobrir itens, use a aba "Extracao por Classes" e, ao final,
    envie os CATMATs encontrados para a extracao nesta aba.
    Nessa aba, marcando "Extrair precos direto por PDM" o programa pula a
    expansao PDM -> CATMAT e consulta a Pesquisa de Preco com tipo=codigoPdm,
    o que reduz drasticamente o numero de requisicoes.

  - Utilize os filtros de data (DD-MM-AAAA) para restringir os resultados
    a um periodo especifico de compras (Data de Inicio e Data Final).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Acompanhe todo o processo em tempo real neste log. Bom trabalho!
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""


BEMVINDO_CONSOLIDACAO = """\
Consolidacao DW + DA  —  remocao de duplicatas

Junta os dados do DW (SIASG) e do DA (Dados Abertos / Compras.gov) em uma
planilha por Classe, com duas abas (dw-XXXX e da-XXXX), descartando do DA
todo registro que ja exista no DW.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Regra de duplicidade
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

A chave e a identificacao do item da compra, com 22 digitos:

  DW  -> coluna "Identif Item Compra" (ja vem com 22 digitos)
  DA  -> idCompra (completado com zeros a esquerda ate 17)
         + numeroItemCompra (completado ate 5)

Se a chave do DA existir no DW, a linha do DA sai: o DW e a fonte
preferencial, por trazer mais informacao.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Como usar
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  1. Adicione as pastas (ou arquivos) com os CSVs/XLSX do DW e do DA.
     Com "Detectar", a origem de cada arquivo e descoberta pelo cabecalho;
     use DW ou DA para forcar a classificacao.
  2. Escolha a pasta de saida.
  3. Clique em "Consolidar e Remover Duplicatas".

Alem das planilhas por classe, sao gerados o Relatorio_Consolidacao.xlsx
(contagens por classe) e, quando houver, o linhas_em_quarentena.csv com as
linhas corrompidas na origem — preservadas na integra e fora das planilhas.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""


def _lbl(parent, text, size=12, weight="normal", color=C_TEXT, **kw):
    return ctk.CTkLabel(parent, text=text, font=("Segoe UI", size, weight),
                        text_color=color, **kw)


def _btn(parent, text, command, variant="secondary", width=0, **kw):
    pal = {
        "primary":   (C_SURFACE,  C_ACCENT,  C_SURFACE,  C_ACCENT_H),
        "success":   (C_SURFACE,  C_GREEN,   C_SURFACE,  C_GREEN_H),
        "danger":    (C_SURFACE,  C_RED,     C_SURFACE,  "#992B1E"),
        "secondary": (C_TEXT,     "#E4E7EF", C_TEXT,     C_BORDER),
        "ghost":     (C_ACCENT,   "transparent", C_ACCENT_H, "#E8EDF8"),
    }
    tc, bg, htc, hbg = pal.get(variant, pal["secondary"])
    return ctk.CTkButton(parent, text=text, command=command,
                         font=("Segoe UI", 12), fg_color=bg, text_color=tc,
                         hover_color=hbg, corner_radius=6,
                         width=width, height=32, **kw)


def _entry(parent, textvariable=None, placeholder="", width=200, **kw):
    return ctk.CTkEntry(parent, textvariable=textvariable,
                        placeholder_text=placeholder,
                        font=("Segoe UI", 12),
                        fg_color=C_SURFACE, text_color=C_TEXT,
                        border_color=C_BORDER, border_width=1,
                        corner_radius=6, width=width,
                        placeholder_text_color=C_TEXT_LIGHT, **kw)


def _sep(parent, pady=(6,6)):
    ctk.CTkFrame(parent, height=1, fg_color=C_BORDER,
                 corner_radius=0).pack(fill="x", padx=14, pady=pady)


def _card(parent, title="", **kw):
    outer = ctk.CTkFrame(parent, fg_color=C_SURFACE, corner_radius=8,
                         border_width=1, border_color=C_BORDER, **kw)
    if title:
        _lbl(outer, title, size=11, weight="bold", color=C_TEXT_MED)\
            .pack(anchor="w", padx=14, pady=(10,4))
        _sep(outer, pady=(0,6))
    return outer


# =============================================================================
# APLICATIVO PRINCIPAL
# =============================================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Extrator de CATMATs Pro  |  BPS / DESID")
        self.withdraw()                          # esconde até centralizar
        self.geometry("1100x860")
        self.minsize(960, 720)
        self.configure(fg_color=C_BG)

        # estado
        self.processing           = False
        self.codes_iterator       = None
        self.writer               = None
        self.codigos_lista: List  = []
        self.paginas_corrompidas  = {}
        self.registros_esperados  = {}
        self.registros_baixados   = {}
        self.total_baixados       = 0
        self.count_corrigidas     = 0
        self.count_reparadas      = 0
        self._modo_por_classe     = False
        self.count_vazios         = 0
        self._data_inicio         = None
        self._data_fim            = None
        self._tipo_busca          = TIPO_CATMAT
        self.lista_pdms_df        = pd.DataFrame()
        self.lista_catmats: List  = []

        # Construir interface ANTES de centralizar
        self._build_header()
        self._build_tabs()

        # Centralizar após tudo construído — delay generoso para o Tkinter
        # calcular dimensões reais antes de exibir
        self.after(200, self._centralizar)

    def _centralizar(self):
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        # winfo_width retorna 1 até a janela aparecer; usar reqwidth como fallback
        ww = self.winfo_reqwidth()  or 1100
        wh = self.winfo_reqheight() or 800
        # Respeitar o geometry definido (1100x800)
        ww = max(ww, 1100)
        wh = max(wh, 860)
        x = max(0, (sw - ww) // 2)
        y = max(0, (sh - wh) // 2)
        self.geometry(f"{ww}x{wh}+{x}+{y}")
        self.deiconify()

    # ── HEADER ────────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=C_ACCENT, corner_radius=0, height=50)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        _lbl(hdr, "  Extrator de CATMATs Pro", size=15, weight="bold",
             color=C_SURFACE).pack(side="left", padx=6)
        _lbl(hdr, "BPS · DESID · Ministério da Saúde  ",
             size=10, color="#A8BFDF").pack(side="right")
        # faixa amarela
        ctk.CTkFrame(self, height=3, fg_color=C_YELLOW,
                     corner_radius=0).pack(fill="x")

    # ── TABS ──────────────────────────────────────────────────────────────────
    def _build_tabs(self):
        self.tabs = ctk.CTkTabview(
            self, fg_color=C_BG,
            segmented_button_fg_color=C_BORDER,
            segmented_button_selected_color=C_ACCENT,
            segmented_button_selected_hover_color=C_ACCENT_H,
            segmented_button_unselected_color=C_BORDER,
            segmented_button_unselected_hover_color="#C5CAD5",
            text_color=C_TEXT, text_color_disabled=C_TEXT_LIGHT,
            corner_radius=0)
        self.tabs.pack(fill="both", expand=True)
        self.tabs.add("  Extração por CATMAT  ")
        self.tabs.add("  Extração por Classes  ")
        self.tabs.add("  Consolidação DW + DA  ")
        self._build_tab_extracao(self.tabs.tab("  Extração por CATMAT  "))
        self._build_tab_explorador(self.tabs.tab("  Extração por Classes  "))
        self._build_tab_consolidacao(self.tabs.tab("  Consolidação DW + DA  "))

    # ── ABA 1 ─────────────────────────────────────────────────────────────────
    def _build_tab_extracao(self, parent):
        parent.configure(fg_color=C_BG)
        # Frame normal sem scroll — tudo deve caber na tela
        wrap = ctk.CTkFrame(parent, fg_color=C_BG, corner_radius=0)
        wrap.pack(fill="both", expand=True, padx=12, pady=8)

        # — Card 1: entrada —
        c1 = _card(wrap, "1.  Dados para a Extração")
        c1.pack(fill="x", pady=(0,8))
        inn = ctk.CTkFrame(c1, fg_color="transparent")
        inn.pack(fill="x", padx=14, pady=(0,12))

        # arquivo
        r = ctk.CTkFrame(inn, fg_color="transparent"); r.pack(fill="x", pady=3)
        _lbl(r, "Arquivo de Códigos:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_arquivo = tk.StringVar()
        _entry(r, textvariable=self.var_arquivo,
               placeholder="Selecione .xlsx ou .csv", width=420)\
            .pack(side="left", expand=True, fill="x")
        _btn(r, "Procurar…", self._escolher_arquivo, variant="ghost", width=90)\
            .pack(side="left", padx=(8,0))

        _sep(inn)

        # tipo de busca — espelha o parâmetro "tipo" da API
        rt = ctk.CTkFrame(inn, fg_color="transparent"); rt.pack(fill="x", pady=3)
        _lbl(rt, "Buscar por:", color=C_TEXT_MED).pack(side="left", padx=(0,10))
        self.var_tipo1 = tk.StringVar(value=ROTULO_TIPO[TIPO_CATMAT])
        ctk.CTkSegmentedButton(
            rt, values=[ROTULO_TIPO[TIPO_CATMAT], ROTULO_TIPO[TIPO_PDM]],
            variable=self.var_tipo1, command=self._on_tipo_extracao,
            font=("Segoe UI", 12), width=180, corner_radius=6,
            fg_color=C_BG, selected_color=C_ACCENT, selected_hover_color=C_ACCENT_H,
            unselected_color=C_BG, unselected_hover_color=C_BORDER,
            text_color=C_TEXT).pack(side="left")
        self.lbl_hint_tipo = _lbl(rt, "coluna esperada no arquivo: codigoItemCatalogo",
                                  size=10, color=C_TEXT_LIGHT)
        self.lbl_hint_tipo.pack(side="left", padx=(12,0))

        # um arquivo por classe — a classe vem do próprio registro (codigoClasse),
        # ou de uma coluna "classe" no arquivo de entrada, quando houver
        rpc = ctk.CTkFrame(inn, fg_color="transparent"); rpc.pack(fill="x", pady=3)
        self._row_por_classe1 = rpc
        self.var_por_classe1 = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(rpc, text="Salvar um arquivo por classe",
                        variable=self.var_por_classe1,
                        command=self._toggle_pasta_classe1,
                        font=("Segoe UI", 12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER).pack(side="left")
        _lbl(rpc, "  (separa a saída em classe_XXXX; a classe vem dos próprios "
                  "registros ou da coluna \"classe\" do arquivo)",
             size=10, color=C_TEXT_LIGHT).pack(side="left")

        # Pasta de destino — os arquivos de cada classe são gravados direto nela,
        # sem diálogo ao final
        self.frame_pasta_classe1 = ctk.CTkFrame(inn, fg_color="transparent")
        row_pc = ctk.CTkFrame(self.frame_pasta_classe1, fg_color="transparent")
        row_pc.pack(fill="x")
        _lbl(row_pc, "Pasta de destino:", color=C_TEXT_MED, size=11)\
            .pack(side="left", padx=(0,8))
        self.var_pasta_classe1 = tk.StringVar()
        _entry(row_pc, textvariable=self.var_pasta_classe1,
               placeholder="Selecione a pasta onde os arquivos serão salvos",
               width=380).pack(side="left", expand=True, fill="x")
        _btn(row_pc, "📂  Procurar", self._escolher_pasta_classe1,
             variant="ghost", width=100).pack(side="left", padx=(8,0))

        _sep(inn)

        # datas
        r2 = ctk.CTkFrame(inn, fg_color="transparent"); r2.pack(fill="x", pady=3)
        _lbl(r2, "Data de Início:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_ini1 = tk.StringVar()
        _entry(r2, textvariable=self.var_ini1, placeholder="DD-MM-AAAA", width=130)\
            .pack(side="left")
        _lbl(r2, "Data Final:", color=C_TEXT_MED).pack(side="left", padx=(20,6))
        self.var_fim1 = tk.StringVar()
        _entry(r2, textvariable=self.var_fim1, placeholder="DD-MM-AAAA", width=130)\
            .pack(side="left")

        _sep(inn)

        # formato
        r3 = ctk.CTkFrame(inn, fg_color="transparent"); r3.pack(fill="x", pady=3)
        _lbl(r3, "Formato de Saída:", color=C_TEXT_MED).pack(side="left", padx=(0,12))
        self.var_fmt = tk.StringVar(value="xlsx")
        for txt, val in [("Excel (.xlsx)","xlsx"), ("CSV (.csv)","csv")]:
            ctk.CTkRadioButton(r3, text=txt, variable=self.var_fmt, value=val,
                               font=("Segoe UI",12), text_color=C_TEXT,
                               fg_color=C_ACCENT, border_color=C_BORDER)\
                .pack(side="left", padx=(0,16))

        _sep(inn)

        # corrompidos
        r4 = ctk.CTkFrame(inn, fg_color="transparent"); r4.pack(fill="x", pady=3)
        self.var_salvar_corr = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(r4, text="Salvar cópias dos CSV corrompidos",
                        variable=self.var_salvar_corr, command=self._toggle_pasta,
                        font=("Segoe UI",12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER)\
            .pack(side="left")
        self.frame_pasta = ctk.CTkFrame(inn, fg_color="transparent")
        rp = ctk.CTkFrame(self.frame_pasta, fg_color="transparent")
        rp.pack(fill="x", pady=3)
        _lbl(rp, "Pasta:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_pasta = tk.StringVar()
        _entry(rp, textvariable=self.var_pasta,
               placeholder="Pasta de destino", width=380)\
            .pack(side="left", expand=True, fill="x")
        _btn(rp, "Procurar…", self._escolher_pasta, variant="ghost", width=90)\
            .pack(side="left", padx=(8,0))

        # — Card 2: estatísticas —
        c2 = _card(wrap, "2.  Resumo da Execução")
        c2.pack(fill="x", pady=(0,8))
        grid = ctk.CTkFrame(c2, fg_color="transparent")
        grid.pack(fill="x", padx=14, pady=(0,8))
        stats = [
            ("Códigos Processados",   "k_proc",  C_ACCENT),
            ("Registros Consolidados","k_reg",   C_GREEN),
            ("Reparadas · Com Perda", "k_corr",  C_ORANGE),
            ("Códigos sem Dados",     "k_vaz",   C_RED),
        ]
        self._stats = {}
        for col, (nome, key, cor) in enumerate(stats):
            cell = ctk.CTkFrame(grid, fg_color=C_BG, corner_radius=6,
                                border_width=1, border_color=C_BORDER)
            cell.grid(row=0, column=col, padx=5, pady=4, sticky="ew")
            grid.grid_columnconfigure(col, weight=1)
            _lbl(cell, nome, size=10, color=C_TEXT_MED).pack(pady=(6,1))
            lv = ctk.CTkLabel(cell, text="0", font=("Segoe UI",17,"bold"),
                              text_color=cor)
            lv.pack(pady=(0,6))
            self._stats[key] = lv

        # — Card 3: log —
        c3 = _card(wrap, "3.  Log e Progresso")
        c3.pack(fill="x", pady=(0,4))

        brow = ctk.CTkFrame(c3, fg_color="transparent")
        brow.pack(fill="x", padx=14, pady=(0,4))
        self.lbl_status = _lbl(brow, "Status: Ocioso", size=11,
                                color=C_TEXT_MED, anchor="w")
        self.lbl_status.pack(side="left", expand=True, fill="x")
        self.lbl_pct = _lbl(brow, "0%", size=11, weight="bold", color=C_GREEN)
        self.lbl_pct.pack(side="right", padx=(8,0))

        self.progress = ctk.CTkProgressBar(c3, fg_color=C_BORDER,
                                           progress_color=C_GREEN,
                                           corner_radius=3, height=6)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=14, pady=(0,8))

        log_wrap = ctk.CTkFrame(c3, fg_color=C_LOG_BG, corner_radius=6)
        log_wrap.pack(fill="x", padx=14, pady=(0,10))
        self.log = scrolledtext.ScrolledText(
            log_wrap, bg=C_LOG_BG, fg=C_LOG_FG,
            font=("Consolas",10), wrap="word",
            relief="flat", bd=0, state="normal",
            height=9,
            insertbackground=C_LOG_FG)
        self.log.pack(fill="x", padx=6, pady=6)
        for tag, cor in [("ok","#4EC94E"),("warn","#F4A11D"),
                         ("err","#E05C5C"),("info","#7EB8F7"),
                         ("date","#FFCD07")]:
            self.log.tag_config(tag, foreground=cor)
        self._log(WELCOME, "info")

        # botões ficam no wrap (fora do card), sempre visíveis
        br = ctk.CTkFrame(wrap, fg_color="transparent")
        br.pack(fill="x", pady=(4,4))
        self.btn_start = _btn(br, "▶  Iniciar Extração", self._start,
                              variant="primary", width=160)
        self.btn_start.pack(side="left", padx=(0,8))
        self.btn_cancel = _btn(br, "✖  Cancelar", self._cancelar,
                               variant="secondary", width=100)
        self.btn_cancel.configure(state="disabled")
        self.btn_cancel.pack(side="left", padx=(0,8))
        self.btn_pause = _btn(br, "⏸  Pausar", self._pausar,
                              variant="secondary", width=100)
        self.btn_pause.configure(state="disabled")
        self.btn_pause.pack(side="left", padx=(0,8))
        self.btn_log = _btn(br, "💾  Salvar Log", self._salvar_log,
                            variant="secondary", width=120)
        self.btn_log.configure(state="disabled")
        self.btn_log.pack(side="left")

    # ── ABA 2 ─────────────────────────────────────────────────────────────────
    def _build_tab_explorador(self, parent):
        parent.configure(fg_color=C_BG)

        # ── Card 1: Classes (topo) ────────────────────────────────────────────
        c1 = _card(parent, "1.  Buscar PDMs por Classes")
        c1.pack(fill="x", padx=12, pady=(8,6))
        inn1 = ctk.CTkFrame(c1, fg_color="transparent")
        inn1.pack(fill="x", padx=14, pady=(0,10))
        _lbl(inn1, "Informe as Classes para extração separadas por  ;",
             color=C_TEXT_MED, size=11).pack(anchor="w", pady=(0,4))
        r = ctk.CTkFrame(inn1, fg_color="transparent")
        r.pack(fill="x")
        self.var_classe = tk.StringVar()
        ent = _entry(r, textvariable=self.var_classe,
                     placeholder="ex.: 20115 ; 20116 ; 20117", width=400)
        ent.pack(side="left", expand=True, fill="x")
        ent.bind("<Return>", lambda e: self._buscar_pdms())
        _btn(r, "Buscar PDMs", self._buscar_pdms,
             variant="primary", width=140).pack(side="left", padx=(10,0))
        _btn(r, "⚡  Buscar e Extrair", self._buscar_e_extrair_classes,
             variant="success", width=160).pack(side="left", padx=(8,0))
        self.lbl_pdm_count = _lbl(r, "", size=11, color=C_GREEN)
        self.lbl_pdm_count.pack(side="right", padx=8)
        row_chk = ctk.CTkFrame(inn1, fg_color="transparent")
        row_chk.pack(fill="x", pady=(8,0))
        self.var_arquivo_por_classe = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(row_chk,
                        text="Deseja salvar um arquivo por classe?",
                        variable=self.var_arquivo_por_classe,
                        command=self._toggle_pasta_por_classe,
                        font=("Segoe UI", 12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER)            .pack(side="left")
        _lbl(row_chk, "  (gera um arquivo separado para cada classe informada)",
             size=10, color=C_TEXT_LIGHT).pack(side="left")

        # Extração direta por PDM — usa tipo=codigoPdm na Pesquisa de Preço
        row_pdm = ctk.CTkFrame(inn1, fg_color="transparent")
        row_pdm.pack(fill="x", pady=(4,0))
        self.var_extrair_por_pdm = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(row_pdm,
                        text="Extrair preços direto por PDM (tipo=codigoPdm)",
                        variable=self.var_extrair_por_pdm,
                        font=("Segoe UI", 12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER).pack(side="left")
        _lbl(row_pdm, "  (dispensa a expansão PDM → CATMAT: muito mais rápido)",
             size=10, color=C_TEXT_LIGHT).pack(side="left")
        # Linha de pasta de destino — visível só quando checkbox marcado
        self.frame_pasta_classes = ctk.CTkFrame(inn1, fg_color="transparent")
        row_pasta = ctk.CTkFrame(self.frame_pasta_classes, fg_color="transparent")
        row_pasta.pack(fill="x")
        _lbl(row_pasta, "Pasta de destino:", color=C_TEXT_MED, size=11)            .pack(side="left", padx=(0,8))
        self.var_pasta_classes = tk.StringVar()
        _entry(row_pasta, textvariable=self.var_pasta_classes,
               placeholder="Selecione a pasta onde os arquivos serão salvos",
               width=380).pack(side="left", expand=True, fill="x")
        _btn(row_pasta, "📂  Procurar", self._escolher_pasta_classes,
             variant="ghost", width=100).pack(side="left", padx=(8,0))

        # ── Área central: tabela (esquerda) + busca avulsa (direita) ─────────
        mid = ctk.CTkFrame(parent, fg_color=C_BG)
        mid.pack(fill="both", expand=True, padx=12, pady=(0,4))

        # Painel lateral DIREITO: Busca Avulsa por PDMs
        # — deve ser empacotado ANTES do c2 para reservar espaço antes do expand
        cav = _card(mid, "Busca Avulsa por PDMs")
        cav.pack(side="right", fill="y", padx=(6,0))
        _lbl(cav, "Códigos PDM (um por linha):",
             size=11, color=C_TEXT_MED).pack(anchor="w", padx=14, pady=(0,4))
        self.txt_avulso = ctk.CTkTextbox(cav, font=("Consolas",11),
                                         fg_color=C_SURFACE, text_color=C_TEXT,
                                         border_width=1, border_color=C_BORDER,
                                         width=190, corner_radius=6)
        self.txt_avulso.pack(fill="both", expand=True, padx=14)
        _btn(cav, "🔍  Buscar CATMATs\n(PDMs da lista)",
             self._buscar_avulso, variant="primary")            .pack(fill="x", padx=14, pady=(8,10))

        # Card 2: PDMs Encontrados — expande para ocupar o espaço restante
        c2 = _card(mid, "2.  PDMs Encontrados")
        c2.pack(side="left", fill="both", expand=True)

        frow = ctk.CTkFrame(c2, fg_color="transparent")
        frow.pack(fill="x", padx=14, pady=(0,4))
        _lbl(frow, "Filtro:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_filtro = tk.StringVar(value="todos")
        for txt, val in [("Todos","todos"),("Ativos","ativo"),("Inativos","inativo")]:
            ctk.CTkRadioButton(frow, text=txt, variable=self.var_filtro, value=val,
                               command=self._filtrar,
                               font=("Segoe UI",11), text_color=C_TEXT,
                               fg_color=C_ACCENT, border_color=C_BORDER)                .pack(side="left", padx=(0,10))
        _btn(frow, "🔍  Buscar CATMATs (PDMs da tabela)", self._buscar_catmats,
             variant="secondary", width=240).pack(side="left", padx=(16,0))
        _btn(frow, "Exportar PDMs", self._exp_pdms,
             variant="ghost", width=120).pack(side="right")

        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("BPS.Treeview", background=C_SURFACE, foreground=C_TEXT,
                        fieldbackground=C_SURFACE, rowheight=26,
                        font=("Segoe UI",10), borderwidth=0)
        style.configure("BPS.Treeview.Heading", background=C_BG,
                        foreground=C_TEXT_MED, font=("Segoe UI",10,"bold"),
                        relief="flat")
        style.map("BPS.Treeview",
                  background=[("selected", C_ACCENT)],
                  foreground=[("selected", C_SURFACE)])

        tf = ctk.CTkFrame(c2, fg_color=C_SURFACE, corner_radius=0)
        tf.pack(fill="both", expand=True, padx=8, pady=(0,8))
        vsb = ttk.Scrollbar(tf, orient="vertical", command=None)
        vsb.pack(side="right", fill="y")
        self.tree = ttk.Treeview(tf, columns=("cod","desc","status"),
                                 show="headings", style="BPS.Treeview",
                                 selectmode="extended",
                                 yscrollcommand=vsb.set)
        vsb.configure(command=self.tree.yview)
        self.tree.heading("cod",    text="Cód. PDM")
        self.tree.heading("desc",   text="Descrição")
        self.tree.heading("status", text="Status")
        self.tree.column("cod",    width=90,   anchor="center", stretch=False)
        self.tree.column("desc",   width=9999, anchor="w",      stretch=True)
        self.tree.column("status", width=80,   anchor="center", stretch=False)
        self.tree.pack(fill="both", expand=True)

        # ── Card 3: Ações (rodapé) ────────────────────────────────────────────
        c3 = _card(parent, "3.  Ações")
        c3.pack(fill="x", padx=12, pady=(0,8))

        ar = ctk.CTkFrame(c3, fg_color="transparent")
        ar.pack(fill="x", padx=14, pady=(0,4))
        _btn(ar, "⚡  Buscar e Extrair", self._buscar_e_extrair,
             variant="primary").pack(side="left", padx=(0,8))
        self.btn_exp_cat = _btn(ar, "📥  Exportar CATMATs Encontrados",
                                self._exp_catmats, variant="ghost")
        self.btn_exp_cat.configure(state="disabled")
        self.btn_exp_cat.pack(side="left", padx=(0,16))
        self.lbl_exp_status = _lbl(ar, "", size=11, color=C_TEXT_MED)
        self.lbl_exp_status.pack(side="left", expand=True, fill="x")

        dr = ctk.CTkFrame(c3, fg_color="transparent")
        dr.pack(fill="x", padx=14, pady=(0,4))
        _lbl(dr, "Data de Início:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_ini2 = tk.StringVar()
        _entry(dr, textvariable=self.var_ini2, placeholder="DD-MM-AAAA", width=130)            .pack(side="left")
        _lbl(dr, "Data Final:", color=C_TEXT_MED).pack(side="left", padx=(16,6))
        self.var_fim2 = tk.StringVar()
        _entry(dr, textvariable=self.var_fim2, placeholder="DD-MM-AAAA", width=130)            .pack(side="left")

        cr = ctk.CTkFrame(c3, fg_color="transparent")
        cr.pack(fill="x", padx=14, pady=(0,10))
        self.btn_pb = _btn(cr, "⏸  Pausar Busca", self._pausar_busca,
                           variant="secondary", width=130)
        self.btn_pb.configure(state="disabled")
        self.btn_pb.pack(side="left", padx=(0,8))
        self.btn_cb = _btn(cr, "✖  Cancelar Busca", self._cancelar_busca,
                           variant="danger", width=130)
        self.btn_cb.configure(state="disabled")
        self.btn_cb.pack(side="left", padx=(0,16))
        self.btn_ini_exp = _btn(cr,
            "▶  Iniciar Extração com CATMATs Encontrados",
            self._iniciar_exp, variant="success")
        self.btn_ini_exp.configure(state="disabled")
        self.btn_ini_exp.pack(side="left")

    # ── ABA 3 ─────────────────────────────────────────────────────────────────
    def _build_tab_consolidacao(self, parent):
        """Consolida DW (SIASG) + DA (Compras.gov) removendo do DA o que já está
        no DW. Motor: consolidar_dw_da.consolidar()."""
        parent.configure(fg_color=C_BG)

        # estado próprio da aba — nada é compartilhado com as abas 1 e 2
        self._cons_entradas   = []      # [{"caminho": str, "origem": "auto|dw|da"}]
        self._cons_rodando    = False
        self._cons_cancelar   = False
        self._cons_ultima_saida = ""

        # ── Card 1: entradas ─────────────────────────────────────────────────
        c1 = _card(parent, "1.  Arquivos de Entrada  (CSV/XLSX do DW e do DA)")
        c1.pack(fill="x", padx=12, pady=(8,6))
        inn1 = ctk.CTkFrame(c1, fg_color="transparent")
        inn1.pack(fill="x", padx=14, pady=(0,10))

        rb = ctk.CTkFrame(inn1, fg_color="transparent"); rb.pack(fill="x", pady=(0,6))
        _lbl(rb, "Origem:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_cons_origem = tk.StringVar(value="Detectar")
        ctk.CTkSegmentedButton(
            rb, values=["Detectar", "DW", "DA"], variable=self.var_cons_origem,
            font=("Segoe UI", 12), width=200, corner_radius=6,
            fg_color=C_BG, selected_color=C_ACCENT, selected_hover_color=C_ACCENT_H,
            unselected_color=C_BG, unselected_hover_color=C_BORDER,
            text_color=C_TEXT).pack(side="left")
        _btn(rb, "📁  Pasta…", self._cons_add_pasta, variant="primary", width=100)\
            .pack(side="left", padx=(12,6))
        _btn(rb, "📄  Arquivos…", self._cons_add_arquivos, variant="secondary",
             width=110).pack(side="left", padx=(0,6))
        _btn(rb, "Remover", self._cons_remover, variant="ghost", width=90)\
            .pack(side="left", padx=(0,6))
        _btn(rb, "Limpar", self._cons_limpar, variant="ghost", width=80)\
            .pack(side="left")
        _lbl(rb, "  \"Detectar\" descobre DW/DA pelo cabeçalho do arquivo",
             size=10, color=C_TEXT_LIGHT).pack(side="left", padx=(10,0))

        tf = ctk.CTkFrame(inn1, fg_color=C_SURFACE, corner_radius=0)
        tf.pack(fill="x")
        vsb_c = ttk.Scrollbar(tf, orient="vertical")
        vsb_c.pack(side="right", fill="y")
        self.tree_cons = ttk.Treeview(tf, columns=("origem","caminho"),
                                      show="headings", style="BPS.Treeview",
                                      selectmode="extended", height=5,
                                      yscrollcommand=vsb_c.set)
        vsb_c.configure(command=self.tree_cons.yview)
        self.tree_cons.heading("origem",  text="Origem")
        self.tree_cons.heading("caminho", text="Pasta / Arquivo")
        self.tree_cons.column("origem",  width=90, anchor="center", stretch=False)
        self.tree_cons.column("caminho", width=9999, anchor="w", stretch=True)
        self.tree_cons.pack(fill="x")

        # ── Card 2: saída e opções ───────────────────────────────────────────
        c2 = _card(parent, "2.  Saída e Opções")
        c2.pack(fill="x", padx=12, pady=(0,6))
        inn2 = ctk.CTkFrame(c2, fg_color="transparent")
        inn2.pack(fill="x", padx=14, pady=(0,10))

        rs = ctk.CTkFrame(inn2, fg_color="transparent"); rs.pack(fill="x", pady=3)
        _lbl(rs, "Pasta de saída:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_cons_saida = tk.StringVar()
        _entry(rs, textvariable=self.var_cons_saida,
               placeholder="Onde as planilhas por classe serão gravadas",
               width=420).pack(side="left", expand=True, fill="x")
        _btn(rs, "📂  Procurar", self._cons_escolher_saida, variant="ghost",
             width=100).pack(side="left", padx=(8,0))

        ro = ctk.CTkFrame(inn2, fg_color="transparent"); ro.pack(fill="x", pady=3)
        _lbl(ro, "Prefixo:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_cons_prefixo = tk.StringVar(value="bps_dw_da__Classe_")
        _entry(ro, textvariable=self.var_cons_prefixo, width=180).pack(side="left")
        _lbl(ro, "Sufixo:", color=C_TEXT_MED).pack(side="left", padx=(14,6))
        self.var_cons_sufixo = tk.StringVar()
        _entry(ro, textvariable=self.var_cons_sufixo,
               placeholder="ex.: _2021_a_2026", width=140).pack(side="left")
        _lbl(ro, "Ano mín.:", color=C_TEXT_MED).pack(side="left", padx=(14,6))
        self.var_cons_ano_min = tk.StringVar()
        _entry(ro, textvariable=self.var_cons_ano_min, placeholder="AAAA",
               width=70).pack(side="left")
        _lbl(ro, "Ano máx.:", color=C_TEXT_MED).pack(side="left", padx=(10,6))
        self.var_cons_ano_max = tk.StringVar()
        _entry(ro, textvariable=self.var_cons_ano_max, placeholder="AAAA",
               width=70).pack(side="left")

        rc = ctk.CTkFrame(inn2, fg_color="transparent"); rc.pack(fill="x", pady=3)
        self.var_cons_dup = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(rc, text="Gerar CSV de auditoria com as duplicatas removidas",
                        variable=self.var_cons_dup, font=("Segoe UI",12),
                        text_color=C_TEXT, fg_color=C_ACCENT,
                        border_color=C_BORDER).pack(side="left", padx=(0,20))
        self.var_cons_dedup_interno = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(rc, text="Também remover repetições internas do DA",
                        variable=self.var_cons_dedup_interno, font=("Segoe UI",12),
                        text_color=C_TEXT, fg_color=C_ACCENT,
                        border_color=C_BORDER).pack(side="left")
        _lbl(rc, "  (atenção: costumam ser registros distintos, com fornecedor "
                 "e preço diferentes)", size=10, color=C_TEXT_LIGHT).pack(side="left")

        # ── Resumo ───────────────────────────────────────────────────────────
        grid = ctk.CTkFrame(parent, fg_color=C_BG)
        grid.pack(fill="x", padx=12, pady=(0,6))
        stats = [
            ("Linhas do DW",        "c_dw",   C_ACCENT),
            ("Linhas do DA lidas",  "c_lidas", C_TEXT_MED),
            ("Duplicatas Removidas","c_dup",  C_ORANGE),
            ("Linhas do DA Mantidas","c_mant", C_GREEN),
        ]
        self._stats_cons = {}
        for col, (nome, key, cor) in enumerate(stats):
            cell = ctk.CTkFrame(grid, fg_color=C_SURFACE, corner_radius=6,
                                border_width=1, border_color=C_BORDER)
            cell.grid(row=0, column=col, padx=5, pady=0, sticky="ew")
            grid.grid_columnconfigure(col, weight=1)
            _lbl(cell, nome, size=10, color=C_TEXT_MED).pack(pady=(6,1))
            lv = ctk.CTkLabel(cell, text="0", font=("Segoe UI",17,"bold"),
                              text_color=cor)
            lv.pack(pady=(0,6))
            self._stats_cons[key] = lv

        # ── Card 3: log e progresso ──────────────────────────────────────────
        c3 = _card(parent, "3.  Log e Progresso")
        c3.pack(fill="both", expand=True, padx=12, pady=(0,4))

        brow = ctk.CTkFrame(c3, fg_color="transparent")
        brow.pack(fill="x", padx=14, pady=(0,4))
        self.lbl_cons_status = _lbl(brow, "Status: Ocioso", size=11,
                                    color=C_TEXT_MED, anchor="w")
        self.lbl_cons_status.pack(side="left", expand=True, fill="x")
        self.lbl_cons_pct = _lbl(brow, "0%", size=11, weight="bold", color=C_GREEN)
        self.lbl_cons_pct.pack(side="right", padx=(8,0))

        self.progress_cons = ctk.CTkProgressBar(c3, fg_color=C_BORDER,
                                                progress_color=C_GREEN,
                                                corner_radius=3, height=6)
        self.progress_cons.set(0)
        self.progress_cons.pack(fill="x", padx=14, pady=(0,8))

        log_wrap = ctk.CTkFrame(c3, fg_color=C_LOG_BG, corner_radius=6)
        log_wrap.pack(fill="both", expand=True, padx=14, pady=(0,10))
        self.log_cons = scrolledtext.ScrolledText(
            log_wrap, bg=C_LOG_BG, fg=C_LOG_FG, font=("Consolas",10),
            wrap="word", relief="flat", bd=0, state="normal", height=7,
            insertbackground=C_LOG_FG)
        self.log_cons.pack(fill="both", expand=True, padx=6, pady=6)
        for tag, cor in [("ok","#4EC94E"),("warn","#F4A11D"),
                         ("err","#E05C5C"),("info","#7EB8F7")]:
            self.log_cons.tag_config(tag, foreground=cor)

        # ── Botões ───────────────────────────────────────────────────────────
        br = ctk.CTkFrame(parent, fg_color="transparent")
        br.pack(fill="x", padx=12, pady=(0,8))
        self.btn_cons_start = _btn(br, "▶  Consolidar e Remover Duplicatas",
                                   self._cons_iniciar, variant="primary", width=250)
        self.btn_cons_start.pack(side="left", padx=(0,8))
        self.btn_cons_cancel = _btn(br, "✖  Cancelar", self._cons_cancelar_click,
                                    variant="secondary", width=100)
        self.btn_cons_cancel.configure(state="disabled")
        self.btn_cons_cancel.pack(side="left", padx=(0,8))
        self.btn_cons_abrir = _btn(br, "📂  Abrir Pasta de Saída",
                                   self._cons_abrir_pasta, variant="ghost", width=170)
        self.btn_cons_abrir.configure(state="disabled")
        self.btn_cons_abrir.pack(side="left", padx=(0,8))
        self.btn_cons_log = _btn(br, "💾  Salvar Log", self._cons_salvar_log,
                                 variant="secondary", width=120)
        self.btn_cons_log.configure(state="disabled")
        self.btn_cons_log.pack(side="left")

        self._log_cons(BEMVINDO_CONSOLIDACAO, "info")

        # ── LOG HELPERS ───────────────────────────────────────────────────────────
    def _log(self, msg: str, tag: str = ""):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    def set_status(self, txt): self.lbl_status.configure(text=txt)
    def set_status_explorador(self, txt): self.lbl_exp_status.configure(text=txt)
    def _stat(self, key, val): self._stats[key].configure(text=str(val))

    def _atualizar_stat_paginas(self):
        """Card de páginas: reparadas (recuperadas) · com perda (registros sumiram)."""
        self._stat("k_corr", f"{self.count_reparadas} · {self.count_corrigidas}")



    # ── CALLBACKS ABA 1 ───────────────────────────────────────────────────────
    def _escolher_arquivo(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel/CSV","*.xlsx *.csv"),("Todos","*.*")])
        if p: self.var_arquivo.set(p)

    def _escolher_pasta(self):
        p = filedialog.askdirectory()
        if p: self.var_pasta.set(p)

    def _toggle_pasta(self):
        if self.var_salvar_corr.get():
            self.frame_pasta.pack(fill="x", padx=14, pady=3)
        else:
            self.frame_pasta.pack_forget()

    def _toggle_pasta_classe1(self):
        """Mostra o campo de pasta apenas quando o modo por classe está ativo."""
        if self.var_por_classe1.get():
            # after= ancora o campo logo abaixo do checkbox; sem isso o pack
            # jogaria o frame para o fim da seção, longe do que ele controla
            self.frame_pasta_classe1.pack(fill="x", padx=0, pady=(2,0),
                                          after=self._row_por_classe1)
        else:
            self.frame_pasta_classe1.pack_forget()
            self.var_pasta_classe1.set("")

    def _escolher_pasta_classe1(self):
        p = filedialog.askdirectory(
            title="Escolha a pasta de destino dos arquivos por classe")
        if p:
            self.var_pasta_classe1.set(p)

    def _on_tipo_extracao(self, valor=None):
        """Atualiza a dica de coluna quando o usuário troca CATMAT ↔ PDM."""
        tipo = TIPO_POR_ROTULO.get(self.var_tipo1.get(), TIPO_CATMAT)
        self.lbl_hint_tipo.configure(
            text=f"coluna esperada no arquivo: {tipo}")

    def _tipo_extracao(self):
        """Tipo de busca selecionado na aba de extração."""
        return TIPO_POR_ROTULO.get(self.var_tipo1.get(), TIPO_CATMAT)

    def _start(self):
        arq = self.var_arquivo.get().strip()
        if not arq:
            messagebox.showerror("Arquivo obrigatório",
                                 "Selecione um arquivo de códigos."); return
        tipo = self._tipo_extracao()
        por_classe = self.var_por_classe1.get()
        mapa = {}
        try:
            df_c = pd.read_excel(arq) if arq.lower().endswith(".xlsx") \
                   else pd.read_csv(arq, sep=";")
            # Aceita a coluna do tipo escolhido ou a coluna genérica 'codigo'
            col = next((c for c in (tipo, "codigo") if c in df_c.columns), None)
            if col is None:
                messagebox.showerror("Coluna ausente",
                    f"O arquivo deve ter a coluna '{tipo}' (ou 'codigo')."); return
            codigos = pd.Series(df_c[col]).dropna()\
                        .astype(int).drop_duplicates().tolist()

            # Agrupamento explícito: se o arquivo trouxer a classe de cada código,
            # ela manda. Sem essa coluna, a classe é lida do próprio registro.
            if por_classe:
                col_cl = next((c for c in ("classe", "Classe", "codigoClasse")
                               if c in df_c.columns), None)
                if col_cl:
                    val = df_c[[col, col_cl]].dropna()
                    for cl, grupo in val.groupby(val[col_cl].astype(str)
                                                    .str.strip().str.replace(r"\.0$", "", regex=True)):
                        mapa[cl] = grupo[col].astype(int).drop_duplicates().tolist()
        except Exception as e:
            messagebox.showerror("Erro ao ler arquivo", str(e)); return
        d_i, d_f, err = validar_e_obter_datas(self.var_ini1.get(), self.var_fim1.get())
        if err: messagebox.showerror("Data inválida", err); return

        # Modo por classe grava vários arquivos: a pasta precisa ser conhecida
        # ANTES de começar, para que cada classe seja salva em ato contínuo.
        pasta = ""
        if por_classe:
            pasta = self.var_pasta_classe1.get().strip()
            if not pasta:
                messagebox.showinfo("Pasta de destino",
                    "A opção 'um arquivo por classe' gera vários arquivos.\n\n"
                    "Escolha a pasta onde eles serão salvos.")
                pasta = filedialog.askdirectory(
                    title="Escolha a pasta de destino dos arquivos por classe")
                if not pasta:
                    messagebox.showwarning("Extração não iniciada",
                        "Nenhuma pasta escolhida. Selecione a pasta de destino "
                        "ou desmarque 'Salvar um arquivo por classe'."); return
                self.var_pasta_classe1.set(pasta)
            if not os.path.isdir(pasta):
                messagebox.showerror("Pasta inválida",
                    f"A pasta não existe:\n{pasta}"); return
            origem = (f"coluna '{col_cl}' do arquivo" if mapa
                      else "campo codigoClasse dos registros")
            self._log(f"🗂️  Um arquivo por classe — origem da classe: {origem}.", "info")
            self._log(f"📂 Destino: {pasta}", "info")

        self._iniciar_processo(codigos, self.var_fmt.get(), d_i, d_f,
                               catmats_por_classe=mapa, tipo=tipo,
                               por_classe=por_classe, pasta_destino=pasta)

    def _cancelar(self):
        if not self.processing: return
        self.processing = False
        pausar_extracao.set()   # desbloqueia wait() na thread para ela poder sair

    def _pausar(self):
        if pausar_extracao.is_set():
            # Estava rodando → pausar
            pausar_extracao.clear()
            self.btn_pause.configure(text="▶  Retomar")
            self.set_status("Status: Pausado")
        else:
            # Estava pausado → retomar
            pausar_extracao.set()
            self.btn_pause.configure(text="⏸  Pausar")
            self.set_status("Status: Retomando…")

    def _pausar_por_conexao(self):
        """Pausa automática ao detectar queda de rede — NÃO cancela a extração.
        Retoma automaticamente em 60s ou imediatamente se o usuário clicar Retomar."""
        if not self.processing: return
        # Só loga/pausa se ainda não estava pausado por conexão
        if not pausar_extracao.is_set(): return  # já está pausado
        pausar_extracao.clear()
        self.btn_pause.configure(state="normal", text="▶  Retomar agora")
        self.set_status("Status: Sem conexão — retentando em 60s")
        self._log("\n⚠️  Rede indisponível — retentando automaticamente em 60s.\n"
                  "   Clique em ▶  Retomar agora para tentar imediatamente.", "warn")
        # Atualizar contagem regressiva no status a cada segundo
        def _countdown(seg):
            if not self.processing or pausar_extracao.is_set(): return
            self.set_status(f"Status: Sem conexão — retentando em {seg}s")
            if seg > 0:
                self.after(1000, lambda: _countdown(seg - 1))
        self.after(1000, lambda: _countdown(59))

    def _salvar_log(self):
        p = filedialog.asksaveasfilename(defaultextension=".txt",
                                         filetypes=[("Texto","*.txt")])
        if p:
            try:
                with open(p,"w",encoding="utf-8") as f: f.write(self.log.get("1.0","end"))
                messagebox.showinfo("Salvo", f"Log salvo em:\n{p}")
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    # ── CALLBACKS ABA 2 ───────────────────────────────────────────────────────
    # ── SPINNER (overlay translúcido) ─────────────────────────────────────────
    def _show_spinner(self, msg="Buscando…"):
        """Exibe overlay com spinner animado sobre a aba."""
        self._spinner_active = True
        self._spinner_frame = ctk.CTkFrame(self, fg_color="#FFFFFF",
                                           corner_radius=12,
                                           border_width=1, border_color=C_BORDER)
        self._spinner_frame.place(relx=0.5, rely=0.5, anchor="center")
        self._spinner_chars = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
        self._spinner_idx   = 0
        self._lbl_spin_icon = ctk.CTkLabel(self._spinner_frame,
            text=self._spinner_chars[0],
            font=("Segoe UI", 28), text_color=C_ACCENT)
        self._lbl_spin_icon.pack(padx=40, pady=(22,4))
        self._lbl_spin_msg = ctk.CTkLabel(self._spinner_frame,
            text=msg, font=("Segoe UI", 13), text_color=C_TEXT_MED)
        self._lbl_spin_msg.pack(padx=40, pady=(0,22))
        self._animate_spinner()

    def _animate_spinner(self):
        if not self._spinner_active: return
        self._spinner_idx = (self._spinner_idx + 1) % len(self._spinner_chars)
        self._lbl_spin_icon.configure(text=self._spinner_chars[self._spinner_idx])
        self.after(80, self._animate_spinner)

    def _hide_spinner(self):
        self._spinner_active = False
        if hasattr(self, "_spinner_frame") and self._spinner_frame.winfo_exists():
            self._spinner_frame.destroy()

    def _buscar_pdms(self, acao_pos_busca=None):
        entrada = self.var_classe.get().strip()
        if not entrada:
            messagebox.showerror("Campo vazio", "Informe ao menos um código de Classe."); return

        partes = [p.strip() for p in entrada.split(";") if p.strip()]
        invalidas = [p for p in partes if not p.isdigit()]
        if invalidas:
            messagebox.showerror("Código inválido",
                f"Valores não numéricos: {', '.join(invalidas)}\n"
                "Use apenas números separados por ;")
            return

        self._show_spinner(f"Buscando PDMs de {len(partes)} classe(s)…")
        self.lbl_pdm_count.configure(text="")

        def _thread():
            todos_dfs = []; erros = []
            for cod in partes:
                self.after(0, lambda c=cod: self._lbl_spin_msg.configure(
                    text=f"Buscando classe {c}…"))
                res = buscar_pdms_por_classe(int(cod), URL_BASE, TIMEOUT)
                if res is None:
                    erros.append(cod)
                else:
                    df_c, _ = res
                    todos_dfs.append(df_c)
            self.after(0, lambda: self._on_pdms_carregados(todos_dfs, erros, acao_pos_busca))

        threading.Thread(target=_thread, daemon=True).start()

    def _on_pdms_carregados(self, todos_dfs, erros, acao_pos_busca=None):
        self._hide_spinner()
        if not todos_dfs:
            self.lbl_pdm_count.configure(text="Nenhum PDM encontrado.")
            self.lista_pdms_df = pd.DataFrame(); self._fill_tree([]); return

        df = pd.concat(todos_dfs, ignore_index=True).drop_duplicates(subset=["codigoPdm"])
        df["_col_codigo"] = df["codigoPdm"]
        df["_col_desc"]   = df["nomePdm"]
        df["_col_status"] = df["statusPdm"]
        self.lista_pdms_df        = df
        self._todos_dfs_por_classe = todos_dfs  # preservar para mapa classe→catmats
        self._fill_tree([[r["_col_codigo"], r["_col_desc"], r["_col_status"]]
                         for _, r in df.iterrows()])

        msg = f"{len(df)} PDMs de {len(todos_dfs)} classe(s)"
        if erros: msg += f"  ·  ⚠ Falha: {', '.join(erros)}"
        self.lbl_pdm_count.configure(text=msg)
        if erros:
            messagebox.showwarning("Classes com falha",
                f"Não foi possível buscar: {', '.join(erros)}")
        self.var_filtro.set("todos")
        self.btn_exp_cat.configure(state="disabled")
        self.btn_ini_exp.configure(state="disabled")
        self.lista_catmats = []

        if acao_pos_busca == "extrair":
            self._continuar_busca_e_extrai(df)

    def _filtrar(self):
        if self.lista_pdms_df.empty: return
        f = self.var_filtro.get(); df = self.lista_pdms_df
        if f == "ativo":    df = df[df["_col_status"] == "Ativo"]
        elif f == "inativo": df = df[df["_col_status"] == "Inativo"]
        self._fill_tree([[r["_col_codigo"], r["_col_desc"], r["_col_status"]]
                         for _, r in df.iterrows()])
        # Manter checkbox em sincronia após filtrar
        if hasattr(self, "var_sel_todos"): self.var_sel_todos.set(False)

    def _toggle_selecionar_todos(self):
        """Seleciona ou deseleciona todos os itens visíveis na tabela."""
        items = self.tree.get_children()
        if self.var_sel_todos.get():
            self.tree.selection_set(items)
        else:
            self.tree.selection_remove(items)

    def _fill_tree(self, rows):
        for i in self.tree.get_children(): self.tree.delete(i)
        for row in rows:
            tag = "ativo" if str(row[2]).lower() == "ativo" else "inativo"
            self.tree.insert("", "end", values=row, tags=(tag,))
        self.tree.tag_configure("ativo",   foreground=C_GREEN)
        self.tree.tag_configure("inativo", foreground=C_TEXT_LIGHT)
        # Resetar checkbox ao recarregar
        if hasattr(self, "var_sel_todos"): self.var_sel_todos.set(False)

    def _exp_pdms(self):
        rows = [self.tree.item(i,"values") for i in self.tree.get_children()]
        if not rows: messagebox.showerror("Vazio","Nenhum PDM."); return
        p = filedialog.asksaveasfilename(defaultextension=".csv",
                                         initialfile="PDMs_exportados.csv",
                                         filetypes=[("CSV","*.csv")])
        if p:
            pd.DataFrame(rows, columns=["Código PDM","Descrição","Status"])\
              .to_csv(p, index=False, sep=";", encoding="utf-8-sig")
            messagebox.showinfo("Exportado", f"Salvo em:\n{p}")

    def _pdms_selecionados_codigos(self):
        """Retorna lista de codigoPdm dos itens selecionados na árvore."""
        sel = self.tree.selection()
        return [int(self.tree.item(i,"values")[0]) for i in sel] if sel else []

    def _pdms_sel(self):
        sel = self.tree.selection()
        return [int(self.tree.item(i,"values")[0]) for i in sel] if sel else []

    def _buscar_avulso(self):
        txt = self.txt_avulso.get("1.0", "end").strip()
        if not txt: messagebox.showerror("Vazio","Informe ao menos um código PDM."); return
        pdms = []; inv = []
        # Aceita separador ; ou nova linha
        for parte in txt.replace("\n", ";").split(";"):
            parte = parte.strip()
            if not parte: continue
            try: pdms.append(int(parte))
            except ValueError: inv.append(parte)
        if inv: messagebox.showwarning("Inválidos", f"Ignorados: {', '.join(inv)}")
        if pdms: self._start_busca(pdms, "apenas_buscar")

    def _buscar_catmats(self):
        pdms = self._pdms_sel()
        if not pdms: messagebox.showerror("Nenhum selecionado","Selecione PDMs."); return
        self._start_busca(pdms, "apenas_buscar")

    def _buscar_e_extrair(self):
        pdms = self._pdms_sel()
        if not pdms: messagebox.showerror("Nenhum selecionado","Selecione PDMs."); return
        # Modo direto: pula a descoberta de CATMATs e consulta a Pesquisa de
        # Preço com tipo=codigoPdm
        if self.var_extrair_por_pdm.get():
            d_i, d_f, err = validar_e_obter_datas(self.var_ini2.get(), self.var_fim2.get())
            if err: messagebox.showerror("Data inválida", err); return
            self._log(f"⏩ {len(pdms)} PDMs — extração direta (tipo=codigoPdm).", "info")
            self._iniciar_processo(pdms, self.var_fmt.get(), d_i, d_f, tipo=TIPO_PDM)
            self.after(100, lambda: self.tabs.set("  Extração por CATMAT  "))
            return
        self._start_busca(pdms, "extrair")

    def _start_busca(self, pdms, acao):
        global cancelar_busca_catmat
        cancelar_busca_catmat = False
        pausar_busca_catmat.set()
        self.btn_pb.configure(state="normal")
        self.btn_cb.configure(state="normal")
        self.btn_pb.configure(text="⏸  Pausar Busca")
        self.set_status_explorador("Iniciando busca…")
        threading.Thread(target=self._thread_busca,
                         args=(pdms, acao), daemon=True).start()

    def _thread_busca(self, pdms, acao):
        df1, err1 = buscar_catmats_por_pdm(pdms, URL_BASE, TIMEOUT, self)
        df2 = None; err2 = []
        if err1 and not cancelar_busca_catmat:
            self.after(0, lambda: self.set_status_explorador(
                f"2ª tentativa para {len(err1)} PDMs…"))
            time.sleep(2)
            df2, err2 = buscar_catmats_por_pdm(err1, URL_BASE, TIMEOUT, self)
        dfs = [d for d in [df1, df2] if d is not None and not d.empty]
        df_final = pd.concat(dfs, ignore_index=True) if dfs else None
        self.after(0, lambda: self._on_busca(df_final, err2, acao))

    def _on_busca(self, df, falhas, acao):
        self.btn_pb.configure(state="disabled")
        self.btn_cb.configure(state="disabled")
        if df is not None and "codigoItem" in df.columns:
            self.lista_catmats = df["codigoItem"].dropna().astype(int).tolist()
            n = len(self.lista_catmats)
            msg = f"✅ {n} CATMATs encontrados"
            if falhas: msg += f" · ⚠ {len(falhas)} PDMs com falha"
            self.set_status_explorador(msg)
            self.btn_exp_cat.configure(state="normal")
            self.btn_ini_exp.configure(state="normal", text=f"▶  Iniciar Extração com {n} CATMATs Encontrados")
            if falhas:
                messagebox.showwarning("PDMs com falha",
                    f"Falha persistente em:\n{', '.join(map(str,falhas))}")
            if acao == "extrair": self._iniciar_exp()
        else:
            self.lista_catmats = []
            if not cancelar_busca_catmat:
                self.set_status_explorador("Nenhum CATMAT encontrado.")
            self.btn_exp_cat.configure(state="disabled")
            self.btn_ini_exp.configure(state="disabled")

    def _exp_catmats(self):
        if not self.lista_catmats:
            messagebox.showerror("Vazio","Nenhum CATMAT."); return
        p = filedialog.asksaveasfilename(defaultextension=".csv",
                                         initialfile="CATMATs_descobertos.csv",
                                         filetypes=[("CSV","*.csv")])
        if p:
            pd.DataFrame(self.lista_catmats, columns=["codigoItemCatalogo"])\
              .to_csv(p, index=False, sep=";", encoding="utf-8-sig")
            messagebox.showinfo("Exportado", f"{len(self.lista_catmats)} CATMATs:\n{p}")

    def _pausar_busca(self):
        if pausar_busca_catmat.is_set():
            pausar_busca_catmat.clear()
            self.btn_pb.configure(text="▶  Retomar Busca")
            self.set_status_explorador("Busca pausada.")
        else:
            pausar_busca_catmat.set()
            self.btn_pb.configure(text="⏸  Pausar Busca")
            self.set_status_explorador("Retomando…")

    def _toggle_pasta_por_classe(self):
        """Mostra/oculta o campo de pasta quando o checkbox é marcado."""
        if self.var_arquivo_por_classe.get():
            self.frame_pasta_classes.pack(fill="x", padx=0, pady=(6,0))
        else:
            self.frame_pasta_classes.pack_forget()
            self.var_pasta_classes.set("")

    def _escolher_pasta_classes(self):
        p = filedialog.askdirectory(title="Escolha a pasta de destino dos arquivos por classe")
        if p:
            self.var_pasta_classes.set(p)

    def _cancelar_busca(self):
        global cancelar_busca_catmat
        cancelar_busca_catmat = True

    def _iniciar_exp(self):
        if not self.lista_catmats:
            messagebox.showerror("Vazio","Nenhum CATMAT disponível."); return
        d_i, d_f, err = validar_e_obter_datas(self.var_ini2.get(), self.var_fim2.get())
        if err: messagebox.showerror("Data inválida", err); return
        self._log(f"🔎 {len(self.lista_catmats)} CATMATs via explorador.", "info")
        por_classe = self.var_arquivo_por_classe.get()
        # Usa o mapa já construído em _on_busca_e_extrai (se existir)
        mapa = getattr(self, "_catmats_por_classe", {}) if por_classe else {}
        if por_classe and not mapa:
            messagebox.showwarning("Aviso",
                "Use Buscar e Extrair para gerar arquivos por classe.")
            return
        self._iniciar_processo(self.lista_catmats, self.var_fmt.get(), d_i, d_f,
                               catmats_por_classe=mapa, tipo=TIPO_CATMAT)
        self.after(100, lambda: self.tabs.set("  Extração por CATMAT  "))

    def _buscar_e_extrair_classes(self):
        """Fluxo automatizado: para cada classe, faz PDMs → CATMATs → Extração → Salva."""
        entrada = self.var_classe.get().strip()
        if not entrada:
            messagebox.showerror("Campo vazio", "Informe ao menos um código de Classe."); return

        partes = [p.strip() for p in entrada.split(";") if p.strip()]
        invalidas = [p for p in partes if not p.isdigit()]
        if invalidas:
            messagebox.showerror("Código inválido",
                "Valores nao numericos: " + ", ".join(invalidas)); return

        por_classe = self.var_arquivo_por_classe.get()
        pasta_dest = self.var_pasta_classes.get().strip() if por_classe else ""

        if por_classe and len(partes) > 1 and not pasta_dest:
            messagebox.showerror("Pasta obrigatória",
                "Selecione uma pasta de destino para salvar os arquivos por classe."); return

        d_i, d_f, err = validar_e_obter_datas(self.var_ini2.get(), self.var_fim2.get())
        if err: messagebox.showerror("Data inválida", err); return

        fmt = self.var_fmt.get()
        self._salvar_corr = self.var_salvar_corr.get()   # capturado na thread principal
        self._pasta_corr  = self.var_pasta.get()
        self._pasta_classes_destino = pasta_dest
        self.processing   = True
        self.total_baixados = 0
        self.count_corrigidas = 0
        self.count_reparadas  = 0
        self.count_vazios = 0
        pausar_extracao.set()
        pausar_busca_catmat.set()  # necessário para o fluxo automatizado
        global cancelar_busca_catmat
        cancelar_busca_catmat = False   # zera resíduo de cancelamento anterior

        self.log.configure(state="normal"); self.log.delete("1.0","end")
        self.log.configure(state="disabled")
        for k, v in [("k_proc","0"),("k_reg","0"),("k_corr","0 · 0"),("k_vaz","0")]:
            self._stat(k, v)
        self.progress.set(0); self.lbl_pct.configure(text="0%")
        self.set_status("Status: Iniciando…")
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal", fg_color=C_RED,
                                   hover_color="#992B1E", text_color=C_SURFACE)
        self.btn_pause.configure(state="normal", text="⏸  Pausar")
        self.btn_log.configure(state="disabled")

        self._log("💾 Formato: " + ("CSV" if fmt == "csv" else "Excel"), "info")
        if d_i or d_f:
            def fd(s):
                p = s.split("-"); return p[2]+"-"+p[1]+"-"+p[0]
            txt = "📅 Filtro de datas:"
            if d_i: txt += "  Início: " + fd(d_i)
            if d_f: txt += "  Fim: "    + fd(d_f)
            self._log(txt, "date")
        self._log("📋 " + str(len(partes)) + " classe(s): " + " | ".join(partes) + "\n", "info")
        tipo_busca = TIPO_PDM if self.var_extrair_por_pdm.get() else TIPO_CATMAT
        self._tipo_busca = tipo_busca
        self._log("🎯 Tipo de busca: " + ROTULO_TIPO[tipo_busca] +
                  "  (tipo=" + tipo_busca + ")", "info")
        if pasta_dest:
            self._log("📂 Destino: " + pasta_dest, "info")

        self.after(100, lambda: self.tabs.set("  Extração por CATMAT  "))
        threading.Thread(
            target=self._fluxo_classes_thread,
            args=(partes, pasta_dest, fmt, d_i, d_f, tipo_busca),
            daemon=True
        ).start()

    def _fluxo_classes_thread(self, classes_lista, pasta_dest, fmt, d_ini, d_fim,
                              tipo_busca=TIPO_CATMAT):
        """
        Thread principal do fluxo por classe.
        tipo_busca = TIPO_CATMAT → Classe → PDMs → CATMATs → Registros de Preços
        tipo_busca = TIPO_PDM    → Classe → PDMs → Registros de Preços (direto)
        Retry em todos os níveis: classes, PDMs e CATMATs com erro.
        """
        ext          = "csv" if fmt == "csv" else "xlsx"
        total_classes_orig = len(classes_lista)
        salvar_corr  = getattr(self, "_salvar_corr", False)
        pasta_corr   = getattr(self, "_pasta_corr", "")
        total_catmats_acum = 0
        classes_com_falha  = []   # classes que não retornaram PDMs

        # ── Helper: extrai os Registros de Preços de uma lista de códigos ─────
        # (CATMATs quando tipo_busca=TIPO_CATMAT, PDMs quando TIPO_PDM)
        def _extrair_codigos(classe, idx_c, total_c, codigos_lista):

            # ── 3. Extração dos Registros de Preços ───────────────────────────
            nome_arq = "classe_" + classe + "." + ext
            caminho  = os.path.join(pasta_dest, nome_arq) if pasta_dest else nome_arq
            writer   = CSVChunkWriter(caminho) if fmt == "csv" else ExcelChunkWriter(caminho)

            reg_baixados    = {}
            reg_esperados   = {}
            pag_corrompidas = {}
            total_baixados_classe = 0
            vazios_classe         = 0
            catmats_com_erro      = []
            total_cat  = len(codigos_lista)
            writer_lock = threading.Lock()
            state_lock  = threading.Lock()
            comp_count  = [0]

            def _processar_resultado(codigo, dfs_e_meta, tipo, reg_esp, pag_corr):
                nonlocal total_baixados_classe, vazios_classe, total_catmats_acum
                if tipo == "conexao":
                    pass  # já tratado dentro de _fetch_catmat_registros via pausa automática
                elif tipo == "erro":
                    with state_lock:
                        catmats_com_erro.append(codigo)
                    etxt = "⚠️  CATMAT " + str(codigo) + ": erro na API, será retentado."
                    self._ui(lambda t=etxt: self._log(t, "warn"))
                elif tipo == "vazio":
                    vtxt = "ℹ️  " + str(codigo) + ": 0 registros."
                    self._ui(lambda t=vtxt: self._log(t, "info"))
                    with state_lock:
                        vazios_classe += 1
                        self.count_vazios += 1
                        v = self.count_vazios
                    self._ui(lambda vv=v: self._stat("k_vaz", vv))
                else:
                    baixados = sum(len(df) for df, _, _ in dfs_e_meta)
                    with writer_lock:
                        for df_proc, _, _ in dfs_e_meta:
                            writer.write_dataframe(df_proc)
                    with state_lock:
                        total_baixados_classe += baixados
                        reg_baixados[codigo]   = baixados
                        reg_esperados[codigo]  = reg_esp
                        self.total_baixados   += baixados
                        for k, v in pag_corr.items():
                            pag_corrompidas.setdefault(k, []).extend(v)
                        n_corr = sum(len(v) for v in pag_corr.values())
                        self.count_corrigidas += n_corr
                        reg = self.total_baixados
                    self._ui(lambda r=reg: self._stat("k_reg", f"{r:,}".replace(",",".")))
                    n_rep = sum(1 for _, mk, _ in dfs_e_meta if mk == "reparada")
                    if n_rep:
                        with state_lock:
                            self.count_reparadas += n_rep
                    for _, marca, pag in dfs_e_meta:
                        if marca == "perda":
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"❌  Cód {cod} Pág {p}: registros perdidos.", "err"))
                        elif marca == "reparada":
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"⚠️  Cód {cod} Pág {p}: reparada (íntegra).", "warn"))
                    if pag_corr or n_rep:
                        self._ui(self._atualizar_stat_paginas)
                with state_lock:
                    comp_count[0] += 1
                    total_catmats_acum += 1
                    comp = comp_count[0]
                    tca  = total_catmats_acum
                self._ui(lambda v=tca: self._stat("k_proc", v))
                pct = (idx_c - 1 + comp / total_cat) / total_c
                self._ui(lambda c=classe, i=comp, t=total_cat, p=pct:
                    (self.set_status("Status: Classe " + c + " — " + str(i) + "/" + str(t)),
                     self.progress.set(p),
                     self.lbl_pct.configure(text=str(int(p*100)) + "%")))
                return tipo

            # Extração sequencial (max_workers=1 — paralelismo sobrecarrega a API)
            with ThreadPoolExecutor(max_workers=1) as executor:
                futures = {
                    executor.submit(_fetch_catmat_registros, cod,
                                    d_ini, d_fim, salvar_corr, pasta_corr,
                                    self._pausar_por_conexao, tipo_busca,
                                    lambda: not self.processing): cod
                    for cod in codigos_lista
                }
                for future in as_completed(futures):
                    pausar_extracao.wait()
                    if not self.processing:
                        executor.shutdown(wait=False, cancel_futures=True); break
                    cod, dfs_m, tipo, reg_e, pag_c = future.result()
                    res = _processar_resultado(cod, dfs_m, tipo, reg_e, pag_c)
                    if res == "conexao": break

            # Retry sequencial dos códigos com erro
            rotulo = ROTULO_TIPO.get(tipo_busca, "código")
            for espera in [15, 30]:
                if not catmats_com_erro or not self.processing: break
                n_err = len(catmats_com_erro)
                self._ui(lambda n=n_err, e=espera, rt=rotulo:
                    self._log("♻️  Retry " + str(n) + " " + rt + "(s) com erro (aguardando " +
                              str(e) + "s)…", "warn"))
                time.sleep(espera)
                retry_list = list(catmats_com_erro); catmats_com_erro.clear()
                with ThreadPoolExecutor(max_workers=1) as executor:
                    futures = {
                        executor.submit(_fetch_catmat_registros, cod,
                                        d_ini, d_fim, salvar_corr, pasta_corr,
                                        self._pausar_por_conexao, tipo_busca,
                                        lambda: not self.processing): cod
                        for cod in retry_list
                    }
                    for future in as_completed(futures):
                        if not self.processing: break
                        cod, dfs_m, tipo, reg_e, pag_c = future.result()
                        _processar_resultado(cod, dfs_m, tipo, reg_e, pag_c)

            if catmats_com_erro:
                n_def = len(catmats_com_erro)
                self._ui(lambda n=n_def, rt=rotulo:
                    self._log("❌  " + str(n) + " " + rt + "(s) sem resposta após 3 tentativas.", "err"))

            if not self.processing: return True

            # ── 4. Finalizar arquivo desta classe ─────────────────────────────
            parts = writer.finalize()
            arqs  = ", ".join(os.path.basename(p) for p in parts) if parts else "(sem dados)"
            self._ui(lambda c=classe, a=arqs, n=total_baixados_classe:
                self._log("📁  Classe " + c + " — " + str(n) + " registros → " + a, "info"))

            # ── 5. Relatório de integridade desta classe ──────────────────────
            rel_nome    = "Relatorio_Integridade_" + classe + ".xlsx"
            rel_caminho = os.path.join(pasta_dest, rel_nome) if pasta_dest else rel_nome
            try:
                wb = Workbook(); ws = wb.active; ws.title = "Integridade_" + classe
                ws.append([tipo_busca,"esperados","baixados","paginas","status"])
                for c in codigos_lista:
                    bx = int(reg_baixados.get(c, 0))
                    ex = int(reg_esperados.get(c, 0))
                    pg = pag_corrompidas.get(c, [])
                    d  = abs(ex - bx)
                    st = ("OK" if d == 0 else
                          "OK (divergencia: " + str(bx) + "/" + str(ex) + ")" if d <= 2 else
                          "Inconsistencia Grave (" + str(bx) + "/" + str(ex) + ")")
                    ws.append([c, ex, bx, ", ".join(map(str, pg)), st])
                if catmats_com_erro:
                    ws.append([])
                    ws.append(["--- " + rotulo + "s sem resposta apos 3 tentativas ---"])
                    for c in catmats_com_erro:
                        ws.append([c, 0, 0, "", "ERRO_API_PERSISTENTE"])
                wb.save(rel_caminho)
                self._ui(lambda r=rel_nome:
                    self._log("📊  Relatório: " + r, "info"))
            except Exception as e:
                etxt = "⚠️ Relatório classe " + classe + " não salvo: " + str(e)
                self._ui(lambda t=etxt: self._log(t, "warn"))
            return True

        # ── Helper: processa uma classe completa ──────────────────────────────
        def _processar_classe(classe, idx_c, total_c):
            ic, tc = idx_c, total_c
            sep = "─" * 50
            hdr = sep + "\n📦  CLASSE " + classe + "  (" + str(ic) + "/" + str(tc) + ")\n" + sep
            self._ui(lambda h=hdr: self._log("\n" + h, "date"))
            self._ui(lambda c=classe, i=ic, t=tc:
                self.set_status("Status: Classe " + c + " (" + str(i) + "/" + str(t) + ") — PDMs…"))

            # ── 1. PDMs ───────────────────────────────────────────────────────
            resultado = buscar_pdms_por_classe(int(classe), URL_BASE, TIMEOUT)
            if resultado is None:
                return False   # falhou — será retentada
            df_pdms, _ = resultado
            pdms_lista  = df_pdms["codigoPdm"].astype(int).tolist()
            n_pdms = len(pdms_lista)
            self._ui(lambda c=classe, n=n_pdms:
                self._log("✅  Classe " + c + ": " + str(n) + " PDMs.", "ok"))

            # ── 2. CATMATs ────────────────────────────────────────────────────
            # No modo tipo=codigoPdm a expansão PDM → CATMAT é desnecessária:
            # a própria Pesquisa de Preço devolve todos os itens do PDM.
            if tipo_busca == TIPO_PDM:
                self._ui(lambda c=classe, n=n_pdms:
                    self._log("⏩  Classe " + c + ": extração direta de " + str(n) +
                              " PDMs (sem expandir CATMATs).", "info"))
                return _extrair_codigos(classe, idx_c, total_c, pdms_lista)

            self._ui(lambda c=classe, n=n_pdms:
                self.set_status("Status: Classe " + c + " — CATMATs (" + str(n) + " PDMs)…"))

            _log_pdm = lambda m: self._log("  🔍 " + m, "info")
            df_cat1, erros1 = buscar_catmats_por_pdm(pdms_lista, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
            df_cat2 = None; erros2 = []
            if erros1 and not cancelar_busca_catmat:
                n_e1 = len(erros1)
                self._ui(lambda n=n_e1:
                    self._log("♻️  2ª tentativa (10s) para " + str(n) + " PDMs com erro…", "warn"))
                time.sleep(3)
                df_cat2, erros2 = buscar_catmats_por_pdm(erros1, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
            # 3ª tentativa para PDMs que ainda falharam
            df_cat3 = None; erros3 = []
            if erros2 and not cancelar_busca_catmat:
                n_e2 = len(erros2)
                self._ui(lambda n=n_e2:
                    self._log("♻️  3ª tentativa (20s) para " + str(n) + " PDMs com erro…", "warn"))
                time.sleep(8)
                df_cat3, erros3 = buscar_catmats_por_pdm(erros2, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
                if erros3:
                    falha_pdm = ", ".join(map(str, erros3))
                    self._ui(lambda t=falha_pdm:
                        self._log("❌  PDMs sem resposta após 3 tentativas: " + t, "err"))

            dfs_c = [d for d in [df_cat1, df_cat2, df_cat3] if d is not None and not d.empty]
            df_catmats = pd.concat(dfs_c, ignore_index=True) if dfs_c else None
            if df_catmats is None or "codigoItem" not in df_catmats.columns:
                self._ui(lambda c=classe:
                    self._log("⚠️  Classe " + c + ": nenhum CATMAT. Pulando.", "warn"))
                return True   # PDMs foram encontrados mas sem CATMATs — não é falha de PDMs

            catmats_lista = df_catmats["codigoItem"].dropna().astype(int).tolist()
            n_cat = len(catmats_lista)
            self._ui(lambda c=classe, n=n_cat:
                self._log("✅  Classe " + c + ": " + str(n) + " CATMATs.", "ok"))

            return _extrair_codigos(classe, idx_c, total_c, catmats_lista)

        # ── Loop principal por classe ─────────────────────────────────────────
        for idx_classe, classe in enumerate(classes_lista, 1):
            if not self.processing: break
            ok = _processar_classe(classe, idx_classe, total_classes_orig)
            if not ok:
                self._ui(lambda c=classe:
                    self._log("⚠️  Classe " + c + ": sem PDMs na 1ª tentativa. Fila de retry.", "warn"))
                classes_com_falha.append(classe)

        # ── Retry classes sem PDMs (10s → 20s) ───────────────────────────────
        for num_tent, espera in enumerate([3, 8], 2):
            if not classes_com_falha or not self.processing: break
            n_f = len(classes_com_falha)
            self._ui(lambda n=n_f, e=espera, t=num_tent:
                self._log("\n♻️  " + str(n) + " classe(s) sem PDMs — tentativa " +
                          str(t) + "/3 (aguardando " + str(e) + "s)…", "warn"))
            time.sleep(espera)
            ainda_falha = []
            for idx, classe in enumerate(classes_com_falha, 1):
                if not self.processing: break
                ok = _processar_classe(classe, idx, len(classes_com_falha))
                if not ok:
                    ainda_falha.append(classe)
            classes_com_falha = ainda_falha

        if classes_com_falha:
            falhas = ", ".join(classes_com_falha)
            self._ui(lambda t=falhas:
                self._log("❌  Classes sem PDMs após 3 tentativas: " + t, "err"))

        # ── Todas as classes processadas ──────────────────────────────────────
        self._ui(self._finalizar_fluxo_classes)

    def _finalizar_fluxo_classes(self):
        """Chamado ao término de todas as classes."""
        foi_cancelado = not self.processing
        self.processing = False
        self.progress.set(1.0); self.lbl_pct.configure(text="100%")
        self.set_status("Status: Concluído!" if not foi_cancelado else "Status: Cancelado")
        self._log(
            "\n🎉 Todas as classes processadas com sucesso!" if not foi_cancelado
            else "\n🛑 Extração cancelada.", "info")

        self.btn_start.configure(state="normal")
        self.btn_cancel.configure(state="disabled",
                                   fg_color="#E4E7EF", hover_color=C_BORDER,
                                   text_color=C_TEXT)
        self.btn_pause.configure(state="disabled", text="⏸  Pausar")
        self.btn_log.configure(state="normal")

        pasta = getattr(self, "_pasta_classes_destino", "").strip()
        if pasta:
            messagebox.showinfo("Concluído",
                "Extração finalizada!\nTodos os arquivos foram salvos em:\n" + pasta)
        else:
            messagebox.showinfo("Concluído", "Extração finalizada!")

        # ── MOTOR DE EXTRAÇÃO ─────────────────────────────────────────────────────
    def _iniciar_processo(self, codigos, fmt, d_ini, d_fim, catmats_por_classe=None,
                          tipo=TIPO_CATMAT, por_classe=None, pasta_destino=None):
        if not codigos: return
        self.processing            = True
        self.codigos_lista         = codigos
        self._data_inicio          = d_ini
        self._data_fim             = d_fim
        self._fmt                  = fmt
        self._tipo_busca           = tipo
        # Tk não é thread-safe: capturar aqui, na thread principal
        self._salvar_corr          = self.var_salvar_corr.get()
        self._pasta_corr           = self.var_pasta.get()
        self._catmats_por_classe_ativo = catmats_por_classe or {}
        # Um arquivo por classe. Quando não há mapa prévio de códigos→classe,
        # a classe é lida do campo codigoClasse de cada registro.
        self._modo_por_classe = (bool(self._catmats_por_classe_ativo)
                                 if por_classe is None else bool(por_classe))
        # Pasta escolhida pelo usuário para salvar os arquivos por classe.
        # pasta_destino explícito evita que a pasta da aba 2 vaze para a aba 1.
        if pasta_destino is not None:
            self._pasta_classes_destino = pasta_destino
        else:
            self._pasta_classes_destino = (self.var_pasta_classes.get().strip()
                                           if hasattr(self, "var_pasta_classes") else "")
        self.paginas_corrompidas   = {}
        self.registros_esperados   = {}
        self.registros_baixados    = {}
        self.total_baixados        = 0
        self.count_corrigidas      = 0
        self.count_reparadas       = 0
        self.count_vazios          = 0
        pausar_extracao.set()

        # Se há arquivo por classe, usamos um writer por classe (criados sob demanda)
        # Senão, writer único
        if self._modo_por_classe:
            self.writer = None  # será None; usamos self._writers_por_classe
            self._writers_por_classe = {}  # classe → writer
        else:
            self.writer = CSVChunkWriter("dados_completos_extraidos.csv") \
                          if fmt == "csv" else \
                          ExcelChunkWriter("dados_completos_extraidos.xlsx")
            self._writers_por_classe = {}

        self.log.configure(state="normal"); self.log.delete("1.0","end")
        self.log.configure(state="disabled")
        self._log(f"💾 Formato: {'CSV' if fmt == 'csv' else 'Excel'}", "info")
        if d_ini or d_fim:
            def fd(s): p = s.split("-"); return f"{p[2]}-{p[1]}-{p[0]}"
            txt = "📅 Filtro de datas:"
            if d_ini: txt += f"  Início: {fd(d_ini)}"
            if d_fim:  txt += f"  Fim: {fd(d_fim)}"
            self._log(txt, "date")
        self._log(f"🎯 Tipo de busca: {ROTULO_TIPO.get(tipo, tipo)}  (tipo={tipo})", "info")
        self._log(f"🔎 {len(codigos)} códigos carregados.\n", "info")

        for k, v in [("k_proc",f"0 / {len(codigos)}"),
                     ("k_reg","0"),("k_corr","0 · 0"),("k_vaz","0")]:
            self._stat(k, v)
        self.progress.set(0); self.lbl_pct.configure(text="0%")
        self.set_status("Status: Processando…")
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal", fg_color=C_RED,
                                   hover_color="#992B1E", text_color=C_SURFACE)
        self.btn_pause.configure(state="normal", text="⏸  Pausar")
        self.btn_log.configure(state="disabled")
        # Lança extração em thread separada — UI continua responsiva
        self._extracao_thread_obj = threading.Thread(
            target=self._extracao_thread, daemon=True)
        self._extracao_thread_obj.start()

    # ─────────────────────────────────────────────────────────────────────────
    # MOTOR DE EXTRAÇÃO — roda 100% em thread separada
    # Comunicação com UI exclusivamente via self.after(0, callback)
    # ─────────────────────────────────────────────────────────────────────────

    def _get_writer_para(self, codigo: int, classe: str = None):
        if not self._modo_por_classe:
            return self.writer
        # Mapa explícito (explorador ou coluna "classe" do arquivo) tem prioridade;
        # sem ele, vale a classe lida do próprio registro.
        classe_do_cod = None
        for cl, cats in self._catmats_por_classe_ativo.items():
            if codigo in cats:
                classe_do_cod = cl; break
        if classe_do_cod is None:
            classe_do_cod = classe or "sem_classe"
        classe_do_cod = re.sub(r'[\\/:*?"<>|]', "_", str(classe_do_cod)).strip() or "sem_classe"
        if classe_do_cod not in self._writers_por_classe:
            ext   = "csv" if self._fmt == "csv" else "xlsx"
            # Salva direto na pasta escolhida pelo usuário (se informada)
            pasta = getattr(self, "_pasta_classes_destino", "").strip()
            nome_arq = f"classe_{classe_do_cod}.{ext}"
            caminho  = os.path.join(pasta, nome_arq) if pasta else nome_arq
            self._writers_por_classe[classe_do_cod] = (
                CSVChunkWriter(caminho) if self._fmt == "csv"
                else ExcelChunkWriter(caminho))
        return self._writers_por_classe[classe_do_cod]

    def _ui(self, fn):
        """Agenda fn() na thread principal de forma segura."""
        self.after(0, fn)

    def _extracao_thread(self):
        """Thread de extração paralela — 4 CATMATs simultâneos."""
        codigos     = self.codigos_lista
        total       = len(codigos)
        salvar_corr = getattr(self, "_salvar_corr", False)
        pasta_corr  = getattr(self, "_pasta_corr", "")
        d_ini       = self._data_inicio
        d_fim       = self._data_fim
        tipo_busca  = getattr(self, "_tipo_busca", TIPO_CATMAT)
        writer_locks: dict = {}   # id(writer) → Lock
        state_lock  = threading.Lock()
        comp_count  = [0]

        def _wlock(codigo, classe=None):
            w = self._get_writer_para(codigo, classe)
            k = id(w)
            if k not in writer_locks:
                writer_locks[k] = threading.Lock()
            return w, writer_locks[k]

        def _escrever(codigo, df_proc):
            """Roteia o DataFrame para o writer certo, quebrando por classe."""
            if (not self._modo_por_classe or self._catmats_por_classe_ativo
                    or "codigoClasse" not in df_proc.columns):
                w, lk = _wlock(codigo)
                with lk: w.write_dataframe(df_proc)
                return
            # Classe lida do próprio registro: uma partição por classe encontrada
            chaves = (df_proc["codigoClasse"].astype(str).str.strip()
                      .replace({"": "sem_classe", "nan": "sem_classe",
                                "None": "sem_classe"}))
            for classe, parte in df_proc.groupby(chaves, sort=False):
                w, lk = _wlock(codigo, str(classe))
                with lk: w.write_dataframe(parte)

        def _flush_periodico():
            """Descarrega em disco o que já foi montado (throttled no writer)."""
            alvos = (list(self._writers_por_classe.values())
                     if self._modo_por_classe else
                     ([self.writer] if self.writer else []))
            for w in alvos:
                lk = writer_locks.get(id(w))
                if lk is None:
                    continue
                with lk:
                    w.flush()

        with ThreadPoolExecutor(max_workers=1) as executor:
            futures = {
                executor.submit(_fetch_catmat_registros, cod,
                                d_ini, d_fim, salvar_corr, pasta_corr,
                                self._pausar_por_conexao, tipo_busca,
                                lambda: not self.processing): cod
                for cod in codigos
            }
            for future in as_completed(futures):
                pausar_extracao.wait()
                if not self.processing:
                    executor.shutdown(wait=False, cancel_futures=True); break

                codigo, dfs_e_meta, tipo, reg_esp, pag_corr = future.result()

                with state_lock:
                    comp_count[0] += 1
                    comp = comp_count[0]

                if tipo == "conexao":
                    pass  # já tratado dentro de _fetch_catmat_registros via pausa automática

                elif tipo in ("erro", "vazio"):
                    txt = (f"ℹ️  {codigo}: sem registro (erro API)." if tipo == "erro"
                           else f"ℹ️  {codigo}: 0 registros.")
                    with state_lock:
                        self.count_vazios += 1
                        v = self.count_vazios
                        self.registros_baixados[codigo]  = 0
                        self.registros_esperados[codigo] = reg_esp
                    self._ui(lambda t=txt: self._log(t, "info"))
                    self._ui(lambda vv=v: self._stat("k_vaz", vv))

                else:  # "ok"
                    baixados = sum(len(df) for df, _, _ in dfs_e_meta)
                    for df_proc, _, _ in dfs_e_meta:
                        _escrever(codigo, df_proc)
                    _flush_periodico()
                    with state_lock:
                        self.total_baixados             += baixados
                        self.registros_baixados[codigo]  = baixados
                        self.registros_esperados[codigo] = reg_esp
                        reg = self.total_baixados
                        for k, v in pag_corr.items():
                            self.paginas_corrompidas.setdefault(k, []).extend(v)
                        n_c = sum(len(v) for v in pag_corr.values())
                        self.count_corrigidas += n_c
                    self._ui(lambda r=reg: self._stat("k_reg", f"{r:,}".replace(",",".")))
                    n_rep = sum(1 for _, mk, _ in dfs_e_meta if mk == "reparada")
                    if n_rep:
                        with state_lock:
                            self.count_reparadas += n_rep
                    for _, marca, pag in dfs_e_meta:
                        if marca == "perda":
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"❌  Cód {cod} Pág {p}: registros perdidos.", "err"))
                        elif marca == "reparada":
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"⚠️  Cód {cod} Pág {p}: reparada (íntegra).", "warn"))
                        else:
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"✅  Cód {cod} Pág {p}: OK.", "ok"))
                    if pag_corr or n_rep:
                        self._ui(self._atualizar_stat_paginas)

                pct = comp / total
                self._ui(lambda p=pct, c=comp, t=total: (
                    self.progress.set(p),
                    self.lbl_pct.configure(text=f"{int(p*100)}%"),
                    self.set_status(f"Status: Processando... ({c}/{t})"),
                    self._stat("k_proc", f"{c} / {t}")
                ))

        self._ui(self._finalizar)

    def _finalizar(self):
        """Chamado na thread principal via after() ao término da extração."""
        foi_cancelado = not self.processing
        self.processing = False
        self.progress.set(1.0); self.lbl_pct.configure(text="100%")
        self.set_status("Status: Concluído!" if not foi_cancelado else "Status: Cancelado")
        self._log("\n🎉 Extração concluída!" if not foi_cancelado
                  else "\n🛑 Extração cancelada.", "info")

        # Finalizar writers
        parts = []
        if self._modo_por_classe and self._writers_por_classe:
            for classe, w in self._writers_por_classe.items():
                p = w.finalize(); parts.extend(p)
                if p: self._log(f"📂 Classe {classe}: {', '.join(p)}", "info")
        elif self.writer:
            parts = self.writer.finalize()

        if parts:
            self._log(f"💾 Arquivos gerados: {', '.join(parts)}", "info")

        # Relatório de integridade
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Relatorio Integridade"
            ws.append([getattr(self, "_tipo_busca", TIPO_CATMAT),
                       "esperados","baixados","paginas","status"])
            for c in self.codigos_lista:
                bx = int(self.registros_baixados.get(c,0))
                ex = int(self.registros_esperados.get(c,0))
                pg = self.paginas_corrompidas.get(c,[])
                d  = abs(ex-bx)
                st = ("OK" if d==0 else
                      f"OK (divergencia: {bx}/{ex})" if d<=2 else
                      f"Inconsistencia Grave ({bx}/{ex})")
                ws.append([c,ex,bx,", ".join(map(str,pg)),st])
            pasta_rel = getattr(self, "_pasta_classes_destino", "").strip()
            cam_rel = (os.path.join(pasta_rel, "Relatorio_Integridade.xlsx")
                       if pasta_rel else "Relatorio_Integridade.xlsx")
            wb.save(cam_rel)
            self._log(f"📊 {cam_rel} gerado.", "info")
        except Exception as e:
            self._log(f"⚠️ Relatório não salvo: {e}", "warn")

        self.btn_start.configure(state="normal")
        self.btn_cancel.configure(state="disabled",
                                   fg_color="#E4E7EF", hover_color=C_BORDER,
                                   text_color=C_TEXT)
        self.btn_pause.configure(state="disabled", text="⏸  Pausar")
        self.btn_log.configure(state="normal")

        if not parts:
            messagebox.showinfo("Sem dados","Nenhum dado válido baixado.")
            return

        n_classes = len(self._writers_por_classe) if self._modo_por_classe else 0
        ext = os.path.splitext(parts[0])[1]

        resumo_linhas = [
            "Processo Concluido!",
            chr(8212)*40,
            f"Codigos Processados:     {len(self.codigos_lista)}",
            f"Registros Consolidados:  {self.total_baixados:,}",
            f"Paginas Reparadas:       {self.count_reparadas}",
            f"Paginas com Perda:       {self.count_corrigidas}",
            f"Codigos sem Registros:   {self.count_vazios}",
        ]
        if n_classes >= 1:
            resumo_linhas.append(f"Arquivos por classe:     {n_classes}")
        messagebox.showinfo("Resumo", "\n".join(resumo_linhas))

        # Modo por classe: mesmo com uma única classe os arquivos já têm nome
        # próprio (classe_XXXX) e destino definido — não faz sentido pedir
        # "salvar como" para eles.
        if n_classes >= 1:
            pasta_dest = getattr(self, "_pasta_classes_destino", "").strip()
            nomes = "\n".join(os.path.basename(p) for p in parts)
            if pasta_dest:
                # Os writers já gravaram direto na pasta ao longo da execução;
                # o relatório de integridade também. Nada a copiar nem a perguntar.
                self._log(f"📁 {len(parts)} arquivo(s) em: {pasta_dest}", "info")
                messagebox.showinfo("Concluído",
                    f"{len(parts)} arquivo(s) salvos em:\n{pasta_dest}\n\n{nomes}")
            else:
                # Sem pasta definida: pede agora e copia
                pasta_dest = filedialog.askdirectory(
                    title=f"Escolha a pasta para salvar os {len(parts)} arquivo(s)")
                if pasta_dest:
                    for arq in parts:
                        shutil.copy(arq, os.path.join(pasta_dest, os.path.basename(arq)))
                    try:
                        shutil.copy("Relatorio_Integridade.xlsx",
                                    os.path.join(pasta_dest, "Relatorio_Integridade.xlsx"))
                    except Exception:
                        pass
                    self._log(f"📁 {len(parts)} arquivo(s) salvos em: {pasta_dest}", "info")
                    messagebox.showinfo("Concluído",
                        f"{len(parts)} arquivo(s) salvos em:\n{pasta_dest}\n\n{nomes}")
                else:
                    messagebox.showwarning("Atenção",
                        f"Nenhuma pasta escolhida. Arquivos na pasta do programa:\n{nomes}")
        else:
            ultimo = parts[-1]
            tipos  = [("Excel","*.xlsx")] if ext == ".xlsx" else [("CSV","*.csv")]
            dest   = filedialog.asksaveasfilename(
                        defaultextension=ext,
                        initialfile=os.path.basename(ultimo),
                        filetypes=tipos)
            if dest:
                if not dest.lower().endswith(ext): dest += ext
                shutil.copy(ultimo, dest)
                messagebox.showinfo("Salvo",
                    f"Dados salvos em:\n{dest}\n\nRelatorio de integridade na pasta do programa.")
            else:
                messagebox.showwarning("Atencao", f"Arquivo permanece em:\n{ultimo}")

    # ─────────────────────────────────────────────────────────────────────────
    # ABA 3 — CONSOLIDAÇÃO DW + DA (remoção de duplicatas)
    # A regra de negócio vive em consolidar_dw_da.py; aqui só há interface.
    # ─────────────────────────────────────────────────────────────────────────

    def _log_cons(self, msg: str, tag: str = ""):
        self.log_cons.configure(state="normal")
        self.log_cons.insert("end", msg + "\n", tag)
        self.log_cons.see("end")
        self.log_cons.configure(state="disabled")

    def _stat_cons(self, key, val):
        self._stats_cons[key].configure(text=str(val))

    @staticmethod
    def _fmt_br(n) -> str:
        """1234567 -> 1.234.567"""
        return f"{n:,}".replace(",", ".")

    # ── entradas ─────────────────────────────────────────────────────────────
    _ROTULO_ORIGEM = {"Detectar": "auto", "DW": "dw", "DA": "da"}

    def _cons_add(self, caminhos):
        origem = self._ROTULO_ORIGEM.get(self.var_cons_origem.get(), "auto")
        ja = {e["caminho"] for e in self._cons_entradas}
        novos = 0
        for c in caminhos:
            if c and c not in ja:
                self._cons_entradas.append({"caminho": c, "origem": origem})
                ja.add(c); novos += 1
        if novos:
            self._cons_atualizar_tree()

    def _cons_add_pasta(self):
        p = filedialog.askdirectory(
            title="Pasta com os CSVs/XLSX do DW e/ou do DA")
        if p:
            self._cons_add([p])

    def _cons_add_arquivos(self):
        arqs = filedialog.askopenfilenames(
            title="Arquivos do DW e/ou do DA",
            filetypes=[("CSV/Excel","*.csv *.xlsx *.xlsm"),("Todos","*.*")])
        if arqs:
            self._cons_add(list(arqs))

    def _cons_atualizar_tree(self):
        self.tree_cons.delete(*self.tree_cons.get_children())
        rotulo = {v: k for k, v in self._ROTULO_ORIGEM.items()}
        for i, e in enumerate(self._cons_entradas):
            self.tree_cons.insert("", "end", iid=str(i),
                                  values=(rotulo.get(e["origem"], "Detectar"),
                                          e["caminho"]))

    def _cons_remover(self):
        sel = self.tree_cons.selection()
        if not sel:
            messagebox.showinfo("Nada selecionado",
                                "Selecione na lista o que deseja remover.")
            return
        for i in sorted((int(s) for s in sel), reverse=True):
            del self._cons_entradas[i]
        self._cons_atualizar_tree()

    def _cons_limpar(self):
        self._cons_entradas = []
        self._cons_atualizar_tree()

    def _cons_escolher_saida(self):
        p = filedialog.askdirectory(title="Pasta de saída da consolidação")
        if p:
            self.var_cons_saida.set(p)

    def _cons_abrir_pasta(self):
        pasta = self._cons_ultima_saida or self.var_cons_saida.get().strip()
        if not pasta:
            return
        try:
            os.startfile(pasta)                     # Windows
        except Exception:
            messagebox.showinfo("Pasta de saída", pasta)

    def _cons_salvar_log(self):
        p = filedialog.asksaveasfilename(defaultextension=".txt",
                                         filetypes=[("Texto","*.txt")])
        if p:
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(self.log_cons.get("1.0", "end"))
                messagebox.showinfo("Salvo", f"Log salvo em:\n{p}")
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    # ── execução ─────────────────────────────────────────────────────────────
    def _cons_iniciar(self):
        if self._cons_rodando:
            return
        if not self._cons_entradas:
            messagebox.showerror("Sem entradas",
                "Adicione ao menos uma pasta ou arquivo do DW/DA."); return

        saida = self.var_cons_saida.get().strip()
        if not saida:
            messagebox.showerror("Pasta de saída",
                "Escolha a pasta onde as planilhas serão gravadas."); return

        def _ano(var, nome):
            v = var.get().strip()
            if not v:
                return None, True
            if not (v.isdigit() and len(v) == 4):
                messagebox.showerror("Ano inválido",
                    f"{nome} deve ter 4 dígitos (ex.: 2021)."); return None, False
            return int(v), True

        ano_min, ok = _ano(self.var_cons_ano_min, "Ano mínimo")
        if not ok: return
        ano_max, ok = _ano(self.var_cons_ano_max, "Ano máximo")
        if not ok: return
        if ano_min and ano_max and ano_min > ano_max:
            messagebox.showerror("Filtro de ano",
                "O ano mínimo não pode ser maior que o ano máximo."); return

        # A saída dentro da entrada faz as planilhas geradas voltarem como
        # entrada numa segunda execução — serão ignoradas, mas custam leitura.
        saida_abs = os.path.abspath(saida)
        for e in self._cons_entradas:
            ent_abs = os.path.abspath(e["caminho"])
            if os.path.isdir(ent_abs) and \
               (saida_abs == ent_abs or saida_abs.startswith(ent_abs + os.sep)):
                if not messagebox.askyesno("Saída dentro da entrada",
                    "A pasta de saída está dentro de uma pasta de entrada.\n\n"
                    "Os arquivos gerados serão relidos (e ignorados) em uma "
                    "próxima execução.\n\nDeseja continuar mesmo assim?"):
                    return
                break

        # Tk não é thread-safe: tudo é lido aqui, na thread principal.
        params = dict(
            entradas=[e["caminho"] for e in self._cons_entradas if e["origem"] == "auto"],
            dw=[e["caminho"] for e in self._cons_entradas if e["origem"] == "dw"],
            da=[e["caminho"] for e in self._cons_entradas if e["origem"] == "da"],
            saida=saida,
            prefixo=self.var_cons_prefixo.get().strip() or "bps_dw_da__Classe_",
            sufixo=self.var_cons_sufixo.get().strip(),
            ano_min=ano_min, ano_max=ano_max,
            salvar_duplicatas=bool(self.var_cons_dup.get()),
            dedup_interno_da=bool(self.var_cons_dedup_interno.get()),
        )

        self._cons_rodando  = True
        self._cons_cancelar = False
        self._cons_ultima_saida = saida
        self.log_cons.configure(state="normal")
        self.log_cons.delete("1.0", "end")
        self.log_cons.configure(state="disabled")
        for k in ("c_dw", "c_lidas", "c_dup", "c_mant"):
            self._stat_cons(k, "0")
        self.progress_cons.set(0)
        self.lbl_cons_pct.configure(text="0%")
        self.lbl_cons_status.configure(text="Status: Consolidando…")
        self.btn_cons_start.configure(state="disabled")
        self.btn_cons_cancel.configure(state="normal", fg_color=C_RED,
                                       hover_color="#992B1E", text_color=C_SURFACE)
        self.btn_cons_abrir.configure(state="disabled")
        self.btn_cons_log.configure(state="disabled")

        self._log_cons(f"📁 Saída: {saida}", "info")
        if ano_min or ano_max:
            self._log_cons(f"📅 Filtro de ano: "
                           f"{ano_min or '—'} a {ano_max or '—'}", "info")
        if params["dedup_interno_da"]:
            self._log_cons("⚠ Dedup interno do DA ativo: repetições da mesma "
                           "chave dentro do próprio DA também serão removidas.",
                           "warn")
        self._log_cons("", "")

        threading.Thread(target=self._cons_thread, args=(params,),
                         daemon=True).start()

    def _cons_cancelar_click(self):
        if not self._cons_rodando:
            return
        self._cons_cancelar = True
        self.lbl_cons_status.configure(text="Status: Cancelando…")
        self._log_cons("\n🛑 Cancelamento solicitado — encerrando…", "warn")

    def _cons_thread(self, params):
        """Roda o motor fora da thread da UI; toda a volta é via self.after(0, …)."""
        try:
            res = consolidar(
                log=lambda m: self._ui(lambda: self._log_cons(m)),
                progresso=lambda f, r="": self._ui(
                    lambda: self._cons_progresso(f, r)),
                cancelado=lambda: self._cons_cancelar,
                **params)
            self._ui(lambda: self._cons_finalizar(res, None))
        except Exception as e:
            self._ui(lambda e=e: self._cons_finalizar(None, e))

    def _cons_progresso(self, fracao, rotulo=""):
        fracao = max(0.0, min(1.0, float(fracao)))
        self.progress_cons.set(fracao)
        self.lbl_cons_pct.configure(text=f"{int(fracao * 100)}%")
        if rotulo:
            self.lbl_cons_status.configure(text=f"Status: {rotulo}")

    def _cons_finalizar(self, res, erro):
        self._cons_rodando = False
        self.btn_cons_start.configure(state="normal")
        self.btn_cons_cancel.configure(state="disabled", fg_color="#E4E7EF",
                                       hover_color=C_BORDER, text_color=C_TEXT)
        self.btn_cons_log.configure(state="normal")

        if erro is not None:
            self.lbl_cons_status.configure(text="Status: Erro")
            self._log_cons(f"\n❌ Falha na consolidação: "
                           f"{type(erro).__name__}: {erro}", "err")
            messagebox.showerror("Erro na consolidação", str(erro))
            return

        if res.get("cancelado"):
            self.lbl_cons_status.configure(text="Status: Cancelado")
            self.progress_cons.set(0); self.lbl_cons_pct.configure(text="0%")
            return

        t = res.get("totais", {})
        self._stat_cons("c_dw",    self._fmt_br(t.get("dw", 0)))
        self._stat_cons("c_lidas", self._fmt_br(t.get("da_lidas", 0)))
        self._stat_cons("c_dup",   self._fmt_br(t.get("da_dup", 0)))
        self._stat_cons("c_mant",  self._fmt_br(t.get("da_mantidas", 0)))

        gerados = res.get("gerados", {})
        if not gerados:
            self.lbl_cons_status.configure(text="Status: Nada a consolidar")
            n_dw, n_da = res.get("arquivos_dw", 0), res.get("arquivos_da", 0)
            if not (n_dw or n_da):
                detalhe = ("Nenhum arquivo das entradas foi reconhecido como DW "
                           "ou DA.\nConfira os cabeçalhos ou classifique "
                           "manualmente em \"Origem\".")
            else:
                detalhe = (f"{n_dw} arquivo(s) do DW e {n_da} do DA foram lidos, "
                           "mas nenhuma linha válida sobrou.\nVeja o "
                           "linhas_em_quarentena.csv e o filtro de ano.")
            self._log_cons("\n⚠ " + detalhe.replace("\n", "\n   "), "warn")
            messagebox.showwarning("Nada a consolidar", detalhe)
            return

        self.progress_cons.set(1.0); self.lbl_cons_pct.configure(text="100%")
        self.lbl_cons_status.configure(text="Status: Concluído!")
        self.btn_cons_abrir.configure(state="normal")

        pct = (t.get("da_dup", 0) / t["da_lidas"] * 100) if t.get("da_lidas") else 0
        self._log_cons("\n" + "═" * 58, "info")
        self._log_cons(f"✅ {len(gerados)} planilha(s) gerada(s) em: "
                       f"{res.get('pasta_saida','')}", "ok")
        self._log_cons(f"   Relatório: {res.get('relatorio','')}", "ok")
        self._log_cons(f"   DA removido (já estava no DW): "
                       f"{self._fmt_br(t.get('da_dup', 0))}  ({pct:.2f}% do DA)", "ok")
        if res.get("arquivo_duplicatas"):
            self._log_cons(f"   Auditoria das duplicatas: "
                           f"{res['arquivo_duplicatas']}", "ok")
        if res.get("ignorados"):
            self._log_cons(f"⚠ {len(res['ignorados'])} arquivo(s) ignorado(s) "
                           f"(cabeçalho não reconhecido).", "warn")
        if res.get("quarentena"):
            self._log_cons(f"⚠ {self._fmt_br(res['quarentena'])} linha(s) "
                           f"corrompida(s) fora das planilhas — conteúdo "
                           f"preservado em {res.get('arquivo_quarentena','')}.",
                           "warn")
        if res.get("modalidades_desconhecidas"):
            self._log_cons("⚠ Códigos de modalidade não mapeados: "
                           + ", ".join(res["modalidades_desconhecidas"])
                           + "\n   Complete MAPA_MODALIDADE_DA em "
                             "consolidar_dw_da.py.", "warn")
        self._log_cons("═" * 58, "info")

        messagebox.showinfo("Concluído",
            f"Consolidação finalizada!\n\n"
            f"Planilhas geradas: {len(gerados)}\n"
            f"Linhas do DW: {self._fmt_br(t.get('dw', 0))}\n"
            f"Duplicatas removidas do DA: {self._fmt_br(t.get('da_dup', 0))}\n"
            f"Linhas do DA mantidas: {self._fmt_br(t.get('da_mantidas', 0))}\n\n"
            f"Pasta: {res.get('pasta_saida','')}")


# =============================================================================
# =============================================================================
#  MOTOR DA ABA 3 — CONSOLIDAÇÃO DW + DA (remoção de duplicatas)
#
#  Antes vivia em consolidar_dw_da.py; foi trazido para cá para o programa
#  ser um arquivo único. Nada abaixo depende da interface: consolidar() é
#  chamada tanto pela aba 3 quanto pela linha de comando.
# =============================================================================
# =============================================================================


# =============================================================================
# CONFIGURAÇÃO  (ajuste aqui se precisar)
# =============================================================================

# Limite físico de linhas de uma aba do Excel (inclui o cabeçalho).
# Ao estourar, o script cria automaticamente "dw-6505 (2)", "dw-6505 (3)"...
LIMITE_LINHAS_PLANILHA = 1_048_576

# Formatos aplicados na planilha final (iguais aos do modelo enviado).
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_QTDE = "#,##0"
FORMATO_MOEDA = 'R$ #,##0.00'

# True  -> CATMAT gravado como texto, preservando zeros à esquerda ("000183").
# False -> CATMAT gravado como número (183). Atenção: perde os zeros.
CATMAT_COMO_TEXTO = True

# Data do DA usada para preencher as colunas "Ano" e "dataCompra".
# Alterne para "dataResultado" se preferir alinhar com o "Ano Resultado Compra"
# do DW.
CAMPO_DATA_DA = "dataCompra"

# Formatar a data em dd/mm/aaaa custa ~35% de tempo a mais na gravação.
# Com False, a data sai como aaaa-mm-dd (mais rápido, ainda é data de verdade).
FORMATAR_DATA_BR = True

# Modalidades do DA (códigos do SIASG). Complete se aparecerem códigos novos —
# o script avisa no fim quais códigos não estavam mapeados.
MAPA_MODALIDADE_DA = {
    "1": "Convite",
    "2": "Tomada de Preços",
    "3": "Concorrência",
    "4": "Concorrência Internacional",
    "5": "Pregão",
    "6": "Dispensa de Licitação",
    "7": "Inexigibilidade de Licitação",
    "20": "Concurso",
    "22": "Regime Diferenciado de Contratações",
    "99": "Não informado",
}

# Sigla -> nome da unidade de fornecimento (mesmo mapa do Extrator de CATMATs).
_SIGLA_NOME = {
    "FR-AM": "Frasco-Ampola", "FR": "Frasco", "CAPS": "Cápsula",
    "COMPR": "Comprimido", "AM": "Ampola", "UN": "Unidade",
    "SER": "Seringa", "BIS": "Bisnaga", "BLIS": "Blister",
    "BOL": "Bolsa", "BOM": "Bombona", "CA": "Cartucho",
    "CI": "Curie", "CJ": "Conjunto", "DOSE(S)": "Dose(s)",
    "DOSES": "Dose(s)", "DRAG": "Drágea", "EMB": "Embalagem",
    "EMP": "Emplastro", "ENV": "Envelope", "FLAC": "Flaconete",
    "G": "Grama", "GL": "Galão", "GLOB": "Glóbulo",
    "KG": "Quilograma", "L": "Litro", "MCG": "Micrograma",
    "MCU": "Milicurie", "MG": "Miligrama",
    "MIL CTE": "Milheiro de Cartelas", "ML": "Mililitro",
    "PAST": "Pastilha", "POTE": "Pote", "RO": "Rolo",
    "SAC": "Sachê", "SUP": "Supositório", "TAB": "Tablete",
    "TBO": "Tubo", "TBTE": "Tubete", "UI": "Unid. Internacional",
}

_MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# Em algumas exportações do DW as colunas de código vêm com o cabeçalho EM BRANCO,
# logo depois da coluna descritiva correspondente (ex.: "Descrição Material
# Servico" seguida de uma coluna sem nome que contém o CATMAT). Este mapa
# reconstrói esses nomes.
_DW_CODIGO_SEGUINTE = {
    "descricao material servico": "catmat",
    "classe material": "classe",
    "padrao desc material": "inc",
    "orgao sup unid partic": "uasgsup",
    "orgao unid partic": "uasguni",
    "orgao vinc uresp compra": "uasgvinc",
    "nome uresp compra": "uasg",
}

# Tamanhos aceitos para as chaves (fora disso a linha vai para a quarentena).
TAM_CHAVE_DW = 22
TAM_ID_COMPRA_DA = (16, 17)
TAM_ITEM_DA = (1, 5)

# -----------------------------------------------------------------------------
# Layout das abas de saída: (título, tipo, largura)
# -----------------------------------------------------------------------------
COLS_DW = [
    ("IdentififItemCompra", "texto", 24),
    ("Catmat", "texto", 12),
    ("Descrição", "texto", 60),
    ("U.F.", "texto", 18),
    ("Classe", "inteiro", 9),
    ("Órgão Sup Unid Partic", "texto", 32),
    ("Órgão Unid Partic", "texto", 32),
    ("Órgão Vinc UResp Compra", "texto", 32),
    ("Nome UResp Compra", "texto", 36),
    ("Municipio UResp Compra", "texto", 22),
    ("UF UResp Compra", "texto", 8),
    ("Esfera Unid Partic", "texto", 14),
    ("CPF/CNPJ Fornecedor", "texto", 18),
    ("Nome Fornecedor", "texto", 40),
    ("Fabric Material Compra", "texto", 24),
    ("Marca Material Compra", "texto", 24),
    ("Ano Resultado Compra", "inteiro", 10),
    ("Dia Resultado Compra", "data", 14),
    ("Modalidade Compra", "texto", 26),
    ("Qtde Comprada Item", "qtde", 12),
    ("Valor Preço Unit Item", "moeda", 14),
    ("Preço Total", "moeda", 14),
]

COLS_DA = [
    ("IdentififItemCompra", "texto", 24),
    ("Catmat", "texto", 12),
    ("Descrição", "texto", 60),
    ("U.F.", "texto", 18),
    ("Classe", "inteiro", 9),
    ("nomeUasg", "texto", 36),
    ("municipio", "texto", 22),
    ("estado", "texto", 8),
    ("nomeOrgao", "texto", 32),
    ("poder", "texto", 8),
    ("esfera", "texto", 8),
    ("niFornecedor", "texto", 18),
    ("nomeFornecedor", "texto", 40),
    ("marca", "texto", 24),
    ("Ano", "inteiro", 8),
    ("dataCompra", "data", 14),
    ("modalidade", "texto", 26),
    ("quantidade", "qtde", 12),
    ("precoUnitario", "moeda", 14),
    ("Preço Total", "moeda", 14),
]

_FORMATO_POR_TIPO = {"data": FORMATO_DATA, "qtde": FORMATO_QTDE, "moeda": FORMATO_MOEDA}


# =============================================================================
# UTILIDADES
# =============================================================================

def norm(s) -> str:
    """Normaliza nome de coluna: sem acento, minúsculo, espaços colapsados."""
    if s is None:
        return ""
    s = str(s).replace("\ufeff", "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def txt(v) -> str:
    """Converte qualquer valor de célula em texto limpo, sem '.0' de float."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:                      # NaN
            return ""
        if v.is_integer():
            return str(int(v))
        return repr(v)
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else s


def num(v):
    """Converte texto em número, aceitando '1.234,56', '1234.56', '165,00', '1 '."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(" ", "").replace("\xa0", "")
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    neg = s.startswith("-")
    s = s.lstrip("+-").replace("R$", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")       # 1.234,56
    elif "," in s:
        s = s.replace(",", ".")                        # 1234,56
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3 and len(s.split(".")[0]) <= 3:
        s = s.replace(".", "")                         # 1.234 -> milhar
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def inteiro(v):
    f = num(v)
    return int(f) if f is not None else None


def data_dw(v):
    """'19 Out 1999' -> date(1999,10,19). Aceita também 19/10/1999 e ISO."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = txt(v)
    if not s:
        return None
    partes = s.replace("/", " ").replace("-", " ").split()
    if len(partes) == 3:
        d, m, a = partes
        mes = _MESES_PT.get(norm(m)[:3])
        if mes is None:
            try:
                mes = int(m)
            except ValueError:
                mes = None
        try:
            if mes and len(d) == 4:                    # formato ISO: aaaa mm dd
                return date(int(d), mes, int(a))
            if mes:
                return date(int(a), mes, int(d))
        except ValueError:
            return None
    return None


def data_iso(v):
    """'2025-07-04' ou '2025-07-04T12:00:00' -> date."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = txt(v)
    if len(s) >= 10:
        try:
            return date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except ValueError:
            pass
    return None


def so_digitos(s: str) -> str:
    return "".join(c for c in s if c.isdigit())


# =============================================================================
# LEITURA DE ARQUIVOS (CSV e XLSX) EM STREAMING
# =============================================================================

try:
    csv.field_size_limit(sys.maxsize)
except OverflowError:
    csv.field_size_limit(2 ** 31 - 1)


def _detectar_encoding(caminho: Path) -> str:
    with open(caminho, "rb") as fb:
        amostra = fb.read(1_048_576)
    if amostra.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    for corte in (0, 1, 2, 3):                 # evita erro por caractere partido
        try:
            (amostra[: len(amostra) - corte]).decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_separador(linha: str) -> str:
    # O DW exporta com ";" e o Extrator de CATMATs usa "@" nos CSVs do DA.
    candidatos = {sep: linha.count(sep) for sep in ("@", ";", "\t", "|", ",")}
    sep = max(candidatos, key=candidatos.get)
    return sep if candidatos[sep] > 0 else ";"


def ler_tabela(caminho: Path):
    """Gerador: entrega o cabeçalho (lista) e, em seguida, cada linha (lista)."""
    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb["Dados CATMAT"] if "Dados CATMAT" in wb.sheetnames else wb[wb.sheetnames[0]]
        try:
            for linha in ws.iter_rows(values_only=True):
                yield list(linha)
        finally:
            wb.close()
    else:
        enc = _detectar_encoding(caminho)
        with open(caminho, "r", encoding=enc, newline="", errors="replace") as f:
            primeira = f.readline()
            sep = _detectar_separador(primeira)
            f.seek(0)
            for linha in csv.reader(f, delimiter=sep):
                yield linha


def indices(cabecalho) -> dict:
    """{nome_normalizado: posição}, reconstruindo cabeçalhos vazios do DW."""
    nomes = [norm(c) for c in cabecalho]
    for i in range(1, len(nomes)):
        if not nomes[i]:
            nomes[i] = _DW_CODIGO_SEGUINTE.get(nomes[i - 1], "")
    idx = {}
    for i, nome in enumerate(nomes):
        if nome and nome not in idx:
            idx[nome] = i
    return idx


def detectar_fonte(idx: dict) -> Optional[str]:
    if "identif item compra" in idx:
        return "DW"
    if "idcompra" in idx and "numeroitemcompra" in idx:
        return "DA"
    return None


# =============================================================================
# TRANSFORMAÇÃO DAS LINHAS
# =============================================================================

def _leitor(linha, idx):
    def g(nome, padrao=""):
        i = idx.get(nome)
        if i is None or i >= len(linha):
            return padrao
        return txt(linha[i])
    return g


def _catmat(valor: str):
    if CATMAT_COMO_TEXTO or not valor:
        return valor
    return inteiro(valor) if so_digitos(valor) == valor else valor


def transformar_dw(linha, idx):
    """Devolve (chave22, classe, linha_de_saida, erro) a partir de uma linha do DW."""
    g = _leitor(linha, idx)

    chave = g("identif item compra")
    classe = g("classe")
    erro = ""
    if not (chave.isdigit() and len(chave) == TAM_CHAVE_DW):
        erro = f"'Identif Item Compra' deveria ter {TAM_CHAVE_DW} dígitos: {chave!r}"
    elif classe and not classe.isdigit():
        erro = f"coluna 'Classe' não numérica ({classe!r}) - provável desalinhamento"

    qtde = num(g("qtde comprada item"))
    preco = num(g("valor preco unit item"))
    total = round(qtde * preco, 2) if (qtde is not None and preco is not None) else None
    saida = [
        chave,
        _catmat(g("catmat")),
        g("descricao material servico"),
        g("unidade fornecimento"),
        inteiro(classe),
        g("orgao sup unid partic"),
        g("orgao unid partic"),
        g("orgao vinc uresp compra"),
        g("nome uresp compra"),
        g("municipio uresp compra"),
        g("uf uresp compra"),
        g("esfera unid partic"),
        g("cpf/cnpj fornecedor"),
        g("nome fornecedor"),
        g("fabric material compra"),
        g("marca material compra"),
        inteiro(g("ano resultado compra")),
        data_dw(g("dia resultado compra")),
        g("modalidade compra"),
        qtde,
        preco,
        total,
    ]
    return chave, classe, saida, erro


def _unidade_fornecimento_da(g):
    """Reproduz a coluna 'Unidade de Fornecimento' do Extrator de CATMATs."""
    pronta = g("unidade de fornecimento")
    if pronta:
        return pronta
    nome = g("nomeunidadefornecimento")
    sigla = g("siglaunidadefornecimento")
    cap = g("capacidadeunidadefornecimento")
    medida = g("siglaunidademedida")
    if not nome and sigla:
        nome = _SIGLA_NOME.get(sigla.upper(), sigla)
    cap_num = num(cap) or 0
    partes = [nome] if nome else []
    if cap and cap_num != 0:
        partes.append(cap)
        if medida:
            partes.append(medida)
    return " ".join(partes)


def transformar_da(linha, idx, modalidades_desconhecidas):
    """Devolve (chave22, classe, linha_de_saida, erro) a partir de uma linha do DA."""
    g = _leitor(linha, idx)

    id_compra = g("idcompra")
    n_item = g("numeroitemcompra")
    classe = g("codigoclasse")
    chave, erro = "", ""

    # Validação estrita: é ela que separa a linha boa da linha corrompida.
    # Linhas desalinhadas costumam trazer o idItemCompra no lugar do idCompra,
    # o que geraria uma chave errada (e uma duplicata não detectada).
    if not (id_compra.isdigit() and TAM_ID_COMPRA_DA[0] <= len(id_compra) <= TAM_ID_COMPRA_DA[1]):
        erro = (f"'idCompra' deveria ter {TAM_ID_COMPRA_DA[0]} ou "
                f"{TAM_ID_COMPRA_DA[1]} dígitos: {id_compra!r}")
    elif not (n_item.isdigit() and TAM_ITEM_DA[0] <= len(n_item) <= TAM_ITEM_DA[1]):
        erro = f"'numeroItemCompra' inválido: {n_item!r}"
    elif classe and not classe.isdigit():
        erro = f"'codigoClasse' não numérico ({classe!r}) - provável desalinhamento"
    else:
        chave = id_compra.zfill(17) + n_item.zfill(5)

    dt = data_iso(g(norm(CAMPO_DATA_DA))) or data_iso(g("datacompra")) or data_iso(g("dataresultado"))
    qtde = num(g("quantidade"))
    preco = num(g("precounitario"))
    total = round(qtde * preco, 2) if (qtde is not None and preco is not None) else None

    cod_mod = g("modalidade")
    modalidade = MAPA_MODALIDADE_DA.get(cod_mod, "")
    if not modalidade and cod_mod:
        if not erro:                       # linha corrompida não conta como código novo
            modalidades_desconhecidas.add(cod_mod)
        modalidade = f"Modalidade {cod_mod}"

    saida = [
        chave,
        _catmat(g("codigoitemcatalogo")),
        g("descricaoitem"),
        _unidade_fornecimento_da(g),
        inteiro(classe),
        g("nomeuasg"),
        g("municipio"),
        g("estado"),
        g("nomeorgao"),
        g("poder"),
        g("esfera"),
        g("nifornecedor"),
        g("nomefornecedor"),
        g("marca"),
        dt.year if dt else None,
        dt,
        modalidade,
        qtde,
        preco,
        total,
    ]
    return chave, classe, saida, erro


# =============================================================================
# ESCRITA DAS PLANILHAS
# =============================================================================

_FONTE_CAB = Font(bold=True, color="FFFFFF")
_FUNDO_CAB = PatternFill("solid", fgColor="1F4E79")
_ALINHA_CAB = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _letra(i: int) -> str:
    letra = ""
    while i >= 0:
        letra = chr(ord("A") + i % 26) + letra
        i = i // 26 - 1
    return letra


class SaidaClasse:
    """Um arquivo .xlsx por classe, com as abas dw-XXXX e da-XXXX."""

    def __init__(self, classe: str):
        self.classe = classe
        self.wb = Workbook(write_only=True)
        self.abas = {"dw": [], "da": []}
        self._nova_aba("dw")                      # garante a aba do DW mesmo vazia

    def _nova_aba(self, tipo: str):
        cols = COLS_DW if tipo == "dw" else COLS_DA
        n = len(self.abas[tipo]) + 1
        nome = f"{tipo}-{self.classe}" + (f" ({n})" if n > 1 else "")
        ws = self.wb.create_sheet(nome[:31])
        for i, (_titulo, tipo_col, largura) in enumerate(cols):
            dim = ws.column_dimensions[_letra(i)]
            dim.width = largura
            if tipo_col in _FORMATO_POR_TIPO and tipo_col != "data":
                dim.number_format = _FORMATO_POR_TIPO[tipo_col]
        cabecalho = []
        for titulo, _t, _l in cols:
            c = WriteOnlyCell(ws, value=titulo)
            c.font, c.fill, c.alignment = _FONTE_CAB, _FUNDO_CAB, _ALINHA_CAB
            cabecalho.append(c)
        # Em modo write_only, tudo o que vem antes de <sheetData> (painéis,
        # larguras, formatos de coluna) precisa ser definido ANTES do primeiro
        # append; o atributo simples ws.freeze_panes não é serializado.
        Worksheet.freeze_panes.fset(ws, "A2")
        ws.append(cabecalho)
        registro = [ws, 1]
        self.abas[tipo].append(registro)
        return registro

    def append(self, tipo: str, valores: list):
        registro = self.abas[tipo][-1] if self.abas[tipo] else self._nova_aba(tipo)
        if registro[1] >= LIMITE_LINHAS_PLANILHA:
            registro = self._nova_aba(tipo)
        ws = registro[0]
        if FORMATAR_DATA_BR:
            cols = COLS_DW if tipo == "dw" else COLS_DA
            valores = list(valores)
            for i, (_t, tipo_col, _l) in enumerate(cols):
                if tipo_col == "data" and valores[i] is not None:
                    c = WriteOnlyCell(ws, value=valores[i])
                    c.number_format = FORMATO_DATA
                    valores[i] = c
        ws.append(valores)
        registro[1] += 1

    def descartar(self):
        """Abandona a planilha sem gravar (cancelamento).

        Em modo write_only cada aba escreve num arquivo temporário através de
        um gerador lxml. Se o objeto simplesmente for coletado pelo garbage
        collector, o gerador é fechado no meio de um elemento XML e o Python
        imprime um "Exception ignored ... LxmlSyntaxError". Fechar na ordem
        certa (aba -> tempfile) evita o ruído e ainda apaga os temporários.
        """
        for registros in self.abas.values():
            for registro in registros:
                ws = registro[0]
                try:
                    ws.close()
                except Exception:
                    pass
                escritor = getattr(ws, "_writer", None)
                if escritor is not None:
                    try:
                        escritor.cleanup()
                    except Exception:
                        pass
        self.abas = {"dw": [], "da": []}

    def salvar(self, caminho: Path):
        for tipo, cols in (("dw", COLS_DW), ("da", COLS_DA)):
            if not self.abas[tipo]:
                self._nova_aba(tipo)
            ultima = _letra(len(cols) - 1)
            for ws, linhas in self.abas[tipo]:
                ws.auto_filter.ref = f"A1:{ultima}{max(linhas, 1)}"
        self.wb.save(caminho)


# =============================================================================
# PROCESSAMENTO
# =============================================================================

class Cancelado(Exception):
    """Interrompe a consolidação a pedido de quem chamou (botão Cancelar da GUI)."""


# Só é consultado de tempos em tempos: a checagem por linha custaria caro em
# arquivos de milhões de registros.
_INTERVALO_CANCELAMENTO = 2000


def coletar_arquivos(entradas, padroes=("*.csv", "*.CSV", "*.xlsx", "*.xlsm")) -> list:
    achados = []
    for entrada in entradas:
        p = Path(entrada)
        if p.is_dir():
            for padrao in padroes:
                achados.extend(sorted(p.rglob(padrao)))
        elif p.exists():
            achados.append(p)
        else:
            achados.extend(sorted(Path(x) for x in glob.glob(entrada)))
    vistos, unicos = set(), []
    for a in achados:
        chave = str(a.resolve()).lower()
        if chave not in vistos and a.suffix.lower() in (".csv", ".xlsx", ".xlsm"):
            vistos.add(chave)
            unicos.append(a)
    return unicos


def _est(estatisticas, classe):
    return estatisticas.setdefault(classe, {
        "dw": 0, "da_lidas": 0, "da_dup": 0, "da_mantidas": 0,
        "dw_invalidas": 0, "da_invalidas": 0, "da_dup_interna": 0,
    })


class Quarentena:
    """Guarda as linhas corrompidas em CSV, sem perder nada do conteúdo original."""

    def __init__(self, caminho: Path):
        self.caminho = caminho
        self._arq = None
        self._csv = None
        self.total = 0

    def registrar(self, fonte, arquivo, n_linha, motivo, linha):
        if self._arq is None:
            self._arq = open(self.caminho, "w", encoding="utf-8-sig", newline="")
            self._csv = csv.writer(self._arq, delimiter=";")
            self._csv.writerow(["Fonte", "Arquivo", "Linha", "Motivo", "Conteúdo original ->"])
        self._csv.writerow([fonte, arquivo, n_linha, motivo] + [txt(v) for v in linha])
        self.total += 1

    def fechar(self):
        if self._arq:
            self._arq.close()


def processar_dw(arquivos, saidas, estatisticas, chaves_dw, ano_min, ano_max,
                 quarentena, verbose=True, log=print, cancelado=None,
                 progresso=None):
    for i_arq, arq in enumerate(arquivos, start=1):
        if progresso:
            progresso(i_arq - 1, len(arquivos), f"DW · {arq.name}")
        if cancelado and cancelado():
            raise Cancelado()
        gen = ler_tabela(arq)
        try:
            idx = indices(next(gen))
        except StopIteration:
            continue
        if "classe" not in idx:
            log(f"  ! {arq.name}: coluna 'Classe' não encontrada - as linhas irão "
                f"para SEM_CLASSE")
        gravadas = ignoradas = invalidas = 0
        for n_linha, linha in enumerate(gen, start=2):
            if cancelado and n_linha % _INTERVALO_CANCELAMENTO == 0 and cancelado():
                gen.close()
                raise Cancelado()
            if not linha or all(v in (None, "") for v in linha):
                continue
            chave, classe, saida, erro = transformar_dw(linha, idx)
            if erro:
                invalidas += 1
                _est(estatisticas, classe if classe.isdigit() else "SEM_CLASSE")["dw_invalidas"] += 1
                quarentena.registrar("DW", arq.name, n_linha, erro, linha)
                continue
            classe = classe or "SEM_CLASSE"
            chaves_dw.add(int(chave))              # int ocupa menos memória que str
            ano = saida[16]
            if (ano_min and ano and ano < ano_min) or (ano_max and ano and ano > ano_max):
                ignoradas += 1
                continue
            if classe not in saidas:
                saidas[classe] = SaidaClasse(classe)
            saidas[classe].append("dw", saida)
            _est(estatisticas, classe)["dw"] += 1
            gravadas += 1
        if verbose:
            extra = f" | {ignoradas} fora do filtro de ano" if ignoradas else ""
            extra += f" | {invalidas} em quarentena" if invalidas else ""
            log(f"  [DW] {arq.name}: {gravadas:,} linhas{extra}".replace(",", "."))
    if progresso:
        progresso(len(arquivos), len(arquivos), "DW concluído")


def processar_da(arquivos, saidas, estatisticas, chaves_dw, ano_min, ano_max,
                 dedup_interno, escritor_dup, modalidades_desconhecidas,
                 quarentena, verbose=True, log=print, cancelado=None,
                 progresso=None):
    vistas_da = set() if dedup_interno else None
    for i_arq, arq in enumerate(arquivos, start=1):
        if progresso:
            progresso(i_arq - 1, len(arquivos), f"DA · {arq.name}")
        if cancelado and cancelado():
            raise Cancelado()
        gen = ler_tabela(arq)
        try:
            idx = indices(next(gen))
        except StopIteration:
            continue
        lidas = dup = mantidas = ignoradas = invalidas = 0
        for n_linha, linha in enumerate(gen, start=2):
            if cancelado and n_linha % _INTERVALO_CANCELAMENTO == 0 and cancelado():
                gen.close()
                raise Cancelado()
            if not linha or all(v in (None, "") for v in linha):
                continue
            primeira = txt(linha[0]).lower()
            if primeira.startswith(("totalregistros", "totalpaginas", "idcompra")):
                continue                            # rodapé da API ou cabeçalho repetido
            lidas += 1
            chave, classe, saida, erro = transformar_da(linha, idx, modalidades_desconhecidas)
            if erro:
                invalidas += 1
                _est(estatisticas, classe if classe.isdigit() else "SEM_CLASSE")["da_invalidas"] += 1
                quarentena.registrar("DA", arq.name, n_linha, erro, linha)
                continue
            classe = classe or "SEM_CLASSE"
            e = _est(estatisticas, classe)
            e["da_lidas"] += 1

            if int(chave) in chaves_dw:
                dup += 1
                e["da_dup"] += 1
                if escritor_dup:
                    escritor_dup.writerow([classe, chave, arq.name, saida[1],
                                           saida[15] or "", saida[12], saida[18]])
                continue
            if vistas_da is not None:
                if int(chave) in vistas_da:
                    e["da_dup_interna"] += 1
                    continue
                vistas_da.add(int(chave))

            ano = saida[14]
            if (ano_min and ano and ano < ano_min) or (ano_max and ano and ano > ano_max):
                ignoradas += 1
                continue
            if classe not in saidas:
                saidas[classe] = SaidaClasse(classe)
            saidas[classe].append("da", saida)
            e["da_mantidas"] += 1
            mantidas += 1
        if verbose:
            extra = f" | {ignoradas} fora do filtro de ano" if ignoradas else ""
            extra += f" | {invalidas} em quarentena" if invalidas else ""
            log(f"  [DA] {arq.name}: {lidas:,} lidas | {dup:,} duplicadas removidas | "
                f"{mantidas:,} mantidas{extra}".replace(",", "."))
    if progresso:
        progresso(len(arquivos), len(arquivos), "DA concluído")


def gravar_relatorio(caminho: Path, estatisticas: dict, arquivos_gerados: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    cabecalho = ["Classe", "Linhas DW", "DA lidas", "DA duplicadas (removidas)",
                 "DA mantidas", "% removido do DA", "DW em quarentena",
                 "DA em quarentena", "Duplicatas internas DA", "Arquivo gerado"]
    ws.append(cabecalho)
    for c in ws[1]:
        c.font, c.fill, c.alignment = _FONTE_CAB, _FUNDO_CAB, _ALINHA_CAB
    for classe in sorted(estatisticas):
        e = estatisticas[classe]
        pct = (e["da_dup"] / e["da_lidas"] * 100) if e["da_lidas"] else 0
        ws.append([classe, e["dw"], e["da_lidas"], e["da_dup"], e["da_mantidas"],
                   round(pct, 2), e["dw_invalidas"], e["da_invalidas"],
                   e["da_dup_interna"], arquivos_gerados.get(classe, "")])
    larguras = [14, 14, 14, 26, 14, 18, 18, 18, 22, 44]
    for i, w in enumerate(larguras):
        ws.column_dimensions[_letra(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_letra(len(cabecalho) - 1)}{ws.max_row}"
    wb.save(caminho)


# =============================================================================
# MOTOR REUTILIZÁVEL
# =============================================================================

def consolidar(entradas=(), dw=(), da=(), saida=".",
               prefixo="bps_dw_da__Classe_", sufixo="",
               ano_min=None, ano_max=None,
               salvar_duplicatas=False, dedup_interno_da=False,
               log=print, progresso=None, cancelado=None) -> dict:
    """Consolida DW + DA e devolve um resumo do que foi feito.

    É o mesmo motor usado pela linha de comando; os três callbacks existem para
    a interface gráfica:
        log(msg)                       -> uma linha de texto para o usuário
        progresso(fracao, rotulo)      -> fracao entre 0.0 e 1.0
        cancelado() -> bool            -> True interrompe (levanta Cancelado)
    """
    dir_saida = Path(saida)
    dir_saida.mkdir(parents=True, exist_ok=True)

    arquivos_dw = coletar_arquivos(dw)
    arquivos_da = coletar_arquivos(da)
    indefinidos = [a for a in coletar_arquivos(entradas)
                   if a not in arquivos_dw and a not in arquivos_da]

    resultado = {
        "cancelado": False, "arquivos_dw": 0, "arquivos_da": 0, "ignorados": [],
        "estatisticas": {}, "gerados": {}, "pasta_saida": str(dir_saida),
        "totais": {"dw": 0, "da_lidas": 0, "da_dup": 0, "da_mantidas": 0},
        "quarentena": 0, "arquivo_quarentena": None, "arquivo_duplicatas": None,
        "modalidades_desconhecidas": [], "relatorio": None,
    }

    if not (arquivos_dw or arquivos_da or indefinidos):
        log("Nenhum arquivo .csv/.xlsx encontrado nas entradas informadas.")
        return resultado

    # ── Fatias da barra de progresso: ler DW, ler DA, gravar planilhas ────────
    def _fracao(base, peso):
        def _cb(feito, total, rotulo=""):
            if progresso:
                progresso(base + peso * (feito / total if total else 1), rotulo)
        return _cb

    saidas, estatisticas, chaves_dw = {}, {}, set()
    modalidades_desconhecidas = set()
    quarentena = Quarentena(dir_saida / "linhas_em_quarentena.csv")
    arq_dup = escritor_dup = None

    try:
        log("Identificando a origem de cada arquivo...")
        ignorados = []
        for arq in indefinidos:
            if cancelado and cancelado():
                raise Cancelado()
            gen = ler_tabela(arq)
            try:
                fonte = detectar_fonte(indices(next(gen)))
            except StopIteration:
                fonte = None
            gen.close()
            if fonte == "DW":
                arquivos_dw.append(arq)
            elif fonte == "DA":
                arquivos_da.append(arq)
            else:
                ignorados.append(arq)
        for arq in ignorados:
            log(f"  ! ignorado (cabeçalho não reconhecido): {arq.name}")
        log(f"  {len(arquivos_dw)} arquivo(s) do DW | "
            f"{len(arquivos_da)} arquivo(s) do DA")
        resultado["arquivos_dw"] = len(arquivos_dw)
        resultado["arquivos_da"] = len(arquivos_da)
        resultado["ignorados"] = [a.name for a in ignorados]

        log("\n[1/3] Lendo o DW e montando o índice de chaves...")
        processar_dw(arquivos_dw, saidas, estatisticas, chaves_dw, ano_min,
                     ano_max, quarentena, log=log, cancelado=cancelado,
                     progresso=_fracao(0.0, 0.45))
        log(f"  -> {len(chaves_dw):,} chaves únicas no DW".replace(",", "."))

        if salvar_duplicatas:
            arq_dup = open(dir_saida / "duplicatas_removidas.csv", "w",
                           encoding="utf-8-sig", newline="")
            escritor_dup = csv.writer(arq_dup, delimiter=";")
            escritor_dup.writerow(["Classe", "IdentififItemCompra", "Arquivo origem",
                                   "Catmat", "dataCompra", "nomeFornecedor",
                                   "precoUnitario"])

        log("\n[2/3] Lendo o DA e removendo as duplicatas...")
        processar_da(arquivos_da, saidas, estatisticas, chaves_dw, ano_min,
                     ano_max, dedup_interno_da, escritor_dup,
                     modalidades_desconhecidas, quarentena, log=log,
                     cancelado=cancelado, progresso=_fracao(0.45, 0.40))
    except Cancelado:
        resultado["cancelado"] = True
        for saida_classe in saidas.values():
            saida_classe.descartar()
        saidas.clear()
        log("\n🛑 Consolidação cancelada — nenhuma planilha foi gravada.")
        return resultado
    finally:
        if arq_dup:
            arq_dup.close()
        quarentena.fechar()

    log("\n[3/3] Gravando as planilhas...")
    gerados = {}
    classes = sorted(saidas)
    for i, classe in enumerate(classes, start=1):
        nome = f"{prefixo}{classe}{sufixo}.xlsx"
        saidas[classe].salvar(dir_saida / nome)
        gerados[classe] = nome
        e = _est(estatisticas, classe)
        log(f"  {nome}: dw={e['dw']:,} | da={e['da_mantidas']:,} "
            f"(removidas {e['da_dup']:,})".replace(",", "."))
        if progresso:
            progresso(0.85 + 0.15 * (i / len(classes)), f"Gravando {nome}")

    nome_relatorio = "Relatorio_Consolidacao.xlsx"
    gravar_relatorio(dir_saida / nome_relatorio, estatisticas, gerados)

    resultado.update({
        "estatisticas": estatisticas,
        "gerados": gerados,
        "relatorio": nome_relatorio,
        "totais": {
            "dw":          sum(e["dw"] for e in estatisticas.values()),
            "da_lidas":    sum(e["da_lidas"] for e in estatisticas.values()),
            "da_dup":      sum(e["da_dup"] for e in estatisticas.values()),
            "da_mantidas": sum(e["da_mantidas"] for e in estatisticas.values()),
        },
        "quarentena": quarentena.total,
        "arquivo_quarentena": quarentena.caminho.name if quarentena.total else None,
        "arquivo_duplicatas": "duplicatas_removidas.csv" if salvar_duplicatas else None,
        "modalidades_desconhecidas": sorted(modalidades_desconhecidas),
    })
    if progresso:
        progresso(1.0, "Concluído")
    return resultado


# =============================================================================
# MAIN
# =============================================================================

def _cli_consolidacao():
    ap = argparse.ArgumentParser(
        description="Consolida DW (SIASG) + DA (Compras.gov) por classe, "
                    "removendo do DA os registros já existentes no DW.")
    ap.add_argument("-e", "--entrada", action="append", default=[],
                    help="pasta (ou arquivo/curinga) com os CSVs do DW e do DA. "
                         "Pode repetir. A origem é detectada pelo cabeçalho.")
    ap.add_argument("--dw", action="append", default=[], help="entradas que são do DW")
    ap.add_argument("--da", action="append", default=[], help="entradas que são do DA")
    ap.add_argument("-s", "--saida", required=True, help="pasta de saída")
    ap.add_argument("--prefixo", default="bps_dw_da__Classe_",
                    help="prefixo do nome dos arquivos gerados")
    ap.add_argument("--sufixo", default="", help="sufixo do nome (ex.: _2020_a_2026)")
    ap.add_argument("--ano-min", type=int, help="descarta registros anteriores a este ano")
    ap.add_argument("--ano-max", type=int, help="descarta registros posteriores a este ano")
    ap.add_argument("--salvar-duplicatas", action="store_true",
                    help="gera CSV de auditoria com as linhas do DA removidas")
    ap.add_argument("--dedup-interno-da", action="store_true",
                    help="também remove repetições da mesma chave dentro do próprio DA "
                         "(atenção: costumam ser registros distintos, de fornecedores "
                         "e preços diferentes)")
    args = ap.parse_args()

    r = consolidar(entradas=args.entrada, dw=args.dw, da=args.da,
                   saida=args.saida, prefixo=args.prefixo, sufixo=args.sufixo,
                   ano_min=args.ano_min, ano_max=args.ano_max,
                   salvar_duplicatas=args.salvar_duplicatas,
                   dedup_interno_da=args.dedup_interno_da)

    if not r["gerados"] and not r["estatisticas"]:
        return 1

    t = r["totais"]
    print("\n" + "=" * 62)
    print(f"DW gravado ..................... {t['dw']:,}".replace(",", "."))
    print(f"DA lido (linhas válidas) ....... {t['da_lidas']:,}".replace(",", "."))
    print(f"DA removido (já estava no DW) .. {t['da_dup']:,}".replace(",", "."))
    print(f"DA mantido ..................... {t['da_mantidas']:,}".replace(",", "."))
    print(f"Relatório ...................... {r['relatorio']}")
    if r["quarentena"]:
        print(f"\n! {r['quarentena']:,} linha(s) corrompida(s) não entraram nas planilhas."
              .replace(",", "."))
        print(f"  Conteúdo preservado em: {r['arquivo_quarentena']}")
    if r["modalidades_desconhecidas"]:
        print(f"\n! Códigos de modalidade não mapeados: "
              f"{', '.join(r['modalidades_desconhecidas'])}"
              f"\n  Complete o dicionário MAPA_MODALIDADE_DA neste arquivo.")
    print("=" * 62)
    return 0


# =============================================================================
if __name__ == "__main__":
    # Com argumentos, roda a consolidação em linha de comando; sem eles,
    # abre a interface. Ex.: python ExtratorCatmat.py -e ENTRADA -s SAIDA
    if len(sys.argv) > 1:
        sys.exit(_cli_consolidacao())
    app = App()
    app.mainloop()
